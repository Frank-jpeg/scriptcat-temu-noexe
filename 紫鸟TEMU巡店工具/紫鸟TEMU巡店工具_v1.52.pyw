from __future__ import annotations

import json
import os
import re
import shutil
import threading
import time
import urllib.error
import urllib.request
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, ttk

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

DESKTOP_DIR = Path.home() / "Desktop"
CONFIG_PATH = Path.home() / ".zclaw" / "config.json"
BACKUP_DIR = Path(r"D:\临时备份")
BASE_URL_DEFAULT = "http://127.0.0.1:9481"
RECENT_DAYS_DEFAULT = 2
ARRIVAL_OVERDUE_RECENT_DAYS = 4
TEMU_HOME_URL = "https://agentseller.temu.com/"
QC_DETAIL_URL = "https://seller.kuajingmaihuo.com/wms/qc-detail"
URGENT_STOCK_URL = "https://agentseller.temu.com/stock/fully-mgt/order-manage-urgency"
GOVERN_DASHBOARD_URL = "https://agentseller.temu.com/govern/dashboard"
SHIPPING_LIST_URL = "https://seller.kuajingmaihuo.com/main/order-manager/shipping-list"
RETURN_ORDER_URL = "https://seller.kuajingmaihuo.com/wms/stock-mgt/return-order-mgt"
FUNDS_CENTER_URL = "https://seller.kuajingmaihuo.com/labor/account"
VIOLATION_MESSAGE_URL = "https://agentseller.temu.com/wms/stock-mgt/violation-message"
PRICE_RULE_URL = "https://agentseller.temu.com/main/adjust-price-manage/order-price"
POST_CHECK_DOCK_URL = "https://agentseller.temu.com/newon/product-select"
POST_CHECK_DOCK_LABEL = "上新生命周期管理"
PRICE_RULE_CONFIG_PATH = Path.home() / ".zclaw" / "temu_price_rule_config.json"
STORE_EMAIL_CONFIG_PATH = Path.home() / ".zclaw" / "temu_store_email_accounts.json"
SHIPPING_STALE_DAYS = 6
WITHDRAW_ALERT_THRESHOLD = 2000.0
PRICE_RULE_PAGE_SIZE = 100
NAVIGATION_RETRY_COUNT = 3
NAVIGATION_RETRY_DELAY = 3.0
PROJECT_RETRY_COUNT = 2
EMAIL_LOGIN_RETRY_COUNT = 3

TITLE_FILL = PatternFill("solid", fgColor="EAF3FF")
SECTION_FILL = PatternFill("solid", fgColor="F4F6F8")
CARD_FILL = PatternFill("solid", fgColor="FFF7E8")
HEADER_FILL = PatternFill("solid", fgColor="F2F2F2")
ALERT_ROW_FILL = PatternFill("solid", fgColor="FFD9D9")
ALERT_STRONG_FILL = PatternFill("solid", fgColor="C62828")
OK_FILL = PatternFill("solid", fgColor="EAF7EA")
THIN = Side(style="thin", color="D9D9D9")
BOX_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

TOOLS_REQUIRED = {
    "resolve_store",
    "list_stores",
    "open_store",
    "close_store",
    "visit_page",
    "execute_script",
    "click_element",
}

INSPECTION_ITEMS = [
    ("qc", "抽检结果明细"),
    ("urgent", "检查JIT是否逾期"),
    ("urgent_declared_price", "检查待发货低申报价"),
    ("govern", "合规中心"),
    ("shipping", f"检查VMI超{SHIPPING_STALE_DAYS}天未收货"),
    ("violation", "检查违规信息待处理"),
    ("price_rule", "价格申报自动助手"),
    ("return_order", "退货包裹查询"),
    ("funds", "检查资金中心余额"),
]
LOW_PRIORITY_INSPECTION_KEYS = {"return_order", "funds"}
HIGH_PRIORITY_INSPECTION_KEYS = [key for key, _label in INSPECTION_ITEMS if key not in LOW_PRIORITY_INSPECTION_KEYS]
ALL_INSPECTION_KEYS = {key for key, _label in INSPECTION_ITEMS}
INSPECTION_LABELS = dict(INSPECTION_ITEMS)

DEFAULT_PRICE_RULE_ROWS = [
    {"kw": "帽", "min": 0, "max": 15.5, "action": "不调整"},
    {"kw": "帽", "min": 15.6, "max": 99, "action": "调整"},
    {"kw": "拉链", "min": 0, "max": 17.5, "action": "不调整"},
    {"kw": "拉链", "min": 17.6, "max": 99, "action": "调整"},
    {"kw": "钱包", "min": 0, "max": 12, "action": "不调整"},
    {"kw": "钱包", "min": 12.1, "max": 99, "action": "调整"},
    {"kw": "背包", "min": 0, "max": 17.5, "action": "不调整"},
    {"kw": "背包", "min": 17.56, "max": 99, "action": "调整"},
    {"kw": "健身", "min": 0, "max": 17.5, "action": "不调整"},
    {"kw": "健身", "min": 17.56, "max": 99, "action": "调整"},
]
LEGACY_DEFAULT_PRICE_RULE_ROWS = [
    {"kw": "帽", "min": 0.0, "max": 15.5, "action": "不调整"},
    {"kw": "帽", "min": 15.6, "max": 99.0, "action": "调整"},
    {"kw": "拉链", "min": 0.0, "max": 17.5, "action": "不调整"},
    {"kw": "拉链", "min": 17.6, "max": 99.0, "action": "调整"},
    {"kw": "钱包", "min": 0.0, "max": 12.0, "action": "不调整"},
    {"kw": "钱包", "min": 12.1, "max": 99.0, "action": "调整"},
]
DEFAULT_PRICE_RULE_CONFIG = {
    "protectDiff": True,
    "protectDiffLimit": 1.0,
    "rules": DEFAULT_PRICE_RULE_ROWS,
}
TEMU_STATE_SCRIPT = r"""
(() => {
  const normalize = (value) => (value || '').replace(/\s+/g, ' ').trim();
  const visible = (el) => {
    const rect = el.getBoundingClientRect();
    const style = getComputedStyle(el);
    return rect.width > 0 && rect.height > 0 && style.visibility !== 'hidden' && style.display !== 'none';
  };
  const title = normalize(document.title || '');
  const body = normalize(document.body.innerText || '');
  const url = location.href || '';
  const hasHomeEntry = body.includes('首页');
  const portalMarkers = [
    'TEMU Agent Center',
    'Seller Central',
    '履约中心',
    '备货管理',
    '服务市场',
    '运营对接',
    '规则中心',
    '学习',
  ];
  const portalReady = portalMarkers.some((marker) => title.includes(marker) || body.includes(marker));
  const regionButtons = Array.from(document.querySelectorAll('a,button,div,span'))
    .filter((el) => visible(el) && normalize(el.innerText) === '商家中心')
    .map((el) => {
      const rect = el.getBoundingClientRect();
      return {
        tag: el.tagName,
        className: el.className || '',
        x: rect.x,
        y: rect.y,
        width: rect.width,
        height: rect.height
      };
    });
  const authButton = Array.from(document.querySelectorAll('button'))
    .find((el) => visible(el) && normalize(el.innerText).includes('授权登录'));
  const confirmForwardButton = Array.from(document.querySelectorAll('button'))
    .find((el) => visible(el) && normalize(el.innerText).includes('确认授权并前往'));
  const checkbox = document.querySelector('input[type="checkbox"], input[value="on"]');
  const authInputs = Array.from(document.querySelectorAll('input'))
    .filter((el) => {
      const type = String(el.type || '').toLowerCase();
      return visible(el) && ['text', 'email', 'tel', 'password'].includes(type);
    });
  const authFilledCount = authInputs.filter((el) => normalize(el.value || '') !== '').length;
  const authPasswordFilledCount = authInputs.filter((el) => String(el.type || '').toLowerCase() === 'password' && normalize(el.value || '') !== '').length;
  const authAccountFilledCount = authInputs.filter((el) => ['text', 'email', 'tel'].includes(String(el.type || '').toLowerCase()) && normalize(el.value || '') !== '').length;
  const shortPhoneInput = authInputs
    .filter((el) => String(el.type || '').toLowerCase() !== 'password')
    .map((el) => normalize(el.value || ''))
    .filter((value) => value !== '+86')
    .map((value) => value.replace(/\D/g, ''))
    .filter((digits) => digits !== '86')
    .find((digits) => digits.length > 0 && digits.length < 11) || '';
  const bodyPhoneMatch = body.match(/\+86\s*(\d{1,10})(?!\d)/);
  const shortPhoneDigits = shortPhoneInput || (bodyPhoneMatch ? bodyPhoneMatch[1] : '');
  const hasPhonePrefix = authInputs.some((el) => normalize(el.value || '') === '+86') || body.includes('+86');
  const phoneFormatError = body.includes('手机号码格式不对');
  const emailLoginPage = body.includes('邮箱登录') && (
    body.includes('仅支持子账号邮箱登录') ||
    authInputs.some((el) => normalize(el.placeholder || '').includes('邮箱') || normalize(el.placeholder || '').includes('账号'))
  );
  const hasPhoneLoginTab = body.includes('手机号登录') || body.includes('手机号码登录');
  const needsEmailLogin = phoneFormatError || (hasPhoneLoginTab && body.includes('邮箱登录') && hasPhonePrefix && shortPhoneDigits.length > 0 && shortPhoneDigits.length < 11);
  const hasRegionPage = body.includes('中国地区') && body.includes('其他地区');
  const authUrl = url.includes('/auth/') || url.includes('/authentication');
  const authPrompt = body.includes('授权登录') && (body.includes('隐私政策') || body.includes('账号ID') || body.includes('店铺名称'));
  const confirmForwardPrompt = body.includes('确认授权并前往') || body.includes('即将前往 Seller Central');
  const loginPrompt = body.includes('扫码登录') || body.includes('账号登录') || authUrl || hasRegionPage || authPrompt || confirmForwardPrompt;
  return {
    url,
    title,
    body,
    hasRegionPage,
    regionButtons,
    hasAuthButton: !!authButton,
    hasConfirmForwardButton: !!confirmForwardButton,
    hasConfirmForwardPrompt: confirmForwardPrompt,
    checkboxChecked: checkbox ? !!checkbox.checked : null,
    authInputCount: authInputs.length,
    authFilledCount,
    authPasswordFilledCount,
    authAccountFilledCount,
    shortPhoneDigits,
    hasPhonePrefix,
    phoneFormatError,
    emailLoginPage,
    needsEmailLogin,
    authCredentialsReady: authPasswordFilledCount > 0 && authAccountFilledCount > 0,
    homeReady: !loginPrompt && !hasRegionPage && !authButton && !confirmForwardButton && (portalReady || (hasHomeEntry && portalMarkers.slice(2).some((marker) => body.includes(marker)))),
    loginPrompt,
  };
})()
"""

CLICK_REGION_SCRIPT = r"""
(() => {
  const normalize = (value) => (value || '').replace(/\s+/g, ' ').trim();
  const visible = (el) => {
    const rect = el.getBoundingClientRect();
    const style = getComputedStyle(el);
    return rect.width > 0 && rect.height > 0 && style.visibility !== 'hidden' && style.display !== 'none';
  };
  const candidates = Array.from(document.querySelectorAll('a,button,div,span'))
    .filter((el) => visible(el) && normalize(el.innerText) === '商家中心')
    .sort((left, right) => {
      const leftRect = left.getBoundingClientRect();
      const rightRect = right.getBoundingClientRect();
      return (rightRect.x - leftRect.x) || (leftRect.width - rightRect.width);
    });
  const target = candidates[0];
  if (!target) {
    return { ok: false, reason: 'region_button_not_found' };
  }
  target.click();
  return { ok: true, className: target.className || '', tag: target.tagName };
})()
"""

CLICK_AUTH_SCRIPT = r"""
(() => {
  const normalize = (value) => (value || '').replace(/\s+/g, ' ').trim();
  const visible = (el) => {
    if (!el) {
      return false;
    }
    const rect = el.getBoundingClientRect();
    const style = getComputedStyle(el);
    return rect.width > 0 && rect.height > 0 && style.visibility !== 'hidden' && style.display !== 'none';
  };
  const body = normalize(document.body.innerText || '');
  const authInputs = Array.from(document.querySelectorAll('input')).filter(visible);
  const accountInputs = authInputs.filter((el) => String(el.type || '').toLowerCase() !== 'password');
  const shortPhoneDigits = accountInputs
    .map((el) => normalize(el.value || ''))
    .filter((value) => value !== '+86')
    .map((value) => value.replace(/\D/g, ''))
    .filter((digits) => digits !== '86')
    .find((digits) => digits.length > 0 && digits.length < 11) || '';
  const hasPhonePrefix = accountInputs.some((el) => normalize(el.value || '') === '+86') || body.includes('+86');
  const hasPhoneLoginTab = body.includes('手机号登录') || body.includes('手机号码登录');
  const shouldSwitchEmail = body.includes('手机号码格式不对') || (hasPhoneLoginTab && body.includes('邮箱登录') && hasPhonePrefix && shortPhoneDigits.length > 0 && shortPhoneDigits.length < 11);
  if (shouldSwitchEmail) {
    const emailTab = Array.from(document.querySelectorAll('button,div,span,a,label'))
      .filter(visible)
      .find((el) => normalize(el.innerText || el.textContent || '') === '邮箱登录')
      || Array.from(document.querySelectorAll('button,div,span,a,label'))
        .filter(visible)
        .find((el) => normalize(el.innerText || el.textContent || '').includes('邮箱登录'));
    if (emailTab) {
      emailTab.click();
      return { ok: true, stage: body.includes('手机号码格式不对') ? 'switched_to_email_after_phone_error' : 'switched_to_email_before_phone_error', shortPhoneDigits };
    }
    return { ok: false, stage: 'phone_error_email_tab_missing' };
  }
  const checkbox = Array.from(document.querySelectorAll('input[type="checkbox"], input[value="on"]'))[0] || null;
  if (checkbox && !checkbox.checked) {
    checkbox.click();
  }
  const button = Array.from(document.querySelectorAll('button'))
    .find((el) => (el.innerText || '').replace(/\s+/g, ' ').includes('授权登录'));
  if (!button) {
    return { ok: false, reason: 'auth_button_not_found', checked: checkbox ? !!checkbox.checked : null };
  }
  button.click();
  return { ok: true, checked: checkbox ? !!checkbox.checked : null, buttonText: (button.innerText || '').replace(/\s+/g, ' ').trim() };
})()
"""

EMAIL_LOGIN_AUTH_SCRIPT = r"""
(() => {
  const normalize = (value) => (value || '').replace(/\s+/g, ' ').trim();
  const visible = (el) => {
    if (!el) {
      return false;
    }
    const rect = el.getBoundingClientRect();
    const style = getComputedStyle(el);
    return rect.width > 0 && rect.height > 0 && style.visibility !== 'hidden' && style.display !== 'none';
  };
  const clickReal = (el) => {
    if (!el) {
      return;
    }
    const rect = el.getBoundingClientRect();
    const options = {
      bubbles: true,
      cancelable: true,
      view: window,
      clientX: rect.left + rect.width / 2,
      clientY: rect.top + rect.height / 2,
    };
    for (const type of ['pointerdown', 'mousedown', 'pointerup', 'mouseup', 'click']) {
      try {
        el.dispatchEvent(new MouseEvent(type, options));
      } catch (_error) {}
    }
    try {
      el.click();
    } catch (_error) {}
  };
  const body = normalize(document.body.innerText || '');
  const phoneFormatError = body.includes('手机号码格式不对');
  const inputs = Array.from(document.querySelectorAll('input')).filter(visible);
  const accountInputs = inputs.filter((el) => String(el.type || '').toLowerCase() !== 'password');
  const phoneDigits = accountInputs
    .map((el) => normalize(el.value || ''))
    .filter((value) => value !== '+86')
    .map((value) => value.replace(/\D/g, ''))
    .filter((digits) => digits !== '86')
    .find((digits) => digits.length > 0 && digits.length < 11) || '';
  const hasPhonePrefix = accountInputs.some((el) => normalize(el.value || '') === '+86') || body.includes('+86');
  const emailInput = accountInputs.find((el) => {
    const placeholder = normalize(el.placeholder || '');
    return placeholder.includes('邮箱') || placeholder.includes('账号');
  }) || null;
  const isEmailLoginPage = body.includes('邮箱登录') && (!!emailInput || body.includes('仅支持子账号邮箱登录'));
  const hasPhoneLoginTab = body.includes('手机号登录') || body.includes('手机号码登录');
  const needsEmailLogin = phoneFormatError || (hasPhoneLoginTab && body.includes('邮箱登录') && hasPhonePrefix && phoneDigits.length > 0 && phoneDigits.length < 11);

  if (needsEmailLogin && !isEmailLoginPage) {
    const emailTab = Array.from(document.querySelectorAll('button,div,span,a,label'))
      .filter(visible)
      .find((el) => normalize(el.innerText || el.textContent || '') === '邮箱登录')
      || Array.from(document.querySelectorAll('button,div,span,a,label'))
        .filter(visible)
        .find((el) => normalize(el.innerText || el.textContent || '').includes('邮箱登录'));
    if (!emailTab) {
      return { ok: false, stage: 'email_tab_missing', phoneDigits };
    }
    clickReal(emailTab);
    return { ok: true, stage: 'switched_to_email', phoneDigits };
  }

  if (!needsEmailLogin && !isEmailLoginPage) {
    return { ok: false, stage: 'not_email_login_page', phoneDigits };
  }

  const accountInput = emailInput || accountInputs[0] || null;
  if (!accountInput) {
    return { ok: false, stage: 'email_account_input_missing', phoneDigits };
  }
  const accountValue = normalize(accountInput.value || '');
  const accountHasEmail = /[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}/.test(accountValue);
  if (!accountHasEmail) {
    clickReal(accountInput);
    try {
      accountInput.focus();
    } catch (_error) {}
    const emailOptions = Array.from(document.querySelectorAll('div,li,span,p'))
      .filter(visible)
      .map((el) => {
        const text = normalize(el.innerText || el.textContent || '');
        const rect = el.getBoundingClientRect();
        return { el, text, area: rect.width * rect.height };
      })
      .filter((item) => /[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}/.test(item.text))
      .sort((left, right) => left.area - right.area);
    if (!emailOptions.length) {
      return { ok: false, stage: 'email_option_missing', phoneDigits, placeholder: normalize(accountInput.placeholder || '') };
    }
    const option = emailOptions[0];
    const emailText = (option.text.match(/[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}/) || [''])[0];
    clickReal(option.el);
    return { ok: true, stage: 'email_selected', emailText, phoneDigits };
  }

  const checkbox = Array.from(document.querySelectorAll('input[type="checkbox"], input[value="on"]'))[0] || null;
  if (checkbox && !checkbox.checked) {
    clickReal(checkbox);
  }
  const authButton = Array.from(document.querySelectorAll('button'))
    .filter(visible)
    .find((el) => normalize(el.innerText || '').includes('授权登录'))
    || Array.from(document.querySelectorAll('button,div,span,a'))
      .filter(visible)
      .find((el) => normalize(el.innerText || el.textContent || '').includes('授权登录'));
  if (!authButton) {
    return { ok: false, stage: 'auth_button_missing', accountValue, checked: checkbox ? !!checkbox.checked : null };
  }
  clickReal(authButton);
  return { ok: true, stage: 'email_auth_clicked', accountValue, checked: checkbox ? !!checkbox.checked : null };
})()
"""

EMAIL_ACCOUNT_LOGIN_AUTH_SCRIPT_TEMPLATE = r"""
(() => {
  const targetEmail = __EMAIL_ACCOUNT__;
  const normalize = (value) => (value || '').replace(/\s+/g, ' ').trim();
  const visible = (el) => {
    if (!el) return false;
    const rect = el.getBoundingClientRect();
    const style = getComputedStyle(el);
    return rect.width > 0 && rect.height > 0 && style.visibility !== 'hidden' && style.display !== 'none';
  };
  const clickReal = (el) => {
    if (!el) return;
    const rect = el.getBoundingClientRect();
    const options = {
      bubbles: true,
      cancelable: true,
      view: window,
      clientX: rect.left + rect.width / 2,
      clientY: rect.top + rect.height / 2,
    };
    for (const type of ['pointerdown', 'mousedown', 'pointerup', 'mouseup', 'click']) {
      try { el.dispatchEvent(new MouseEvent(type, options)); } catch (_error) {}
    }
    try { el.click(); } catch (_error) {}
  };
  const setNativeValue = (el, value) => {
    const descriptor = Object.getOwnPropertyDescriptor(Object.getPrototypeOf(el), 'value')
      || Object.getOwnPropertyDescriptor(HTMLInputElement.prototype, 'value');
    if (descriptor && descriptor.set) descriptor.set.call(el, value);
    else el.value = value;
  };
  const setInputValue = (el, value) => {
    clickReal(el);
    try { el.focus(); el.select(); } catch (_error) {}
    setNativeValue(el, '');
    try { el.dispatchEvent(new Event('input', { bubbles: true })); } catch (_error) {}
    let inserted = false;
    try { inserted = !!document.execCommand && document.execCommand('insertText', false, value); } catch (_error) {}
    if (!inserted || normalize(el.value || '').toLowerCase() !== String(value).toLowerCase()) {
      setNativeValue(el, value);
    }
    try {
      el.dispatchEvent(new InputEvent('input', { bubbles: true, data: value, inputType: 'insertText' }));
    } catch (_error) {
      try { el.dispatchEvent(new Event('input', { bubbles: true })); } catch (_innerError) {}
    }
    try { el.dispatchEvent(new Event('change', { bubbles: true })); } catch (_error) {}
  };
  const body = normalize(document.body.innerText || '');
  if (body.includes('仅支持登录下拉框内的账号')) {
    return { ok: false, stage: 'email_account_protected' };
  }
  const inputs = Array.from(document.querySelectorAll('input')).filter(visible);
  const accountInputs = inputs.filter((el) => String(el.type || '').toLowerCase() !== 'password');
  const phoneDigits = accountInputs
    .map((el) => normalize(el.value || ''))
    .filter((value) => value !== '+86')
    .map((value) => value.replace(/\D/g, ''))
    .filter((digits) => digits !== '86')
    .find((digits) => digits.length > 0 && digits.length < 11) || '';
  const hasPhonePrefix = accountInputs.some((el) => normalize(el.value || '') === '+86') || body.includes('+86');
  const emailInput = accountInputs.find((el) => {
    const placeholder = normalize(el.placeholder || '');
    return placeholder.includes('邮箱') || placeholder.includes('账号');
  }) || null;
  const hasPhoneLoginTab = body.includes('手机号登录') || body.includes('手机号码登录');
  const needsEmailLogin = body.includes('手机号码格式不对') || (hasPhoneLoginTab && body.includes('邮箱登录') && hasPhonePrefix && phoneDigits.length > 0 && phoneDigits.length < 11);
  const isEmailLoginPage = body.includes('邮箱登录') && (!!emailInput || body.includes('仅支持子账号邮箱登录'));

  if (needsEmailLogin && !isEmailLoginPage) {
    const emailTab = Array.from(document.querySelectorAll('button,div,span,a,label'))
      .filter(visible)
      .find((el) => normalize(el.innerText || el.textContent || '') === '邮箱登录')
      || Array.from(document.querySelectorAll('button,div,span,a,label'))
        .filter(visible)
        .find((el) => normalize(el.innerText || el.textContent || '').includes('邮箱登录'));
    if (!emailTab) return { ok: false, stage: 'email_tab_missing', phoneDigits };
    clickReal(emailTab);
    return { ok: true, stage: 'switched_to_email', phoneDigits };
  }
  if (!needsEmailLogin && !isEmailLoginPage) return { ok: false, stage: 'not_email_login_page', phoneDigits };
  if (!targetEmail) return { ok: false, stage: 'email_account_empty', phoneDigits };

  const accountInput = emailInput || accountInputs[0] || null;
  if (!accountInput) return { ok: false, stage: 'email_account_input_missing', phoneDigits };
  const accountValue = normalize(accountInput.value || '');
  if (accountValue.toLowerCase() !== String(targetEmail).toLowerCase()) {
    setInputValue(accountInput, targetEmail);
    return { ok: true, stage: 'email_account_filled', accountValue: targetEmail, phoneDigits };
  }

  const checkbox = Array.from(document.querySelectorAll('input[type="checkbox"], input[value="on"]'))[0] || null;
  if (checkbox && !checkbox.checked) {
    checkbox.click();
  }
  const button = Array.from(document.querySelectorAll('button'))
    .find((el) => (el.innerText || '').replace(/\s+/g, ' ').includes('授权登录'));
  if (!button) {
    return { ok: false, stage: 'auth_button_missing', accountValue, checked: checkbox ? !!checkbox.checked : null };
  }
  button.click();
  return { ok: true, stage: 'email_auth_clicked', accountValue, checked: checkbox ? !!checkbox.checked : null };
})()
"""

CLICK_CONFIRM_FORWARD_SCRIPT = r"""
(() => {
  const normalize = (value) => (value || '').replace(/\s+/g, ' ').trim();
  const visible = (el) => {
    const rect = el.getBoundingClientRect();
    const style = getComputedStyle(el);
    return rect.width > 0 && rect.height > 0 && style.visibility !== 'hidden' && style.display !== 'none';
  };
  const checkbox = Array.from(document.querySelectorAll('input[type="checkbox"], input[value="on"]'))[0] || null;
  if (checkbox && !checkbox.checked) {
    checkbox.click();
  }
  const button = Array.from(document.querySelectorAll('button'))
    .find((el) => visible(el) && normalize(el.innerText).includes('确认授权并前往'));
  if (!button) {
    return { ok: false, reason: 'confirm_forward_button_not_found', checked: checkbox ? !!checkbox.checked : null };
  }
  button.click();
  return {
    ok: true,
    checked: checkbox ? !!checkbox.checked : null,
    buttonText: normalize(button.innerText || ''),
  };
})()
"""

QC_EXTRACT_SCRIPT = r"""
(() => {
  const normalize = (value) => (value || '').replace(/\u00a0/g, ' ').replace(/\s+/g, ' ').trim();
  const visible = (el) => {
    const rect = el.getBoundingClientRect();
    const style = getComputedStyle(el);
    return rect.width > 0 && rect.height > 0 && style.visibility !== 'hidden' && style.display !== 'none';
  };
  const rows = Array.from(document.querySelectorAll('tbody tr'))
    .filter((row) => visible(row))
    .map((row) => {
      const cells = Array.from(row.querySelectorAll('td')).map((cell) => normalize(cell.innerText));
      if (cells.length < 4) {
        return null;
      }
      return {
        productInfo: cells[0] || '',
        skuInfo: cells[1] || '',
        prepareOrderNo: cells[2] || '',
        latestQcTime: cells[3] || '',
        operation: cells[4] || '',
      };
    })
    .filter(Boolean);
  const body = normalize(document.body.innerText || '');
  const totalMatch = body.match(/共有\s*(\d+)\s*条/);
  return {
    url: location.href,
    title: document.title,
    totalRowsText: totalMatch ? Number(totalMatch[1]) : null,
    bodyPreview: body.slice(0, 4000),
    rows
  };
})()
"""

QC_READY_SCRIPT = r"""
(() => {
  const normalize = (value) => (value || '').replace(/\u00a0/g, ' ').replace(/\s+/g, ' ').trim();
  const body = normalize(document.body.innerText || '');
  return {
    url: location.href,
    title: normalize(document.title || ''),
    body,
    rowCount: document.querySelectorAll('tbody tr').length,
    hasNoData: body.includes('暂无数据'),
    hasTotal: /共有\s*\d+\s*条/.test(body),
    loginPrompt: body.includes('扫码登录') || body.includes('账号登录') || location.href.includes('/login'),
  };
})()
"""

URGENT_METRICS_SCRIPT = r"""
(() => {
  const normalize = (value) => (value || '').replace(/\u00a0/g, ' ').replace(/\s+/g, ' ').trim();
  const escapeRegExp = (value) => value.replace(/[.*+?^${}()|[\]\\]/g, '\\$&');
  const url = location.href || '';
  const body = normalize(document.body.innerText || '');
  const title = normalize(document.title || '');
  const sectionStart = body.indexOf('快速筛选');
  let section = body;
  if (sectionStart >= 0) {
    const endMarkers = ['备货母单号', '备货单号', '货号', '查询'];
    let sectionEnd = -1;
    for (const marker of endMarkers) {
      const idx = body.indexOf(marker, sectionStart);
      if (idx > sectionStart && (sectionEnd < 0 || idx < sectionEnd)) {
        sectionEnd = idx;
      }
    }
    section = body.slice(sectionStart, sectionEnd > sectionStart ? sectionEnd : body.length);
  }
  const readCount = (label) => {
    const regex = new RegExp(escapeRegExp(label) + '\\s*(\\d[\\d,]*)');
    const match = section.match(regex) || body.match(regex);
    return match ? Number(match[1].replace(/,/g, '')) : null;
  };
  return {
    url,
    title,
    body,
    section,
    shipOverdue: readCount('发货已逾期'),
    arrivalOverdue: readCount('到货已逾期'),
    isUrgentPage: url.includes('/stock/fully-mgt/order-manage-urgency') && body.includes('快速筛选'),
    ready: section.includes('发货已逾期') && section.includes('到货已逾期') && section.includes('到货即将逾期'),
    loginPrompt: body.includes('扫码登录') || body.includes('账号登录') || url.includes('/login'),
  };
})()
"""

URGENT_PENDING_PRICE_STATE_SCRIPT = r"""
(() => {
  const normalize = (value) => (value || '').replace(/\u00a0/g, ' ').replace(/\s+/g, ' ').trim();
  const looksLikeCurrency = (value) => /[¥￥]\s*\d+(?:\.\d+)?/.test(normalize(value));
  const visible = (el) => {
    const rect = el.getBoundingClientRect();
    const style = getComputedStyle(el);
    return rect.width > 0 && rect.height > 0 && style.visibility !== 'hidden' && style.display !== 'none';
  };
  const requiredHeaderAliases = [
    ['备货单号'],
    ['商品信息'],
    ['状态'],
    ['SKU信息'],
    ['申报价格(CNY)', '申报价格', '申报价'],
    ['备货件数'],
    ['送货/入库数'],
    ['备货单创建时间', '创建时间'],
  ];
  const hasHeader = (headers, aliases) => aliases.some((alias) => headers.some((value) => value.includes(alias)));
  const findUrgentTable = () => {
    const candidates = Array.from(document.querySelectorAll('table'))
      .filter((table) => visible(table))
      .map((table) => {
        const headers = Array.from(table.querySelectorAll('thead th')).map((cell) => normalize(cell.innerText || ''));
        const bodyRowCount = table.querySelectorAll('tbody tr').length;
        return { table, headers, bodyRowCount };
      });
    const matches = candidates
      .filter((item) => requiredHeaderAliases.every((aliases) => hasHeader(item.headers, aliases)))
      .sort((left, right) => right.bodyRowCount - left.bodyRowCount);
    return matches[0] || null;
  };
  const isActiveTabNode = (el) => {
    if (!el) return false;
    const chain = [el, el.parentElement, el.parentElement ? el.parentElement.parentElement : null].filter(Boolean);
    return chain.some((node) => {
      const className = String(node.className || '');
      const ariaSelected = String(node.getAttribute && (node.getAttribute('aria-selected') || '')).toLowerCase();
      return className.includes('active') || className.includes('selected') || ariaSelected === 'true';
    });
  };
  const url = location.href || '';
  const body = normalize(document.body.innerText || '');
  const title = normalize(document.title || '');
  const visibleTables = Array.from(document.querySelectorAll('table')).filter((table) => visible(table));
  const tableInfo = findUrgentTable();
  const headers = tableInfo ? tableInfo.headers : [];
  const headersPreview = visibleTables
    .flatMap((table) => Array.from(table.querySelectorAll('thead th')).map((cell) => normalize(cell.innerText || '')).filter(Boolean))
    .slice(0, 20);
  const pendingTabNodes = Array.from(document.querySelectorAll('div,button,span,a'))
    .filter((el) => visible(el) && /^待发货(?:\(\d+\))?$/.test(normalize(el.innerText || '')));
  const activeTabNode = pendingTabNodes.find((el) => isActiveTabNode(el)) || null;
  const pendingTabText = pendingTabNodes[0] ? normalize(pendingTabNodes[0].innerText || '') : '';
  const pendingMatch = pendingTabText.match(/\((\d+)\)/);
  const tableRows = Array.from(document.querySelectorAll('tbody tr')).filter((row) => visible(row));
  const actualRowCount = tableRows.filter((row) => /\bWB\d{6,}\b/.test(normalize(row.innerText || ''))).length;
  return {
    url,
    title,
    body,
    isUrgentPage: url.includes('/stock/fully-mgt/order-manage-urgency') && body.includes('快速筛选'),
    activeTabText: activeTabNode ? normalize(activeTabNode.innerText || '') : '',
    pendingTabSeen: pendingTabNodes.length > 0,
    pendingCount: pendingMatch ? Number(pendingMatch[1]) : 0,
    rowCount: tableRows.length,
    actualRowCount,
    hasUrgentTable: !!tableInfo,
    tableCount: visibleTables.length,
    headersPreview,
    hasPriceHeader: hasHeader(headers, ['申报价格(CNY)', '申报价格', '申报价']),
    hasNoData: body.includes('暂无数据'),
    loading: body.includes('加载中'),
    loginPrompt: body.includes('扫码登录') || body.includes('账号登录') || url.includes('/login'),
  };
})()
"""

URGENT_CLICK_PENDING_TAB_SCRIPT = r"""
(() => {
  const normalize = (value) => (value || '').replace(/\u00a0/g, ' ').replace(/\s+/g, ' ').trim();
  const visible = (el) => {
    const rect = el.getBoundingClientRect();
    const style = getComputedStyle(el);
    return rect.width > 0 && rect.height > 0 && style.visibility !== 'hidden' && style.display !== 'none';
  };
  const clickLikeUser = (el) => {
    if (!el) return false;
    const rect = el.getBoundingClientRect();
    const options = { bubbles: true, cancelable: true, clientX: rect.left + rect.width / 2, clientY: rect.top + rect.height / 2 };
    ['pointerdown', 'mousedown', 'pointerup', 'mouseup', 'click'].forEach((type) => {
      try { el.dispatchEvent(new MouseEvent(type, options)); } catch (_error) {}
    });
    try { el.click(); } catch (_error) {}
    return true;
  };
  const target = Array.from(document.querySelectorAll('div,button,span,a'))
    .find((el) => visible(el) && /^待发货(?:\(\d+\))?$/.test(normalize(el.innerText || '')));
  if (!target) {
    return { clicked: false, reason: 'pending_tab_not_found' };
  }
  clickLikeUser(target);
  clickLikeUser(target.parentElement);
  return { clicked: true, text: normalize(target.innerText || '') };
})()
"""

URGENT_PENDING_PRICE_EXTRACT_SCRIPT = r"""
(() => {
  const normalize = (value) => (value || '').replace(/\u00a0/g, ' ').replace(/\s+/g, ' ').trim();
  const looksLikeCurrency = (value) => /[¥￥]\s*\d+(?:\.\d+)?/.test(normalize(value));
  const visible = (el) => {
    const rect = el.getBoundingClientRect();
    const style = getComputedStyle(el);
    return rect.width > 0 && rect.height > 0 && style.visibility !== 'hidden' && style.display !== 'none';
  };
  const requiredHeaderAliases = [
    ['备货单号'],
    ['商品信息'],
    ['状态'],
    ['SKU信息'],
    ['申报价格(CNY)', '申报价格', '申报价'],
    ['备货件数'],
    ['送货/入库数'],
    ['备货单创建时间', '创建时间'],
  ];
  const hasHeader = (headers, aliases) => aliases.some((alias) => headers.some((value) => value.includes(alias)));
  const headerIndex = (headers, aliases) => headers.findIndex((header) => aliases.some((alias) => header.includes(alias)));
  const findUrgentTable = () => {
    const candidates = Array.from(document.querySelectorAll('table'))
      .filter((table) => visible(table))
      .map((table) => {
        const headers = Array.from(table.querySelectorAll('thead th')).map((cell) => normalize(cell.innerText || ''));
        const bodyRowCount = table.querySelectorAll('tbody tr').length;
        return { table, headers, bodyRowCount };
      });
    const matches = candidates
      .filter((item) => requiredHeaderAliases.every((aliases) => hasHeader(item.headers, aliases)))
      .sort((left, right) => right.bodyRowCount - left.bodyRowCount);
    return matches[0] || null;
  };
  const body = normalize(document.body.innerText || '');
  const url = location.href || '';
  const tableInfo = findUrgentTable();
  const allVisibleTables = Array.from(document.querySelectorAll('table')).filter((table) => visible(table));
  const headersPreview = allVisibleTables
    .flatMap((table) => Array.from(table.querySelectorAll('thead th')).map((cell) => normalize(cell.innerText || '')).filter(Boolean))
    .slice(0, 20);
  const dataRows = Array.from(document.querySelectorAll('tbody tr')).filter((row) => visible(row));
  const parseFallbackRows = () => {
    const fallbackRows = [];
    let currentContext = null;
    for (const row of dataRows) {
      const cells = Array.from(row.querySelectorAll('td')).map((cell) => normalize(cell.innerText || ''));
      const merged = normalize(cells.join(' '));
      if (!merged || merged === '合计' || merged.startsWith('合计 ')) {
        continue;
      }
      const orderMatch = merged.match(/\bWB\d{6,}\b/);
      const priceMatches = Array.from(merged.matchAll(/[¥￥]\s*\d+(?:\.\d+)?/g)).map((match) => match[0]);
      const createdMatch = merged.match(/\d{4}-\d{2}-\d{2} \d{2}:\d{2}(?::\d{2})?/);
      if (orderMatch) {
        currentContext = {
          prepareOrderNo: orderMatch[0],
          productInfo: merged,
          status: merged.includes('待发货') ? '待发货' : '',
          createdTime: createdMatch ? createdMatch[0] : '',
        };
        fallbackRows.push({
          rowType: 'fallback-main',
          cells,
          prepareOrderNo: currentContext.prepareOrderNo,
          productInfo: currentContext.productInfo,
          status: currentContext.status,
          skuInfo: '',
          declaredPriceText: priceMatches[0] || '',
          createdTime: currentContext.createdTime,
        });
        continue;
      }
      if (!currentContext || !priceMatches.length || !(merged.includes('属性') || merged.includes('SKU'))) {
        continue;
      }
      fallbackRows.push({
        rowType: 'fallback-child',
        cells,
        prepareOrderNo: currentContext.prepareOrderNo,
        productInfo: currentContext.productInfo,
        status: currentContext.status,
        skuInfo: merged,
        declaredPriceText: priceMatches[0],
        createdTime: currentContext.createdTime,
      });
    }
    return fallbackRows.filter((item) => looksLikeCurrency(item.declaredPriceText));
  };
  if (!tableInfo) {
    const fallbackRows = parseFallbackRows();
    return {
      url,
      title: document.title,
      bodyPreview: body.slice(0, 5000),
      headers: headersPreview,
      rows: fallbackRows,
      hasNoData: body.includes('暂无数据'),
      rowCount: fallbackRows.length,
      tableFound: fallbackRows.length > 0,
      fallbackUsed: fallbackRows.length > 0,
      tableCount: allVisibleTables.length,
      headersPreview,
      isUrgentPage: url.includes('/stock/fully-mgt/order-manage-urgency') && body.includes('快速筛选'),
    };
  }
  const headers = tableInfo.headers;
  const rows = [];
  let currentContext = null;
  const orderIndex = headerIndex(headers, ['备货单号']);
  const productIndex = headerIndex(headers, ['商品信息']);
  const statusIndex = headerIndex(headers, ['状态']);
  const skuIndex = headerIndex(headers, ['SKU信息']);
  const priceIndex = headerIndex(headers, ['申报价格(CNY)', '申报价格', '申报价']);
  const createdIndex = headerIndex(headers, ['备货单创建时间', '创建时间']);
  for (const row of dataRows) {
    const cells = Array.from(row.querySelectorAll('td')).map((cell) => normalize(cell.innerText || ''));
    if (!cells.length) {
      continue;
    }
    const merged = normalize(cells.join(' '));
    if (!merged || merged === '合计' || merged.startsWith('合计 ')) {
      continue;
    }
    const mapped = {};
    const size = Math.max(headers.length, cells.length);
    for (let i = 0; i < size; i += 1) {
      const key = headers[i] || `col${i}`;
      mapped[key] = cells[i] || '';
    }
    let offset = 0;
    const directOrderText = orderIndex >= 0 && orderIndex < cells.length ? cells[orderIndex] : '';
    if (orderIndex > 0 && !/\bWB\d{6,}\b/.test(directOrderText) && /\bWB\d{6,}\b/.test(cells[orderIndex - 1] || '')) {
      offset = -1;
    }
    const cellByIndex = (index) => {
      if (index < 0) return '';
      const shiftedIndex = index + offset;
      return shiftedIndex >= 0 && shiftedIndex < cells.length ? (cells[shiftedIndex] || '') : '';
    };
    const orderText = orderIndex >= 0 ? cellByIndex(orderIndex) : merged;
    if (/\bWB\d{6,}\b/.test(orderText)) {
      currentContext = {
        prepareOrderNo: cellByIndex(orderIndex) || mapped['备货单号'] || '',
        productInfo: cellByIndex(productIndex) || mapped['商品信息'] || '',
        status: cellByIndex(statusIndex) || mapped['状态'] || '',
        createdTime: cellByIndex(createdIndex) || mapped['备货单创建时间'] || '',
      };
      rows.push({
        rowType: 'main',
        cells,
        prepareOrderNo: currentContext.prepareOrderNo,
        productInfo: currentContext.productInfo,
        status: currentContext.status,
        skuInfo: cellByIndex(skuIndex) || (skuIndex >= 0 ? (mapped['SKU信息'] || '') : ''),
        declaredPriceText: looksLikeCurrency(cellByIndex(priceIndex) || mapped['申报价格(CNY)'] || '') ? (cellByIndex(priceIndex) || mapped['申报价格(CNY)'] || '') : '',
        createdTime: currentContext.createdTime,
      });
      continue;
    }
    if (!currentContext) {
      continue;
    }
    const childSkuInfo = cells[0] || '';
    const childPriceText = cells.length > 1 ? cells[1] : '';
    if (!looksLikeCurrency(childPriceText) || !(childSkuInfo.includes('属性') || childSkuInfo.includes('SKU'))) {
      continue;
    }
    rows.push({
      rowType: 'child',
      cells,
      prepareOrderNo: currentContext.prepareOrderNo,
      productInfo: currentContext.productInfo,
      status: currentContext.status,
      skuInfo: childSkuInfo,
      declaredPriceText: childPriceText,
      createdTime: currentContext.createdTime,
    });
  }
  return {
    url,
    title: document.title,
    bodyPreview: body.slice(0, 5000),
    headers,
    rows,
    hasNoData: body.includes('暂无数据'),
    rowCount: rows.length,
    tableFound: true,
    fallbackUsed: false,
    tableCount: allVisibleTables.length,
    headersPreview,
    isUrgentPage: url.includes('/stock/fully-mgt/order-manage-urgency') && body.includes('快速筛选'),
  };
})()
"""

URGENT_ARRIVAL_OVERDUE_STATE_SCRIPT = r"""
(() => {
  const normalize = (value) => (value || '').replace(/\u00a0/g, ' ').replace(/\s+/g, ' ').trim();
  const visible = (el) => {
    const rect = el.getBoundingClientRect();
    const style = getComputedStyle(el);
    return rect.width > 0 && rect.height > 0 && style.visibility !== 'hidden' && style.display !== 'none';
  };
  const isActualRow = (cells) => {
    if (!cells.length) return false;
    const merged = normalize(cells.join(' '));
    if (!merged || merged === '合计' || merged.startsWith('合计 ')) return false;
    return /\d{4}-\d{2}-\d{2} \d{2}:\d{2}/.test(merged) || merged.includes('WB');
  };
  const rows = Array.from(document.querySelectorAll('tbody tr'))
    .filter((row) => visible(row))
    .map((row) => Array.from(row.querySelectorAll('td')).map((cell) => normalize(cell.innerText || '')));
  const actualRows = rows.filter((cells) => isActualRow(cells));
  const body = normalize(document.body.innerText || '');
  const card = Array.from(document.querySelectorAll('div.quick-overdue-filter_card__plOUM, div, span'))
    .find((el) => visible(el) && /^到货已逾期\s+\d+$/.test(normalize(el.innerText || '')));
  const cardStyle = card ? getComputedStyle(card) : null;
  const borderColor = cardStyle ? String(cardStyle.borderColor || '') : '';
  const selected = borderColor.includes('64, 124, 255');
  return {
    url: location.href,
    title: document.title,
    body,
    hasCard: !!card,
    cardText: card ? normalize(card.innerText || '') : '',
    cardBorderColor: borderColor,
    selected,
    rowCount: rows.length,
    actualRowCount: actualRows.length,
    firstRows: actualRows.slice(0, 3),
    hasNoData: body.includes('暂无数据'),
    loading: body.includes('加载中'),
    loginPrompt: body.includes('扫码登录') || body.includes('账号登录') || location.href.includes('/login'),
  };
})()
"""

URGENT_MARK_ARRIVAL_OVERDUE_CARD_SCRIPT = r"""
(() => {
  const normalize = (value) => (value || '').replace(/\u00a0/g, ' ').replace(/\s+/g, ' ').trim();
  const visible = (el) => {
    const rect = el.getBoundingClientRect();
    const style = getComputedStyle(el);
    return rect.width > 0 && rect.height > 0 && style.visibility !== 'hidden' && style.display !== 'none';
  };
  document.querySelectorAll('#codex-arrival-overdue-card').forEach((el) => el.removeAttribute('id'));
  const target = Array.from(document.querySelectorAll('div.quick-overdue-filter_card__plOUM'))
    .find((el) => visible(el) && /^到货已逾期\s+\d+$/.test(normalize(el.innerText || '')));
  if (!target) {
    return { ok: false, reason: 'arrival_overdue_card_not_found' };
  }
  target.id = 'codex-arrival-overdue-card';
  return { ok: true, text: normalize(target.innerText || '') };
})()
"""

URGENT_ARRIVAL_OVERDUE_EXTRACT_SCRIPT = r"""
(() => {
  const normalize = (value) => (value || '').replace(/\u00a0/g, ' ').replace(/\s+/g, ' ').trim();
  const visible = (el) => {
    const rect = el.getBoundingClientRect();
    const style = getComputedStyle(el);
    return rect.width > 0 && rect.height > 0 && style.visibility !== 'hidden' && style.display !== 'none';
  };
  const headers = Array.from(document.querySelectorAll('thead th')).map((cell) => normalize(cell.innerText || ''));
  const isActualRow = (cells) => {
    if (!cells.length) return false;
    const merged = normalize(cells.join(' '));
    if (!merged || merged === '合计' || merged.startsWith('合计 ')) return false;
    return /\d{4}-\d{2}-\d{2} \d{2}:\d{2}/.test(merged) || merged.includes('WB');
  };
  const rows = Array.from(document.querySelectorAll('tbody tr'))
    .filter((row) => visible(row))
    .map((row) => {
      const cells = Array.from(row.querySelectorAll('td')).map((cell) => normalize(cell.innerText || ''));
      const mapped = {};
      const size = Math.max(headers.length, cells.length);
      for (let i = 0; i < size; i += 1) {
        const key = headers[i] || `col${i}`;
        mapped[key] = cells[i] || '';
      }
      return { cells, mapped };
    })
    .filter((row) => isActualRow(row.cells));
  const body = normalize(document.body.innerText || '');
  return {
    url: location.href,
    title: document.title,
    bodyPreview: body.slice(0, 5000),
    headers,
    rows,
    hasNoData: body.includes('暂无数据'),
    actualRowCount: rows.length,
  };
})()
"""

GOVERN_METRICS_SCRIPT = r"""
(() => {
  const normalize = (value) => (value || '').replace(/\u00a0/g, ' ').replace(/\s+/g, ' ').trim();
  const escapeRegExp = (value) => value.replace(/[.*+?^${}()|[\]\\]/g, '\\$&');
  const body = normalize(document.body.innerText || '');
  const title = normalize(document.title || '');
  const sectionStart = body.indexOf('涉嫌违反政策');
  let section = body;
  if (sectionStart >= 0) {
    const endMarkers = ['展开', '全部消息', '重要通知'];
    let sectionEnd = -1;
    for (const marker of endMarkers) {
      const idx = body.indexOf(marker, sectionStart + 1);
      if (idx > sectionStart && (sectionEnd < 0 || idx < sectionEnd)) {
        sectionEnd = idx;
      }
    }
    section = body.slice(sectionStart, sectionEnd > sectionStart ? sectionEnd : body.length);
  }
  const readCount = (label) => {
    const regex = new RegExp(escapeRegExp(label) + '\\s*(\\d[\\d,]*)');
    const match = section.match(regex);
    return match ? Number(match[1].replace(/,/g, '')) : null;
  };
  return {
    url: location.href,
    title,
    body,
    section,
    ipComplaintCount: readCount('知识产权投诉'),
    troCount: readCount('临时限制令（TRO）'),
    ready: section.includes('涉嫌违反政策') && section.includes('知识产权投诉') && section.includes('临时限制令（TRO）'),
    loginPrompt: body.includes('扫码登录') || body.includes('账号登录') || location.href.includes('/login'),
  };
})()
"""

SHIPPING_FILTER_STATE_SCRIPT = r"""
(() => {
  const normalize = (value) => (value || '').replace(/\u00a0/g, ' ').replace(/\s+/g, ' ').trim();
  const body = normalize(document.body.innerText || '');
  const title = normalize(document.title || '');
  const jitRow = Array.from(document.querySelectorAll('.index-module__row___OoknQ')).find((el) => {
    const label = el.querySelector('.index-module__row_label___3WV-t');
    return label && normalize(label.innerText) === '是否JIT';
  });
  const jitInput = jitRow ? jitRow.querySelector('input[data-testid="beast-core-select-htmlInput"]') : null;
  const activeTab = Array.from(document.querySelectorAll('.TAB_active, .TAB_lineLabelActive_5-117-0'))
    .map((el) => normalize(el.innerText))
    .find(Boolean) || '';
  const totalMatch = body.match(/共有\s*(\d+)\s*条/);
  return {
    url: location.href,
    title,
    body,
    isExpanded: body.includes('是否JIT') && body.includes('发货时间'),
    jitValue: jitInput ? normalize(jitInput.value || '') : '',
    activeTab,
    rowCount: document.querySelectorAll('tbody tr').length,
    totalRowsText: totalMatch ? Number(totalMatch[1]) : null,
    hasNoData: body.includes('暂无数据'),
    loading: body.includes('加载中'),
    loginPrompt: body.includes('扫码登录') || body.includes('账号登录') || location.href.includes('/login'),
  };
})()
"""

SHIPPING_CLICK_EXPAND_SCRIPT = r"""
(() => {
  const normalize = (value) => (value || '').replace(/\s+/g, ' ').trim();
  const visible = (el) => {
    const rect = el.getBoundingClientRect();
    const style = getComputedStyle(el);
    return rect.width > 0 && rect.height > 0 && style.visibility !== 'hidden' && style.display !== 'none';
  };
  if (normalize(document.body.innerText || '').includes('是否JIT')) {
    return { ok: true, state: 'already_expanded' };
  }
  const target = Array.from(document.querySelectorAll('a,button,div,span'))
    .find((el) => visible(el) && normalize(el.innerText) === '展开');
  if (!target) {
    return { ok: false, reason: 'expand_not_found' };
  }
  target.click();
  return { ok: true, state: 'clicked_expand' };
})()
"""

SHIPPING_OPEN_JIT_SELECT_SCRIPT = r"""
(() => {
  const normalize = (value) => (value || '').replace(/\s+/g, ' ').trim();
  const row = Array.from(document.querySelectorAll('.index-module__row___OoknQ')).find((el) => {
    const label = el.querySelector('.index-module__row_label___3WV-t');
    return label && normalize(label.innerText) === '是否JIT';
  });
  if (!row) {
    return { ok: false, reason: 'jit_row_not_found' };
  }
  const input = row.querySelector('input[data-testid="beast-core-select-htmlInput"]');
  if (input && normalize(input.value || '') === '否') {
    return { ok: true, state: 'already_no' };
  }
  const target = row.querySelector('[data-testid="beast-core-select-header"], [data-testid="beast-core-select"]');
  if (!target) {
    return { ok: false, reason: 'jit_select_not_found' };
  }
  target.click();
  return { ok: true, state: 'opened_select' };
})()
"""

SHIPPING_SELECT_JIT_NO_SCRIPT = r"""
(() => {
  const normalize = (value) => (value || '').replace(/\s+/g, ' ').trim();
  const visible = (el) => {
    const rect = el.getBoundingClientRect();
    const style = getComputedStyle(el);
    return rect.width > 0 && rect.height > 0 && style.visibility !== 'hidden' && style.display !== 'none';
  };
  const target = Array.from(document.querySelectorAll('li, div, span'))
    .find((el) => visible(el) && normalize(el.innerText) === '否');
  if (!target) {
    return { ok: false, reason: 'jit_no_option_not_found' };
  }
  target.click();
  return { ok: true, state: 'selected_no' };
})()
"""

SHIPPING_CLICK_QUERY_SCRIPT = r"""
(() => {
  const normalize = (value) => (value || '').replace(/\s+/g, ' ').trim();
  const visible = (el) => {
    const rect = el.getBoundingClientRect();
    const style = getComputedStyle(el);
    return rect.width > 0 && rect.height > 0 && style.visibility !== 'hidden' && style.display !== 'none';
  };
  const target = Array.from(document.querySelectorAll('button, div, span'))
    .find((el) => visible(el) && normalize(el.innerText) === '查询');
  if (!target) {
    return { ok: false, reason: 'query_not_found' };
  }
  target.click();
  return { ok: true, state: 'clicked_query' };
})()
"""

SHIPPING_CLICK_WAITING_TAB_SCRIPT = r"""
(() => {
  const normalize = (value) => (value || '').replace(/\s+/g, ' ').trim();
  const visible = (el) => {
    const rect = el.getBoundingClientRect();
    const style = getComputedStyle(el);
    return rect.width > 0 && rect.height > 0 && style.visibility !== 'hidden' && style.display !== 'none';
  };
  const active = Array.from(document.querySelectorAll('.TAB_active, .TAB_lineLabelActive_5-117-0'))
    .map((el) => normalize(el.innerText))
    .find(Boolean);
  if (active === '待仓库收货') {
    return { ok: true, state: 'already_waiting' };
  }
  const target = Array.from(document.querySelectorAll('div, span, a'))
    .find((el) => visible(el) && normalize(el.innerText) === '待仓库收货');
  if (!target) {
    return { ok: false, reason: 'waiting_tab_not_found' };
  }
  target.click();
  return { ok: true, state: 'clicked_waiting_tab' };
})()
"""

SHIPPING_LIST_EXTRACT_SCRIPT = r"""
(() => {
  const normalize = (value) => (value || '').replace(/\u00a0/g, ' ').replace(/\s+/g, ' ').trim();
  const visible = (el) => {
    const rect = el.getBoundingClientRect();
    const style = getComputedStyle(el);
    return rect.width > 0 && rect.height > 0 && style.visibility !== 'hidden' && style.display !== 'none';
  };
  const rows = Array.from(document.querySelectorAll('tbody tr'))
    .filter((row) => visible(row))
    .map((row) => ({
      text: normalize(row.innerText),
      cells: Array.from(row.querySelectorAll('td')).map((cell) => normalize(cell.innerText)),
    }))
    .filter((row) => row.cells.length >= 10);
  const body = normalize(document.body.innerText || '');
  const totalMatch = body.match(/共有\s*(\d+)\s*条/);
  return {
    url: location.href,
    title: document.title,
    totalRowsText: totalMatch ? Number(totalMatch[1]) : null,
    bodyPreview: body.slice(0, 5000),
    rows,
  };
})()
"""

RETURN_ORDER_STATE_SCRIPT = r"""
(() => {
  const normalize = (value) => (value || '').replace(/\u00a0/g, ' ').replace(/\s+/g, ' ').trim();
  const visible = (el) => {
    const rect = el.getBoundingClientRect();
    const style = getComputedStyle(el);
    return rect.width > 0 && rect.height > 0 && style.visibility !== 'hidden' && style.display !== 'none';
  };
  const title = normalize(document.title || '');
  const body = normalize(document.body.innerText || '');
  const queryButton = Array.from(document.querySelectorAll('button,a,span,div'))
    .find((el) => visible(el) && normalize(el.innerText) === '查询');
  const rowCount = Array.from(document.querySelectorAll('tbody tr')).filter((row) => visible(row)).length;
  const totalMatch = body.match(/(?:共有|共)\s*(\d+)\s*条/);
  return {
    url: location.href,
    title,
    body: body.slice(0, 5000),
    rowCount,
    totalRowsText: totalMatch ? Number(totalMatch[1]) : null,
    hasNoData: body.includes('暂无数据') || body.includes('暂无结果'),
    hasQueryButton: !!queryButton,
    loading: body.includes('加载中') || body.includes('查询中'),
    loginPrompt: body.includes('授权登录') || body.includes('扫码登录') || body.includes('账号登录'),
  };
})()
"""

RETURN_ORDER_CLICK_QUERY_SCRIPT = r"""
(() => {
  const normalize = (value) => (value || '').replace(/\u00a0/g, ' ').replace(/\s+/g, ' ').trim();
  const visible = (el) => {
    const rect = el.getBoundingClientRect();
    const style = getComputedStyle(el);
    return rect.width > 0 && rect.height > 0 && style.visibility !== 'hidden' && style.display !== 'none';
  };
  const button = Array.from(document.querySelectorAll('button,a,span,div'))
    .find((el) => visible(el) && normalize(el.innerText) === '查询');
  if (!button) {
    return false;
  }
  button.click();
  return true;
})()
"""

RETURN_ORDER_EXTRACT_SCRIPT = r"""
(() => {
  const normalize = (value) => (value || '').replace(/\u00a0/g, ' ').replace(/\s+/g, ' ').trim();
  const visible = (el) => {
    const rect = el.getBoundingClientRect();
    const style = getComputedStyle(el);
    return rect.width > 0 && rect.height > 0 && style.visibility !== 'hidden' && style.display !== 'none';
  };
  const headers = Array.from(document.querySelectorAll('thead th')).map((cell) => normalize(cell.innerText));
  const rows = Array.from(document.querySelectorAll('tbody tr'))
    .filter((row) => visible(row))
    .map((row) => {
      const cells = Array.from(row.querySelectorAll('td')).map((cell) => normalize(cell.innerText));
      const mapped = {};
      const size = Math.max(headers.length, cells.length);
      for (let i = 0; i < size; i += 1) {
        const key = headers[i] || `col${i}`;
        mapped[key] = cells[i] || '';
      }
      return {
        text: normalize(row.innerText),
        cells,
        mapped,
      };
    })
    .filter((row) => row.cells.length >= 6);
  const body = normalize(document.body.innerText || '');
  const totalMatch = body.match(/(?:共有|共)\s*(\d+)\s*条/);
  return {
    url: location.href,
    title: document.title,
    totalRowsText: totalMatch ? Number(totalMatch[1]) : null,
    bodyPreview: body.slice(0, 5000),
    headers,
    rows,
  };
})()
"""


FUNDS_BALANCE_SCRIPT = r"""
(() => {
  const normalize = (value) => (value || '').replace(/\u00a0/g, ' ').replace(/\s+/g, ' ').trim();
  const title = normalize(document.title || '');
  const body = normalize(document.body.innerText || '');
  let balanceText = '';
  let availableBalance = null;
  const match = body.match(/可用余额(?:\s*\(CNY\))?\s*[¥￥]?\s*([0-9][0-9,]*(?:\.\d+)?)/);
  if (match) {
    balanceText = match[1];
    availableBalance = Number(match[1].replace(/,/g, ''));
  }
  return {
    url: location.href,
    title,
    body: body.slice(0, 5000),
    balanceText,
    availableBalance,
    hasWithdrawButton: body.includes('提现'),
    ready: body.includes('可用余额') && availableBalance !== null,
    loginPrompt: body.includes('授权登录') || body.includes('扫码登录') || body.includes('账号登录'),
  };
})()
"""

VIOLATION_MESSAGE_STATE_SCRIPT = r"""
(() => {
  const normalize = (value) => (value || '').replace(/\u00a0/g, ' ').replace(/\s+/g, ' ').trim();
  const visible = (el) => {
    const rect = el.getBoundingClientRect();
    const style = getComputedStyle(el);
    return rect.width > 0 && rect.height > 0 && style.visibility !== 'hidden' && style.display !== 'none';
  };
  const title = normalize(document.title || '');
  const body = normalize(document.body.innerText || '');
  const queryButton = Array.from(document.querySelectorAll('button,a,span,div'))
    .find((el) => visible(el) && normalize(el.innerText) === '查询');
  const rowCount = Array.from(document.querySelectorAll('tbody tr')).filter((row) => visible(row)).length;
  const totalMatch = body.match(/(?:共有|共)\s*(\d+)\s*条/);
  return {
    url: location.href,
    title,
    body: body.slice(0, 5000),
    rowCount,
    totalRowsText: totalMatch ? Number(totalMatch[1]) : null,
    hasNoData: body.includes('暂无数据') || body.includes('暂无结果'),
    hasQueryButton: !!queryButton,
    loading: body.includes('加载中') || body.includes('查询中'),
    ready: body.includes('违规发起时间') || body.includes('违规类型') || body.includes('违规金额'),
    loginPrompt: body.includes('授权登录') || body.includes('扫码登录') || body.includes('账号登录'),
  };
})()
"""

VIOLATION_MESSAGE_EXTRACT_SCRIPT = r"""
(() => {
  const normalize = (value) => (value || '').replace(/\u00a0/g, ' ').replace(/\s+/g, ' ').trim();
  const visible = (el) => {
    const rect = el.getBoundingClientRect();
    const style = getComputedStyle(el);
    return rect.width > 0 && rect.height > 0 && style.visibility !== 'hidden' && style.display !== 'none';
  };
  const headers = Array.from(document.querySelectorAll('thead th')).map((cell) => normalize(cell.innerText));
  const rows = Array.from(document.querySelectorAll('tbody tr'))
    .filter((row) => visible(row))
    .map((row) => {
      const cells = Array.from(row.querySelectorAll('td')).map((cell) => normalize(cell.innerText));
      const mapped = {};
      const size = Math.max(headers.length, cells.length);
      for (let i = 0; i < size; i += 1) {
        const key = headers[i] || `col${i}`;
        mapped[key] = cells[i] || '';
      }
      return {
        text: normalize(row.innerText),
        cells,
        mapped,
      };
    })
    .filter((row) => row.cells.length >= 8);
  const body = normalize(document.body.innerText || '');
  const totalMatch = body.match(/(?:共有|共)\s*(\d+)\s*条/);
  return {
    url: location.href,
    title: document.title,
    totalRowsText: totalMatch ? Number(totalMatch[1]) : null,
    bodyPreview: body.slice(0, 5000),
    headers,
    rows,
  };
})()
"""


PRICE_RULE_PAGE_STATE_SCRIPT = r"""
(() => {
  const normalize = (value) => (value || '').replace(/\u00a0/g, ' ').replace(/\s+/g, ' ').trim();
  const visible = (el) => {
    const rect = el.getBoundingClientRect();
    const style = getComputedStyle(el);
    return rect.width > 0 && rect.height > 0 && style.visibility !== 'hidden' && style.display !== 'none';
  };
  const isActiveTabNode = (el) => {
    if (!el) {
      return false;
    }
    const chain = [el, el.parentElement, el.parentElement ? el.parentElement.parentElement : null].filter(Boolean);
    return chain.some((node) => {
      const className = String(node.className || '');
      const ariaSelected = String(node.getAttribute && (node.getAttribute('aria-selected') || '')).toLowerCase();
      return (
        className.includes('TAB_active') ||
        className.includes('tab-active') ||
        className.includes('is-active') ||
        className.includes('active') ||
        ariaSelected === 'true'
      );
    });
  };
  const title = normalize(document.title || '');
  const body = normalize(document.body.innerText || '');
  const tabs = Array.from(document.querySelectorAll('div,button,span,a'))
    .filter((el) => visible(el))
    .map((el) => ({
      element: el,
      text: normalize(el.innerText || ''),
      className: String(el.className || ''),
      active: isActiveTabNode(el),
    }))
    .filter((item) => /^价格申报中\(\d+\)$/.test(item.text) || /^待卖家确认\(\d+\)$/.test(item.text) || /^成功\(\d+\)$/.test(item.text) || /^失败\(\d+\)$/.test(item.text));
  const waitingTab = tabs.find((item) => /^待卖家确认\(\d+\)$/.test(item.text)) || null;
  const activeTab = tabs.find((item) => item.active) || null;
  const waitingCountMatch = waitingTab ? waitingTab.text.match(/\((\d+)\)/) : null;
  const sizeChanger = Array.from(document.querySelectorAll('li')).find((el) => String(el.className || '').includes('PGT_sizeChanger'));
  const sizeInput = sizeChanger ? sizeChanger.querySelector('[data-testid="beast-core-select-htmlInput"], input') : null;
  const pageSize = sizeInput ? normalize(sizeInput.value || sizeInput.getAttribute('value') || '') : '';
  const batchButton = Array.from(document.querySelectorAll('button')).find((el) => visible(el) && normalize(el.innerText || '').includes('批量处理'));
  const table = Array.from(document.querySelectorAll('table')).find((el) => normalize(el.innerText || '').includes('货品信息') && normalize(el.innerText || '').includes('调整后申报价格'));
  const totalMatch = body.match(/(?:共有|共)\s*(\d+)\s*条/);
  const rowCount = Array.from(document.querySelectorAll('tbody tr')).filter((row) => visible(row)).length;
  return {
    url: location.href,
    title,
    body: body.slice(0, 5000),
    waitingCount: waitingCountMatch ? Number(waitingCountMatch[1]) : 0,
    activeTabText: activeTab ? activeTab.text : '',
    waitingTabSeen: !!waitingTab,
    priceTabsReady: tabs.length > 0,
    pageSize,
    totalRowsText: totalMatch ? Number(totalMatch[1]) : null,
    rowCount,
    hasBatchButton: !!batchButton,
    hasTable: !!table,
    hasNoData: body.includes('暂无数据'),
    loginPrompt: body.includes('授权登录') || body.includes('扫码登录') || body.includes('账号登录'),
  };
})()
"""


PRICE_RULE_CLICK_WAITING_TAB_SCRIPT = r"""
(() => {
  const normalize = (value) => (value || '').replace(/\u00a0/g, ' ').replace(/\s+/g, ' ').trim();
  const visible = (el) => {
    const rect = el.getBoundingClientRect();
    const style = getComputedStyle(el);
    return rect.width > 0 && rect.height > 0 && style.visibility !== 'hidden' && style.display !== 'none';
  };
  const clickLikeUser = (el) => {
    if (!el) {
      return false;
    }
    const rect = el.getBoundingClientRect();
    const options = { bubbles: true, cancelable: true, clientX: rect.left + rect.width / 2, clientY: rect.top + rect.height / 2 };
    ['pointerdown', 'mousedown', 'pointerup', 'mouseup', 'click'].forEach((type) => {
      try {
        el.dispatchEvent(new MouseEvent(type, options));
      } catch (_error) {}
    });
    try {
      el.click();
    } catch (_error) {}
    return true;
  };
  const candidates = Array.from(document.querySelectorAll('div,button,span,a'))
    .filter((el) => visible(el) && /^待卖家确认\(\d+\)$/.test(normalize(el.innerText || '')));
  const target = candidates
    .map((el) => el.closest('[role="tab"],button,a,div,span') || el.parentElement || el)
    .find(Boolean) || null;
  if (!target) {
    return { clicked: false };
  }
  clickLikeUser(target);
  clickLikeUser(target.parentElement);
  return { clicked: true, text: normalize(target.innerText || '') };
})()
"""


PRICE_RULE_SET_PAGE_SIZE_100_SCRIPT = r"""
(() => {
  const normalize = (value) => (value || '').replace(/\u00a0/g, ' ').replace(/\s+/g, ' ').trim();
  const visible = (el) => {
    const rect = el.getBoundingClientRect();
    const style = getComputedStyle(el);
    return rect.width > 0 && rect.height > 0 && style.visibility !== 'hidden' && style.display !== 'none';
  };
  const sizeChanger = Array.from(document.querySelectorAll('li')).find((el) => String(el.className || '').includes('PGT_sizeChanger'));
  if (!sizeChanger) {
    return { supported: false, changed: false };
  }
  const input = sizeChanger.querySelector('[data-testid="beast-core-select-htmlInput"], input');
  const current = input ? normalize(input.value || input.getAttribute('value') || '') : '';
  if (current === '100') {
    return { supported: true, changed: false, current };
  }
  const header = sizeChanger.querySelector('[data-testid="beast-core-select-header"]');
  if (!header) {
    return { supported: true, changed: false, current };
  }
  header.click();
  const option = Array.from(document.querySelectorAll('[role="option"], li, div, span'))
    .find((el) => visible(el) && normalize(el.innerText || '') === '100');
  if (!option) {
    return { supported: true, changed: false, current };
  }
  option.click();
  return { supported: true, changed: true, current };
})()
"""


PRICE_RULE_FETCH_STATE_SCRIPT = r"""
(() => window.__temu_price_rule_state || null)()
"""


PRICE_RULE_RUN_SCRIPT_TEMPLATE = r"""
(() => {
  const payload = __PRICE_RULE_PAYLOAD__;
  const stateKey = '__temu_price_rule_state';
  window[stateKey] = { stage: 'starting', message: '启动价格申报自动助手...' };

  (async () => {
    const normalize = (value) => (value || '').replace(/\u00a0/g, ' ').replace(/\s+/g, ' ').trim();
    const sleep = (ms) => new Promise((resolve) => setTimeout(resolve, ms));
    const toNum = (value) => {
      const cleaned = String(value || '').replace(/[^\d.]/g, '');
      return cleaned ? Number(cleaned) : NaN;
    };
    const getCountFromText = (value) => {
      const match = String(value || '').match(/\((\d+)\)/);
      return match ? Number(match[1]) : 0;
    };
    const visible = (el) => {
      const rect = el.getBoundingClientRect();
      const style = getComputedStyle(el);
      return rect.width > 0 && rect.height > 0 && style.visibility !== 'hidden' && style.display !== 'none';
    };
    const setState = (patch) => {
      window[stateKey] = Object.assign({}, window[stateKey] || {}, patch, { updatedAt: Date.now() });
    };

    function getPriceTabs() {
      const isActiveTabNode = (el) => {
        if (!el) {
          return false;
        }
        const chain = [el, el.parentElement, el.parentElement ? el.parentElement.parentElement : null].filter(Boolean);
        return chain.some((node) => {
          const className = String(node.className || '');
          const ariaSelected = String(node.getAttribute && (node.getAttribute('aria-selected') || '')).toLowerCase();
          return (
            className.includes('TAB_active') ||
            className.includes('tab-active') ||
            className.includes('is-active') ||
            className.includes('active') ||
            ariaSelected === 'true'
          );
        });
      };
      return Array.from(document.querySelectorAll('div,button,span,a'))
        .filter((el) => visible(el))
        .map((el) => ({
          element: el,
          text: normalize(el.innerText || ''),
          active: isActiveTabNode(el),
        }))
        .filter((item) => /^价格申报中\(\d+\)$/.test(item.text) || /^待卖家确认\(\d+\)$/.test(item.text) || /^成功\(\d+\)$/.test(item.text) || /^失败\(\d+\)$/.test(item.text));
    }

    function clickLikeUser(el) {
      if (!el) {
        return false;
      }
      const rect = el.getBoundingClientRect();
      const options = { bubbles: true, cancelable: true, clientX: rect.left + rect.width / 2, clientY: rect.top + rect.height / 2 };
      ['pointerdown', 'mousedown', 'pointerup', 'mouseup', 'click'].forEach((type) => {
        try {
          el.dispatchEvent(new MouseEvent(type, options));
        } catch (_error) {}
      });
      try {
        el.click();
      } catch (_error) {}
      return true;
    }

    function isDisabled(el) {
      if (!el) {
        return true;
      }
      const ariaDisabled = String(el.getAttribute && (el.getAttribute('aria-disabled') || '')).toLowerCase();
      const className = String(el.className || '').toLowerCase();
      return el.disabled === true || ariaDisabled === 'true' || className.includes('disabled') || className.includes('is-disabled');
    }

    async function ensureWaitingTab(timeout = 6000) {
      const start = Date.now();
      while (Date.now() - start < timeout) {
        const tabs = getPriceTabs();
        const waitingTab = tabs.find((item) => /^待卖家确认\(\d+\)$/.test(item.text));
        if (!waitingTab) {
          await sleep(200);
          continue;
        }
        if (waitingTab.active) {
          return { ok: true, text: waitingTab.text, alreadyActive: true };
        }
        const target = waitingTab.element.closest('[role="tab"],button,a,div,span') || waitingTab.element.parentElement || waitingTab.element;
        clickLikeUser(target);
        clickLikeUser(target.parentElement);
        await sleep(400);
        const afterTabs = getPriceTabs();
        const activeWaiting = afterTabs.find((item) => /^待卖家确认\(\d+\)$/.test(item.text) && item.active);
        if (activeWaiting) {
          return { ok: true, text: activeWaiting.text, alreadyActive: false };
        }
      }
      return { ok: false };
    }

    function findMainTable() {
      return Array.from(document.querySelectorAll('table')).find((table) => {
        const text = normalize(table.innerText || '');
        return text.includes('货品信息') && text.includes('调整后申报价格') && text.includes('操作');
      }) || null;
    }

    function getMainIdx(table) {
      const headers = Array.from(table.querySelectorAll('thead th, tr th')).map((th) => normalize(th.innerText));
      const idx = (keyword) => headers.findIndex((text) => text.includes(keyword));
      return { name: idx('货品信息'), price: idx('调整后申报价') };
    }

    function checkboxWrap(tr) {
      const td0 = tr.querySelector('td');
      if (!td0) {
        return null;
      }
      return td0.querySelector('[role="checkbox"], input[type="checkbox"], [aria-checked]') || td0.firstElementChild || td0;
    }

    function isChecked(wrap) {
      if (!wrap) {
        return false;
      }
      if (wrap.matches && wrap.matches('input[type="checkbox"]')) {
        return !!wrap.checked;
      }
      const input = wrap.querySelector && wrap.querySelector('input[type="checkbox"]');
      if (input) {
        return !!input.checked;
      }
      const aria = (wrap.matches && wrap.matches('[aria-checked]')) ? wrap : (wrap.querySelector && wrap.querySelector('[aria-checked]'));
      if (aria) {
        return aria.getAttribute('aria-checked') === 'true';
      }
      return false;
    }

    function setChecked(wrap, target) {
      if (!wrap) {
        return false;
      }
      if (isChecked(wrap) === target) {
        return true;
      }
      const candidates = [wrap, wrap.querySelector && wrap.querySelector('[role="checkbox"]'), wrap.querySelector && wrap.querySelector('[aria-checked]'), wrap.querySelector && wrap.querySelector('input[type="checkbox"]'), wrap.parentElement].filter(Boolean);
      for (const el of candidates) {
        try {
          el.click();
        } catch (_error) {}
        if (isChecked(wrap) === target) {
          return true;
        }
      }
      return false;
    }

    function clickBatch() {
      const button = Array.from(document.querySelectorAll('button')).find((el) => normalize(el.innerText || '').includes('批量处理'));
      if (!button || button.disabled) {
        return false;
      }
      button.click();
      return true;
    }

    async function waitModal(timeout = 8000) {
      const start = Date.now();
      while (Date.now() - start < timeout) {
        const dialog = Array.from(document.querySelectorAll('div,[role="dialog"]'))
          .find((el) => /已勾选商品共有|是否调整申报价格|确认批量处理/.test(el.innerText || ''));
        if (dialog) {
          return dialog;
        }
        await sleep(120);
      }
      return null;
    }

    function findConfirmButton(modal) {
      if (!modal) {
        return null;
      }
      const candidates = Array.from(modal.querySelectorAll('button,[role="button"],span,div,a'))
        .filter((el) => visible(el))
        .filter((el) => normalize(el.innerText || el.textContent || '') === '确认');
      for (const item of candidates) {
        const target = item.closest('button,[role="button"],a,div,span') || item;
        if (!isDisabled(target)) {
          return target;
        }
      }
      return null;
    }

    async function waitConfirmButton(modal, timeout = 20000) {
      const start = Date.now();
      while (Date.now() - start < timeout) {
        const button = findConfirmButton(modal);
        if (button) {
          return button;
        }
        await sleep(200);
      }
      return null;
    }

    async function waitModalClosed(modal, timeout = 20000) {
      const start = Date.now();
      while (Date.now() - start < timeout) {
        const exists = modal && document.contains(modal) && visible(modal);
        const anyDialog = Array.from(document.querySelectorAll('div,[role="dialog"]'))
          .some((el) => visible(el) && /已勾选商品共有|是否调整申报价格|确认批量处理/.test(el.innerText || ''));
        if (!exists && !anyDialog) {
          return true;
        }
        await sleep(200);
      }
      return false;
    }

    function getModalIdx(modal) {
      const headers = Array.from(modal.querySelectorAll('thead th, tr th')).map((th) => normalize(th.innerText));
      const idx = (...keywords) => headers.findIndex((text) => keywords.some((keyword) => text.includes(keyword)));
      return {
        currentPrice: idx('当前申报价格', '当前申报价'),
        targetPrice: idx('调整后申报价格', '调整后申报价'),
        action: idx('是否调整申报价格', '操作'),
      };
    }

    function findMatchedRule(name, price, rules) {
      return rules.find((rule) => name.includes(rule.kw) && price >= rule.min && price <= rule.max) || null;
    }

    try {
      const rules = Array.isArray(payload.rules) ? payload.rules : [];
      if (!rules.length) {
        throw new Error('没有可执行的价格规则');
      }
      await ensureWaitingTab();
      const initialWaitingTab = getPriceTabs().find((item) => /^待卖家确认\(\d+\)$/.test(item.text));
      const initialWaitingCount = initialWaitingTab ? getCountFromText(initialWaitingTab.text) : 0;
      setState({ stage: 'running', message: '第1步：筛选勾选中...' });
      let table = null;
      for (let attempt = 0; attempt < 6; attempt += 1) {
        table = findMainTable();
        if (table) {
          break;
        }
        await ensureWaitingTab(1200);
        await sleep(500);
      }
      if (!table) {
        throw new Error('没找到外层列表');
      }
      const idx = getMainIdx(table);
      if (idx.name < 0 || idx.price < 0) {
        throw new Error('价格列表列识别失败');
      }

      const rows = Array.from(table.querySelectorAll('tbody tr'));
      let selected = 0;
      const selectedRules = [];

      for (const tr of rows) {
        const tds = tr.querySelectorAll('td');
        if (tds.length <= Math.max(idx.name, idx.price)) {
          continue;
        }
        const name = normalize(tds[idx.name] ? tds[idx.name].innerText : '');
        const price = toNum(tds[idx.price] ? tds[idx.price].innerText : '');
        const wrap = checkboxWrap(tr);
        if (!wrap || !name || Number.isNaN(price)) {
          continue;
        }
        const matchedRule = findMatchedRule(name, price, rules);
        if (matchedRule) {
          if (setChecked(wrap, true)) {
            selected += 1;
            selectedRules.push({
              name,
              price,
              action: matchedRule.action,
              kw: matchedRule.kw,
              protectDiff: !!matchedRule.protectDiff,
              protectDiffLimit: matchedRule.protectDiffLimit,
            });
          }
        } else {
          setChecked(wrap, false);
        }
      }

      if (selected === 0) {
        setState({
          stage: 'done',
          matchedCount: 0,
          actedCount: 0,
          protectedCount: 0,
          unmatchedModalRows: 0,
          remainingCount: initialWaitingCount,
          selectedPreview: [],
          confirmed: false,
          confirmPending: false,
          message: initialWaitingCount > 0 ? `没有命中项，剩余${initialWaitingCount}条待人工处理` : '没有命中项',
        });
        return;
      }

      if (!clickBatch()) {
        throw new Error('批量处理按钮不可点');
      }

      setState({
        stage: 'waiting_modal',
        matchedCount: selected,
        message: `第2步：勾选${selected}条，等待弹窗...`,
        selectedPreview: selectedRules.slice(0, 20),
      });

      const modal = await waitModal();
      if (!modal) {
        throw new Error('未出现批量弹窗');
      }

      const modalIdx = getModalIdx(modal);
      const modalRows = Array.from(modal.querySelectorAll('table tbody tr'));
      let acted = 0;
      let protectedCount = 0;
      let unmatchedModalRows = 0;
      const modalPreview = [];

      for (let index = 0; index < modalRows.length; index += 1) {
        const tr = modalRows[index];
        const tds = tr.querySelectorAll('td');
        const actionIdx = modalIdx.action >= 0 ? modalIdx.action : 7;
        if (tds.length <= actionIdx) {
          continue;
        }
        const matched = selectedRules[index];
        if (!matched) {
          unmatchedModalRows += 1;
          continue;
        }

        let finalAction = matched.action;
        const currentPrice = modalIdx.currentPrice >= 0 ? toNum(tds[modalIdx.currentPrice] ? tds[modalIdx.currentPrice].innerText : '') : NaN;
        const targetPrice = modalIdx.targetPrice >= 0 ? toNum(tds[modalIdx.targetPrice] ? tds[modalIdx.targetPrice].innerText : '') : NaN;
        if (matched.protectDiff && !Number.isNaN(currentPrice) && !Number.isNaN(targetPrice) && Math.abs(currentPrice - targetPrice) > matched.protectDiffLimit) {
          finalAction = '不调整';
          protectedCount += 1;
        }

        const cell = tds[actionIdx];
        const button = Array.from(cell.querySelectorAll('span,button,div,label')).find((el) => normalize(el.textContent || '') === finalAction);
        if (button) {
          button.click();
          acted += 1;
        }
        modalPreview.push({
          name: matched.name,
          price: matched.price,
          currentPrice,
          targetPrice,
          action: matched.action,
          finalAction,
          keyword: matched.kw,
        });
      }

      const remainingCount = Math.max(initialWaitingCount - acted, 0);
      const canAutoConfirm = acted === selected && unmatchedModalRows === 0;
      if (!canAutoConfirm) {
        setState({
          stage: 'await_confirm',
          matchedCount: selected,
          actedCount: acted,
          protectedCount,
          unmatchedModalRows,
          remainingCount,
          selectedPreview: modalPreview.slice(0, 20),
          confirmed: false,
          confirmPending: true,
          message: `已命中${selected}条，但仅处理${acted}条，请人工检查后确认`,
        });
        return;
      }

      const confirmDelayMs = Math.max(5000, Math.min(20000, 3000 + acted * 450));
      setState({
        stage: 'waiting_before_confirm',
        matchedCount: selected,
        actedCount: acted,
        protectedCount,
        unmatchedModalRows,
        remainingCount,
        selectedPreview: modalPreview.slice(0, 20),
        confirmed: false,
        confirmPending: false,
        confirmDelayMs,
        message: `已处理${acted}条，等待${(confirmDelayMs / 1000).toFixed(1)}秒后自动确认...`,
      });
      await sleep(confirmDelayMs);

      const confirmButton = await waitConfirmButton(modal, 20000);
      if (!confirmButton) {
        setState({
          stage: 'confirm_failed',
          matchedCount: selected,
          actedCount: acted,
          protectedCount,
          unmatchedModalRows,
          remainingCount,
          selectedPreview: modalPreview.slice(0, 20),
          confirmed: false,
          confirmPending: true,
          message: '未找到可点击的确认按钮，请人工确认',
        });
        return;
      }

      setState({
        stage: 'confirming',
        matchedCount: selected,
        actedCount: acted,
        protectedCount,
        unmatchedModalRows,
        remainingCount,
        selectedPreview: modalPreview.slice(0, 20),
        confirmed: false,
        confirmPending: false,
        message: '正在自动点击确认...',
      });
      clickLikeUser(confirmButton);
      clickLikeUser(confirmButton.parentElement);
      const modalClosed = await waitModalClosed(modal, 20000);
      if (!modalClosed) {
        setState({
          stage: 'confirm_failed',
          matchedCount: selected,
          actedCount: acted,
          protectedCount,
          unmatchedModalRows,
          remainingCount,
          selectedPreview: modalPreview.slice(0, 20),
          confirmed: false,
          confirmPending: true,
          message: '确认后弹窗未关闭，请人工确认',
        });
        return;
      }

      setState({
        stage: 'done',
        matchedCount: selected,
        actedCount: acted,
        protectedCount,
        unmatchedModalRows,
        remainingCount,
        selectedPreview: modalPreview.slice(0, 20),
        confirmed: true,
        confirmPending: false,
        message: remainingCount > 0 ? `已自动确认，剩余${remainingCount}条待人工处理` : `已自动确认，已处理${acted}条`,
      });
    } catch (error) {
      setState({
        stage: 'error',
        error: String((error && error.message) || error),
        message: String((error && error.message) || error),
      });
    }
  })();

  return { started: true };
})()
"""


class ZclawError(RuntimeError):
    pass


class StopRequestedError(RuntimeError):
    pass


def normalize_text(value: str) -> str:
    return " ".join((value or "").split())


def state_looks_like_temu_auth_gate(state: dict[str, Any] | None) -> bool:
    if not state:
        return False
    url = normalize_text(str(state.get("url") or "")).lower()
    title = normalize_text(str(state.get("title") or ""))
    body = normalize_text(str(state.get("body") or ""))
    if bool(state.get("loginPrompt")):
        return True
    if "/auth/" in url or "/authentication" in url:
        return True
    if "商家中心" in body and ("中国地区" in body or "其他地区" in body):
        return True
    if "授权登录" in body and ("隐私政策" in body or "账号ID" in body or "店铺名称" in body):
        return True
    if "TEMU Agent Center" in title and ("商家中心" in body or "授权登录" in body):
        return True
    return False


def load_config_json() -> dict[str, Any]:
    if not CONFIG_PATH.exists():
        return {}
    try:
        data = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {}
    return data if isinstance(data, dict) else {}


def get_api_key_state() -> tuple[str, str]:
    key = os.environ.get("ZCLAW_API_KEY", "").strip()
    if key:
        return key, "环境变量"
    data = load_config_json()
    key = str(data.get("ZCLAW_API_KEY") or "").strip()
    if key:
        return key, "本地配置"
    return "", "未配置"


def load_api_key() -> str:
    key, _source = get_api_key_state()
    if key:
        return key
    raise ZclawError("未找到 ZCLAW_API_KEY，请点击“设置 API Key”按钮，或手动配置 C:\\Users\\Administrator\\.zclaw\\config.json。")


def backup_config_file(path: Path) -> Path | None:
    if not path.exists():
        return None
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    backup_path = BACKUP_DIR / f"{path.stem}_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}{path.suffix}"
    shutil.copy2(path, backup_path)
    return backup_path


def save_api_key_to_config(api_key: str) -> Path | None:
    key = api_key.strip()
    if not key:
        raise ZclawError("API Key 不能为空。")
    backup_path = backup_config_file(CONFIG_PATH)
    data = load_config_json()
    data["ZCLAW_API_KEY"] = key
    CONFIG_PATH.parent.mkdir(parents=True, exist_ok=True)
    CONFIG_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    return backup_path


def get_default_price_rule_config() -> dict[str, Any]:
    return json.loads(json.dumps(DEFAULT_PRICE_RULE_CONFIG, ensure_ascii=False))


def normalize_price_rule_config(config: dict[str, Any] | None) -> dict[str, Any]:
    raw = config or {}
    normalized_rules: list[dict[str, Any]] = []
    for item in raw.get("rules") or []:
        keyword = str(item.get("kw") or "").strip()
        if not keyword:
            continue
        min_value = item.get("min", "")
        max_value = item.get("max", "")
        normalized_rules.append(
            {
                "kw": keyword,
                "min": "" if min_value in ("", None) else float(min_value),
                "max": "" if max_value in ("", None) else float(max_value),
                "action": "不调整" if str(item.get("action") or "") == "不调整" else "调整",
            }
        )
    if not normalized_rules:
        normalized_rules = get_default_price_rule_config()["rules"]
    return {
        "protectDiff": bool(raw.get("protectDiff", True)),
        "protectDiffLimit": max(0.0, to_float(raw.get("protectDiffLimit", 1.0), 1.0)),
        "rules": normalized_rules,
    }


def is_legacy_default_price_rule_config(config: dict[str, Any]) -> bool:
    normalized = normalize_price_rule_config(config)
    return normalized.get("protectDiff") is True and abs(to_float(normalized.get("protectDiffLimit"), 1.0) - 1.0) < 1e-9 and normalized.get("rules") == LEGACY_DEFAULT_PRICE_RULE_ROWS


def load_price_rule_config() -> dict[str, Any]:
    if PRICE_RULE_CONFIG_PATH.exists():
        try:
            raw = json.loads(PRICE_RULE_CONFIG_PATH.read_text(encoding="utf-8"))
            normalized = normalize_price_rule_config(raw)
            if is_legacy_default_price_rule_config(normalized):
                upgraded = get_default_price_rule_config()
                save_price_rule_config(upgraded)
                return upgraded
            return normalized
        except Exception:
            return get_default_price_rule_config()
    return get_default_price_rule_config()


def save_price_rule_config(config: dict[str, Any]) -> None:
    normalized = normalize_price_rule_config(config)
    PRICE_RULE_CONFIG_PATH.parent.mkdir(parents=True, exist_ok=True)
    PRICE_RULE_CONFIG_PATH.write_text(json.dumps(normalized, ensure_ascii=False, indent=2), encoding="utf-8")


def normalize_store_email_accounts(raw: Any) -> dict[str, str]:
    if isinstance(raw, dict) and isinstance(raw.get("storeEmailAccounts"), dict):
        raw = raw.get("storeEmailAccounts")
    if not isinstance(raw, dict):
        return {}
    accounts: dict[str, str] = {}
    for store_id, account in raw.items():
        store_key = str(store_id or "").strip()
        email_account = str(account or "").strip()
        if store_key and email_account:
            accounts[store_key] = email_account
    return accounts


def load_store_email_accounts() -> dict[str, str]:
    if not STORE_EMAIL_CONFIG_PATH.exists():
        return {}
    try:
        raw = json.loads(STORE_EMAIL_CONFIG_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {}
    return normalize_store_email_accounts(raw)


def save_store_email_accounts(accounts: dict[str, str]) -> Path | None:
    normalized = normalize_store_email_accounts(accounts)
    existing = load_store_email_accounts()
    if existing == normalized:
        return None
    backup_path = backup_config_file(STORE_EMAIL_CONFIG_PATH)
    payload = {
        "updatedAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "storeEmailAccounts": dict(sorted(normalized.items())),
    }
    STORE_EMAIL_CONFIG_PATH.parent.mkdir(parents=True, exist_ok=True)
    STORE_EMAIL_CONFIG_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return backup_path


def parse_datetime_text(text: str) -> datetime | None:
    value = (text or "").strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    return None


def is_recent(target: datetime | None, recent_days: int, today: date | None = None) -> bool:
    if not target:
        return False
    today = today or date.today()
    start = today - timedelta(days=max(1, recent_days) - 1)
    return start <= target.date() <= today


def is_temu_store(store: dict[str, Any]) -> bool:
    return "temu" in str(store.get("platformName") or "").lower()


def is_transient_navigation_error(error: Exception) -> bool:
    message = str(error)
    return "chrome-extension://" in message or "Navigate failed: expected" in message


def to_int(value: Any, default: int = 0) -> int:
    if value is None:
        return default
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    text = str(value).strip().replace(",", "").replace("+", "")
    if not text:
        return default
    try:
        return int(text)
    except ValueError:
        return default


def to_float(value: Any, default: float = 0.0) -> float:
    if value is None:
        return default
    if isinstance(value, bool):
        return float(int(value))
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "").replace("+", "").replace("¥", "").replace("￥", "")
    if not text:
        return default
    try:
        return float(text)
    except ValueError:
        return default


def looks_like_currency_price(value: Any) -> bool:
    return bool(re.search(r"[¥￥]\s*\d+(?:\.\d+)?", str(value or "")))


def normalize_selected_checks(selected_checks: set[str] | None) -> set[str]:
    if selected_checks is None:
        return set(ALL_INSPECTION_KEYS)
    return {key for key in selected_checks if key in ALL_INSPECTION_KEYS}


def is_check_selected(selected_checks: set[str] | None, check_key: str) -> bool:
    return check_key in normalize_selected_checks(selected_checks)


def format_selected_checks(selected_checks: set[str] | None) -> str:
    enabled = normalize_selected_checks(selected_checks)
    return "、".join(label for key, label in INSPECTION_ITEMS if key in enabled)


def build_manual_problem_items(record: dict[str, Any], recent_days: int, selected_checks: set[str] | None = None) -> list[tuple[str, str]]:
    enabled = normalize_selected_checks(selected_checks)
    qc = record.get("qc") or {}
    urgent = record.get("urgent") or {}
    urgent_pending_price = record.get("urgentPendingPrice") or {}
    govern = record.get("govern") or {}
    shipping = record.get("shipping") or {}
    price_rule = record.get("priceRule") or {}
    return_order = record.get("returnOrder") or {}
    funds = record.get("funds") or {}
    violation = record.get("violation") or {}
    project_errors = record.get("projectErrors") or []

    problem_items: list[tuple[str, str]] = []
    recent_count = to_int(qc.get("recentCount"))
    ship_overdue = to_int(urgent.get("shipOverdue"))
    arrival_overdue_recent_count = to_int(urgent.get("arrivalOverdueRecentCount"))
    low_declared_price_count = to_int(urgent_pending_price.get("lowPriceCount"))
    ip_complaint_count = to_int(govern.get("ipComplaintCount"))
    tro_count = to_int(govern.get("troCount"))
    stale_waiting_count = to_int(shipping.get("staleCount"))
    price_rule_remaining_count = to_int(price_rule.get("remainingCount"))
    price_rule_confirm_pending = bool(price_rule.get("confirmPending"))
    return_order_count = to_int(return_order.get("count"))
    available_balance = to_float(funds.get("availableBalance"))
    violation_pending_count = to_int(violation.get("pendingCount"))

    if "qc" in enabled and recent_count > 0:
        problem_items.append(("qc", f"近{recent_days}天新不合格{recent_count}"))
    if "urgent" in enabled and ship_overdue > 0:
        problem_items.append(("urgent", f"发货已逾期{ship_overdue}"))
    if "urgent" in enabled and arrival_overdue_recent_count > 0:
        problem_items.append(("urgent", f"到货已逾期近{ARRIVAL_OVERDUE_RECENT_DAYS}天创建{arrival_overdue_recent_count}"))
    if "urgent_declared_price" in enabled and low_declared_price_count > 0:
        problem_items.append(("urgent_declared_price", f"待发货低申报价{low_declared_price_count}条"))
    if "govern" in enabled and ip_complaint_count > 0:
        problem_items.append(("govern", f"知识产权投诉{ip_complaint_count}"))
    if "govern" in enabled and tro_count > 0:
        problem_items.append(("govern", f"TRO{tro_count}"))
    if "shipping" in enabled and stale_waiting_count > 0:
        problem_items.append(("shipping", f"VMI超{SHIPPING_STALE_DAYS}天未收货{stale_waiting_count}"))
    if "price_rule" in enabled and price_rule_confirm_pending:
        problem_items.append(("price_rule", "价格申报自动确认失败，请人工确认"))
    elif "price_rule" in enabled and price_rule_remaining_count > 0:
        problem_items.append(("price_rule", f"价格申报剩余{price_rule_remaining_count}条待人工"))
    if "return_order" in enabled and return_order_count > 0:
        problem_items.append(("return_order", f"退货包裹{return_order_count}条"))
    if "funds" in enabled and available_balance > WITHDRAW_ALERT_THRESHOLD:
        problem_items.append(("funds", f"可用余额{available_balance:.2f}待提现"))
    if "violation" in enabled and violation_pending_count > 0:
        problem_items.append(("violation", f"违规信息待处理{violation_pending_count}条"))
    for item in project_errors:
        key = str(item.get("key") or "").strip()
        if key in enabled:
            label = str(item.get("label") or INSPECTION_LABELS.get(key) or key)
            problem_items.append((key, f"{label}检查失败"))
    return problem_items


def build_manual_reasons(record: dict[str, Any], recent_days: int, selected_checks: set[str] | None = None) -> list[str]:
    return [reason for _key, reason in build_manual_problem_items(record, recent_days, selected_checks)]


def extract_text_after_label(text: str, label: str) -> str:
    pattern = rf"{re.escape(label)}[:：]\s*(.+?)(?=\s+\S+[:：]|$)"
    match = re.search(pattern, text)
    return match.group(1).strip() if match else ""


def pick_mapping_value(mapped: dict[str, Any], *candidates: str) -> str:
    for key in candidates:
        value = str(mapped.get(key) or "").strip()
        if value:
            return value
    return ""


def parse_shipping_datetime(text: str, label: str) -> datetime | None:
    match = re.search(rf"{re.escape(label)}[:：]\s*(\d{{4}}-\d{{2}}-\d{{2}} \d{{2}}:\d{{2}}:\d{{2}})", text)
    if not match:
        return None
    return parse_datetime_text(match.group(1))


def is_older_than_days(target: datetime | None, days: int, now: datetime | None = None) -> bool:
    if not target:
        return False
    now = now or datetime.now()
    return now - target >= timedelta(days=days)


def violation_needs_manual(progress: str) -> bool:
    text = normalize_text(progress)
    return text.startswith("公示中") or "逾期未申诉" in text


def safe_sheet_name(name: str) -> str:
    for ch in ['\\', '/', '*', '?', ':', '[', ']']:
        name = name.replace(ch, "_")
    return name[:31] or "店铺"


def make_unique_sheet_name(base_name: str, used_names: set[str]) -> str:
    base_name = safe_sheet_name(base_name)
    candidate = base_name
    key = candidate.lower()
    if key not in used_names:
        used_names.add(key)
        return candidate
    counter = 2
    while True:
        suffix = f"_{counter}"
        candidate = f"{base_name[: max(1, 31 - len(suffix))]}{suffix}"
        key = candidate.lower()
        if key not in used_names:
            used_names.add(key)
            return candidate
        counter += 1


def set_box(sheet, start_row: int, start_col: int, end_row: int, end_col: int, fill: PatternFill | None = None) -> None:
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = sheet.cell(row=row, column=col)
            cell.border = BOX_BORDER
            if fill:
                cell.fill = fill


def write_section_title(sheet, row: int, title: str) -> int:
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    cell = sheet.cell(row=row, column=1, value=title)
    cell.font = Font(bold=True, size=12)
    cell.fill = SECTION_FILL
    cell.alignment = Alignment(vertical="center")
    set_box(sheet, row, 1, row, 8, SECTION_FILL)
    return row + 1


def write_table(sheet, row: int, title: str, headers: list[str], rows: list[dict[str, Any]]) -> int:
    row = write_section_title(sheet, row, title)
    if not headers:
        headers = ["说明"]
        rows = [{"说明": "暂无数据"}]
    elif not rows:
        rows = [{headers[0]: "暂无数据"}]

    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BOX_BORDER
    row += 1

    for item in rows:
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=row, column=col_idx, value=item.get(header, ""))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BOX_BORDER
            if item.get("__recent__"):
                cell.fill = ALERT_ROW_FILL
        row += 1

    return row + 1


def write_store_sheet(sheet, record: dict[str, Any], recent_days: int, selected_checks: set[str] | None = None) -> None:
    enabled = normalize_selected_checks(selected_checks)
    target = record.get("target") or {}
    qc = record.get("qc") or {}
    urgent = record.get("urgent") or {}
    urgent_pending_price = record.get("urgentPendingPrice") or {}
    govern = record.get("govern") or {}
    shipping = record.get("shipping") or {}
    price_rule = record.get("priceRule") or {}
    return_order = record.get("returnOrder") or {}
    funds = record.get("funds") or {}
    violation = record.get("violation") or {}
    project_errors = record.get("projectErrors") or []
    manual_reasons = build_manual_reasons(record, recent_days, enabled)
    need_manual = bool(manual_reasons)
    alert_count_labels = {
        f"近{recent_days}天新增条数",
        "JIT是否逾期-发货已逾期",
        f"JIT到货已逾期-近{ARRIVAL_OVERDUE_RECENT_DAYS}天创建",
        "待发货低申报价条数",
        "合规-知识产权投诉",
        "合规-TRO",
        f"VMI超{SHIPPING_STALE_DAYS}天未收货",
        "价格申报自动助手-剩余待人工",
        "退货包裹查询-记录数",
        "违规信息待处理",
        "项目检查失败",
    }

    for col, width in {
        "A": 22,
        "B": 40,
        "C": 18,
        "D": 22,
        "E": 24,
        "F": 18,
        "G": 18,
        "H": 16,
    }.items():
        sheet.column_dimensions[col].width = width

    sheet.merge_cells("A1:H1")
    title = sheet["A1"]
    title.value = f"{target.get('storeName', '')}  巡店结果"
    title.font = Font(bold=True, size=14)
    title.fill = TITLE_FILL
    title.alignment = Alignment(vertical="center")
    set_box(sheet, 1, 1, 1, 8, TITLE_FILL)

    summary_pairs = [
        ("店铺ID", target.get("storeId", "")),
        ("店铺名称", target.get("storeName", "")),
        ("平台", target.get("platformName", "")),
        ("IP", target.get("ip", "")),
        ("检查时间", record.get("checkedAt", "")),
        ("近2天是否有新不合格", "是" if "qc" in enabled and qc.get("recentCount") else ("否" if "qc" in enabled else "未检查")),
        ("抽检不合格总条数", len(qc.get("rows") or []) if "qc" in enabled else "未检查"),
        (f"近{recent_days}天新增条数", qc.get("recentCount", 0) if "qc" in enabled else "未检查"),
        ("JIT是否逾期-发货已逾期", urgent.get("shipOverdue", 0) if "urgent" in enabled else "未检查"),
        ("JIT是否逾期-到货已逾期", urgent.get("arrivalOverdue", 0) if "urgent" in enabled else "未检查"),
        (
            f"JIT到货已逾期-近{ARRIVAL_OVERDUE_RECENT_DAYS}天创建",
            urgent.get("arrivalOverdueRecentCount", 0) if "urgent" in enabled else "未检查",
        ),
        ("待发货低申报价条数", urgent_pending_price.get("lowPriceCount", 0) if "urgent_declared_price" in enabled else "未检查"),
        ("合规-知识产权投诉", govern.get("ipComplaintCount", 0) if "govern" in enabled else "未检查"),
        ("合规-TRO", govern.get("troCount", 0) if "govern" in enabled else "未检查"),
        (f"VMI超{SHIPPING_STALE_DAYS}天未收货", shipping.get("staleCount", 0) if "shipping" in enabled else "未检查"),
        ("价格申报自动助手-剩余待人工", price_rule.get("remainingCount", 0) if "price_rule" in enabled else "未检查"),
        ("退货包裹查询-记录数", return_order.get("count", 0) if "return_order" in enabled else "未检查"),
        ("违规信息待处理", violation.get("pendingCount", 0) if "violation" in enabled else "未检查"),
        ("资金中心-可用余额(CNY)", funds.get("availableBalance", 0) if "funds" in enabled else "未检查"),
        ("项目检查失败", len(project_errors)),
        ("是否需人工处理", "是" if need_manual else "否"),
        ("需要人工处理项", "；".join(manual_reasons) if manual_reasons else "无"),
    ]

    row = 3
    for idx, (label, value) in enumerate(summary_pairs):
        block_row = row + idx // 2
        block_col = 1 if idx % 2 == 0 else 5
        label_cell = sheet.cell(row=block_row, column=block_col, value=label)
        value_cell = sheet.cell(row=block_row, column=block_col + 1, value=value)
        set_box(sheet, block_row, block_col, block_row, block_col + 1, CARD_FILL)
        label_cell.font = Font(bold=True)
        label_cell.fill = CARD_FILL
        label_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.alignment = Alignment(vertical="center", wrap_text=True)
        if (label == "近2天是否有新不合格" and value == "是") or (label == "是否需人工处理" and value == "是"):
            value_cell.fill = ALERT_STRONG_FILL
            value_cell.font = Font(bold=True, color="FFFFFF")
        elif label == "资金中心-可用余额(CNY)" and to_float(value) > WITHDRAW_ALERT_THRESHOLD:
            value_cell.fill = ALERT_STRONG_FILL
            value_cell.font = Font(bold=True, color="FFFFFF")
        elif label in alert_count_labels and to_int(value) > 0:
            value_cell.fill = ALERT_STRONG_FILL
            value_cell.font = Font(bold=True, color="FFFFFF")
        elif label == "需要人工处理项" and value != "无":
            value_cell.fill = ALERT_ROW_FILL
            value_cell.font = Font(bold=True)
        else:
            value_cell.fill = PatternFill("solid", fgColor="FFFFFF")
        value_cell.border = BOX_BORDER

    row = row + (len(summary_pairs) + 1) // 2 + 1
    if project_errors:
        row = write_table(
            sheet,
            row,
            "项目检查失败",
            ["项目", "错误"],
            [
                {
                    "项目": str(item.get("label") or INSPECTION_LABELS.get(str(item.get("key") or "")) or item.get("key") or ""),
                    "错误": str(item.get("message") or ""),
                    "__recent__": True,
                }
                for item in project_errors
            ],
        )

    headers = ["最新抽检时间", "备货单号", "商品信息", "SKU信息", "操作", "是否近2天新增"]
    rows = []
    for item in qc.get("rows") or []:
        rows.append(
            {
                "最新抽检时间": item.get("latestQcTime", ""),
                "备货单号": item.get("prepareOrderNo", ""),
                "商品信息": item.get("productInfo", ""),
                "SKU信息": item.get("skuInfo", ""),
                "操作": item.get("operation", ""),
                "是否近2天新增": "是" if item.get("isRecent") else "否",
                "__recent__": item.get("isRecent"),
            }
        )
    if "qc" in enabled:
        row = write_table(sheet, row, "抽检不合格", headers, rows)
    else:
        row = write_table(sheet, row, "抽检不合格", ["说明"], [{"说明": "本次未勾选该项目"}])

    urgent_rows = [
        {
            "指标": "发货已逾期",
            "数值": to_int(urgent.get("shipOverdue")),
            "是否正常": "是" if to_int(urgent.get("shipOverdue")) == 0 else "否",
            "__recent__": to_int(urgent.get("shipOverdue")) > 0,
        },
        {
            "指标": "到货已逾期",
            "数值": to_int(urgent.get("arrivalOverdue")),
            "是否正常": "是" if to_int(urgent.get("arrivalOverdueRecentCount")) == 0 else "否",
            "__recent__": False,
        },
        {
            "指标": f"到货已逾期近{ARRIVAL_OVERDUE_RECENT_DAYS}天创建",
            "数值": to_int(urgent.get("arrivalOverdueRecentCount")),
            "是否正常": "是" if to_int(urgent.get("arrivalOverdueRecentCount")) == 0 else "否",
            "__recent__": to_int(urgent.get("arrivalOverdueRecentCount")) > 0,
        },
    ]
    if "urgent" in enabled:
        row = write_table(sheet, row, "检查JIT是否逾期", ["指标", "数值", "是否正常"], urgent_rows)
    else:
        row = write_table(sheet, row, "检查JIT是否逾期", ["说明"], [{"说明": "本次未勾选该项目"}])

    urgent_pending_rows = []
    for item in urgent_pending_price.get("rows") or []:
        urgent_pending_rows.append(
            {
                "备货单号": item.get("prepareOrderNo", ""),
                "商品信息": item.get("productInfo", ""),
                "SKU信息": item.get("skuInfo", ""),
                "状态": item.get("status", ""),
                "申报价格(CNY)": item.get("declaredPrice", ""),
                "备货单创建时间": item.get("createdTime", ""),
                "是否低于10": "是" if item.get("isLowPrice") else "否",
                "备注": item.get("remark", ""),
                "__recent__": item.get("isLowPrice"),
            }
        )
    if "urgent_declared_price" in enabled:
        if urgent_pending_price.get("error"):
            urgent_pending_rows = [
                {
                    "备货单号": "检查失败",
                    "商品信息": str(urgent_pending_price.get("error") or ""),
                    "备注": "该子项目失败，已继续执行后续巡查项目",
                    "__recent__": True,
                }
            ]
        row = write_table(
            sheet,
            row,
            "检查待发货低申报价（紧急备货建议-待发货）",
            ["备货单号", "商品信息", "SKU信息", "状态", "申报价格(CNY)", "备货单创建时间", "是否低于10", "备注"],
            urgent_pending_rows,
        )
    else:
        row = write_table(sheet, row, "检查待发货低申报价", ["说明"], [{"说明": "本次未勾选该项目"}])

    urgent_arrival_rows = []
    for item in urgent.get("arrivalOverdueRows") or []:
        urgent_arrival_rows.append(
            {
                "备货单号": item.get("prepareOrderNo", ""),
                "商品信息": item.get("productInfo", ""),
                "状态": item.get("status", ""),
                "SKU信息": item.get("skuInfo", ""),
                "申报价格(CNY)": item.get("declaredPrice", ""),
                "送货/入库数": item.get("deliveryInfo", ""),
                "备货单创建时间": item.get("createdTime", ""),
                f"是否近{ARRIVAL_OVERDUE_RECENT_DAYS}天创建": "是" if item.get("isRecent") else "否",
                "__recent__": item.get("isRecent"),
            }
        )
    if "urgent" in enabled:
        row = write_table(
            sheet,
            row,
            f"到货已逾期明细（近{ARRIVAL_OVERDUE_RECENT_DAYS}天创建才提醒）",
            ["备货单号", "商品信息", "状态", "SKU信息", "申报价格(CNY)", "送货/入库数", "备货单创建时间", f"是否近{ARRIVAL_OVERDUE_RECENT_DAYS}天创建"],
            urgent_arrival_rows,
        )
    else:
        row = write_table(sheet, row, "到货已逾期明细", ["说明"], [{"说明": "本次未勾选该项目"}])

    govern_rows = [
        {
            "指标": "知识产权投诉",
            "数值": to_int(govern.get("ipComplaintCount")),
            "是否正常": "是" if to_int(govern.get("ipComplaintCount")) == 0 else "否",
            "__recent__": to_int(govern.get("ipComplaintCount")) > 0,
        },
        {
            "指标": "临时限制令（TRO）",
            "数值": to_int(govern.get("troCount")),
            "是否正常": "是" if to_int(govern.get("troCount")) == 0 else "否",
            "__recent__": to_int(govern.get("troCount")) > 0,
        },
    ]
    if "govern" in enabled:
        row = write_table(sheet, row, "合规中心", ["指标", "数值", "是否正常"], govern_rows)
    else:
        row = write_table(sheet, row, "合规中心", ["说明"], [{"说明": "本次未勾选该项目"}])

    shipping_rows = []
    for item in shipping.get("rows") or []:
        shipping_rows.append(
            {
                "发货单号": item.get("shippingOrderNo", ""),
                "备货单号": item.get("prepareOrderNo", ""),
                "发货时间": item.get("shipTimeText", ""),
                "收货时间": item.get("receiveTimeText", ""),
                "状态": item.get("status", ""),
                "商品信息": item.get("productInfo", ""),
                "物流信息": item.get("shippingInfo", item.get("logisticsInfo", "")),
                f"是否超{SHIPPING_STALE_DAYS}天": "是" if item.get("isStale") else "否",
                "__recent__": item.get("isStale"),
            }
        )
    if "shipping" in enabled:
        row = write_table(
            sheet,
            row,
            f"检查VMI超{SHIPPING_STALE_DAYS}天未收货（发货单列表-待仓库收货，是否JIT=否）",
            ["发货单号", "备货单号", "发货时间", "收货时间", "状态", "商品信息", "物流信息", f"是否超{SHIPPING_STALE_DAYS}天"],
            shipping_rows,
        )
    else:
        row = write_table(sheet, row, f"检查VMI超{SHIPPING_STALE_DAYS}天未收货", ["说明"], [{"说明": "本次未勾选该项目"}])

    price_rule_summary_rows = [
        {"指标": "待卖家确认总数", "数值": price_rule.get("waitingCount", 0), "__recent__": False},
        {"指标": "规则命中数", "数值": price_rule.get("matchedCount", 0), "__recent__": False},
        {"指标": "弹窗执行数", "数值": price_rule.get("actedCount", 0), "__recent__": False},
        {"指标": "价差保护数", "数值": price_rule.get("protectedCount", 0), "__recent__": False},
        {"指标": "剩余待人工数", "数值": price_rule.get("remainingCount", 0), "__recent__": to_int(price_rule.get("remainingCount")) > 0},
        {"指标": "自动确认结果", "数值": "待人工确认" if price_rule.get("confirmPending") else ("已自动确认" if price_rule.get("confirmed") else "无需确认"), "__recent__": bool(price_rule.get("confirmPending"))},
        {"指标": "状态", "数值": price_rule.get("stage", ""), "__recent__": False},
    ]
    if "price_rule" in enabled:
        row = write_table(sheet, row, "价格申报自动助手", ["指标", "数值"], price_rule_summary_rows)
        preview_rows = []
        for item in price_rule.get("selectedPreview") or []:
            preview_rows.append(
                {
                    "货品信息": item.get("name", ""),
                    "当前价": item.get("currentPrice", item.get("price", "")),
                    "调整后价": item.get("targetPrice", ""),
                    "规则动作": item.get("action", ""),
                    "最终动作": item.get("finalAction", ""),
                    "__recent__": True,
                }
            )
        row = write_table(sheet, row, "价格申报命中预览", ["货品信息", "当前价", "调整后价", "规则动作", "最终动作"], preview_rows)
    else:
        row = write_table(sheet, row, "价格申报自动助手", ["说明"], [{"说明": "本次未勾选该项目"}])

    return_rows = []
    for item in return_order.get("rows") or []:
        return_rows.append(
            {
                "退货包裹号": item.get("returnPackageNo", ""),
                "快递单号": item.get("trackingNo", ""),
                "状态": item.get("status", ""),
                "物流商": item.get("carrier", ""),
                "打包完成时间": item.get("packCompleteTime", ""),
                "出库时间": item.get("outboundTime", ""),
                "__recent__": True,
            }
        )
    if "return_order" in enabled:
        row = write_table(
            sheet,
            row,
            "退货包裹查询",
            ["退货包裹号", "快递单号", "状态", "物流商", "打包完成时间", "出库时间"],
            return_rows,
        )
    else:
        row = write_table(sheet, row, "退货包裹查询", ["说明"], [{"说明": "本次未勾选该项目"}])

    violation_rows = []
    for item in violation.get("rows") or []:
        violation_rows.append(
            {
                "违规编号": item.get("violationNo", ""),
                "备货单": item.get("prepareOrderNo", ""),
                "备货单类型": item.get("prepareOrderType", ""),
                "违规类型": item.get("violationType", ""),
                "违规发起时间": item.get("violationTime", ""),
                "违规金额(CNY)": item.get("amount", ""),
                "减免后违规金额": item.get("reducedAmount", ""),
                "进度": item.get("progress", ""),
                "操作": item.get("action", ""),
                "是否需人工": "是" if item.get("needsManual") else "否",
                "__recent__": item.get("needsManual"),
            }
        )
    if "violation" in enabled:
        row = write_table(
            sheet,
            row,
            "违规信息（进度命中才提醒）",
            ["违规编号", "备货单", "备货单类型", "违规类型", "违规发起时间", "违规金额(CNY)", "减免后违规金额", "进度", "操作", "是否需人工"],
            violation_rows,
        )
    else:
        row = write_table(sheet, row, "违规信息", ["说明"], [{"说明": "本次未勾选该项目"}])

    funds_rows = [
        {
            "指标": "可用余额(CNY)",
            "数值": funds.get("availableBalance", 0),
            "是否需提现": "是" if to_float(funds.get("availableBalance")) > WITHDRAW_ALERT_THRESHOLD else "否",
            "__recent__": to_float(funds.get("availableBalance")) > WITHDRAW_ALERT_THRESHOLD,
        }
    ]
    if "funds" in enabled:
        row = write_table(sheet, row, "资金中心", ["指标", "数值", "是否需提现"], funds_rows)
    else:
        row = write_table(sheet, row, "资金中心", ["说明"], [{"说明": "本次未勾选该项目"}])
    sheet.freeze_panes = None


def generate_report(records: list[dict[str, Any]], output_path: Path, recent_days: int, selected_checks: set[str] | None = None) -> Path:
    enabled = normalize_selected_checks(selected_checks)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "总览"

    used_sheet_names: set[str] = set()
    sheet_name_map: dict[int, str] = {}
    for record in records:
        target = record.get("target") or {}
        sheet_name_map[id(record)] = make_unique_sheet_name(
            f"{target.get('storeName', '')}_{target.get('storeId', '')}",
            used_sheet_names,
        )

    for col, width in {
        "A": 22,
        "B": 26,
        "C": 4,
        "D": 22,
        "E": 34,
    }.items():
        summary.column_dimensions[col].width = width

    summary.merge_cells("A1:E1")
    cell = summary["A1"]
    cell.value = "紫鸟 TEMU 巡店结果总览"
    cell.font = Font(bold=True, size=14)
    cell.fill = TITLE_FILL
    cell.alignment = Alignment(vertical="center")
    set_box(summary, 1, 1, 1, 5, TITLE_FILL)

    today = date.today()
    summary.merge_cells("A2:E2")
    note_cell = summary["A2"]
    note_cell.value = f"近{recent_days}天按自然日计算：{(today - timedelta(days=recent_days - 1)).isoformat()} 至 {today.isoformat()}；本次巡查项目：{format_selected_checks(enabled)}"
    note_cell.alignment = Alignment(vertical="center", wrap_text=True)
    set_box(summary, 2, 1, 2, 5, SECTION_FILL)

    def write_summary_label(row_idx: int, col_idx: int, text: str) -> None:
        label_cell = summary.cell(row=row_idx, column=col_idx, value=text)
        label_cell.font = Font(bold=True)
        label_cell.fill = HEADER_FILL
        label_cell.alignment = Alignment(vertical="center", wrap_text=True)
        label_cell.border = BOX_BORDER

    def write_summary_value(
        row_idx: int,
        col_idx: int,
        value: Any,
        *,
        alert: bool = False,
        ok: bool = False,
        hyperlink: str | None = None,
    ) -> None:
        value_cell = summary.cell(row=row_idx, column=col_idx, value=value)
        value_cell.alignment = Alignment(vertical="center", wrap_text=True)
        value_cell.border = BOX_BORDER
        if alert:
            value_cell.fill = ALERT_STRONG_FILL
            value_cell.font = Font(bold=True, color="FFFFFF")
        elif ok:
            value_cell.fill = OK_FILL
            value_cell.font = Font(bold=True)
        if hyperlink:
            value_cell.hyperlink = hyperlink
            if not alert:
                value_cell.style = "Hyperlink"

    def write_summary_pair_row(
        row_idx: int,
        left_label: str,
        left_value: Any,
        *,
        left_alert: bool = False,
        left_ok: bool = False,
        left_hyperlink: str | None = None,
        right_label: str = "",
        right_value: Any = "",
        right_alert: bool = False,
        right_ok: bool = False,
        right_hyperlink: str | None = None,
    ) -> int:
        write_summary_label(row_idx, 1, left_label)
        write_summary_value(row_idx, 2, left_value, alert=left_alert, ok=left_ok, hyperlink=left_hyperlink)
        summary.cell(row=row_idx, column=3).border = BOX_BORDER
        if right_label:
            write_summary_label(row_idx, 4, right_label)
            write_summary_value(row_idx, 5, right_value, alert=right_alert, ok=right_ok, hyperlink=right_hyperlink)
        else:
            write_summary_label(row_idx, 4, "")
            write_summary_value(row_idx, 5, right_value, alert=right_alert, ok=right_ok, hyperlink=right_hyperlink)
        return row_idx + 1

    row_no = 4
    for record in records:
        target = record.get("target") or {}
        qc = record.get("qc") or {}
        urgent = record.get("urgent") or {}
        urgent_pending_price = record.get("urgentPendingPrice") or {}
        govern = record.get("govern") or {}
        shipping = record.get("shipping") or {}
        price_rule = record.get("priceRule") or {}
        return_order = record.get("returnOrder") or {}
        violation = record.get("violation") or {}
        funds = record.get("funds") or {}
        project_errors = record.get("projectErrors") or []
        rows = qc.get("rows") or []
        manual_reasons = build_manual_reasons(record, recent_days, enabled)
        need_manual = bool(manual_reasons)
        latest_time = max((item.get("latestQcTime") or "" for item in rows), default="") if "qc" in enabled else "未检查"
        title_fill = ALERT_STRONG_FILL if need_manual else OK_FILL
        title_font = Font(bold=True, size=12, color="FFFFFF" if need_manual else "000000")
        summary.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=5)
        title_cell = summary.cell(
            row=row_no,
            column=1,
            value=f"{target.get('storeName', '')}  |  {'需人工处理' if need_manual else '正常'}",
        )
        title_cell.fill = title_fill
        title_cell.font = title_font
        title_cell.alignment = Alignment(vertical="center")
        set_box(summary, row_no, 1, row_no, 5, title_fill)
        row_no += 1

        row_no = write_summary_pair_row(
            row_no,
            "店铺ID",
            target.get("storeId", ""),
            right_label="平台",
            right_value=target.get("platformName", ""),
        )
        row_no = write_summary_pair_row(
            row_no,
            "抽检不合格总数",
            len(rows) if "qc" in enabled else "未检查",
            right_label=f"近{recent_days}天新增",
            right_value=qc.get("recentCount", 0) if "qc" in enabled else "未检查",
            right_alert=("qc" in enabled and to_int(qc.get("recentCount", 0)) > 0),
        )
        row_no = write_summary_pair_row(
            row_no,
            "JIT发货已逾期",
            urgent.get("shipOverdue", 0) if "urgent" in enabled else "未检查",
            left_alert=("urgent" in enabled and to_int(urgent.get("shipOverdue", 0)) > 0),
            right_label="JIT到货已逾期",
            right_value=urgent.get("arrivalOverdue", 0) if "urgent" in enabled else "未检查",
        )
        row_no = write_summary_pair_row(
            row_no,
            f"JIT到货已逾期近{ARRIVAL_OVERDUE_RECENT_DAYS}天创建",
            urgent.get("arrivalOverdueRecentCount", 0) if "urgent" in enabled else "未检查",
            left_alert=("urgent" in enabled and to_int(urgent.get("arrivalOverdueRecentCount", 0)) > 0),
            right_label="待发货低申报价",
            right_value=urgent_pending_price.get("lowPriceCount", 0) if "urgent_declared_price" in enabled else "未检查",
            right_alert=("urgent_declared_price" in enabled and to_int(urgent_pending_price.get("lowPriceCount", 0)) > 0),
        )
        row_no = write_summary_pair_row(
            row_no,
            "知识产权投诉",
            govern.get("ipComplaintCount", 0) if "govern" in enabled else "未检查",
            left_alert=("govern" in enabled and to_int(govern.get("ipComplaintCount", 0)) > 0),
            right_label="TRO",
            right_value=govern.get("troCount", 0) if "govern" in enabled else "未检查",
            right_alert=("govern" in enabled and to_int(govern.get("troCount", 0)) > 0),
        )
        row_no = write_summary_pair_row(
            row_no,
            f"VMI超{SHIPPING_STALE_DAYS}天未收货",
            shipping.get("staleCount", 0) if "shipping" in enabled else "未检查",
            left_alert=("shipping" in enabled and to_int(shipping.get("staleCount", 0)) > 0),
            right_label="价格申报剩余待人工",
            right_value=price_rule.get("remainingCount", 0) if "price_rule" in enabled else "未检查",
            right_alert=("price_rule" in enabled and to_int(price_rule.get("remainingCount", 0)) > 0),
        )
        row_no = write_summary_pair_row(
            row_no,
            "退货包裹记录数",
            return_order.get("count", 0) if "return_order" in enabled else "未检查",
            left_alert=("return_order" in enabled and to_int(return_order.get("count", 0)) > 0),
            right_label="违规信息待处理",
            right_value=violation.get("pendingCount", 0) if "violation" in enabled else "未检查",
            right_alert=("violation" in enabled and to_int(violation.get("pendingCount", 0)) > 0),
        )
        row_no = write_summary_pair_row(
            row_no,
            "可用余额(CNY)",
            funds.get("availableBalance", 0) if "funds" in enabled else "未检查",
            left_alert=("funds" in enabled and to_float(funds.get("availableBalance", 0)) > WITHDRAW_ALERT_THRESHOLD),
            right_label="最新抽检时间",
            right_value=latest_time,
        )
        row_no = write_summary_pair_row(
            row_no,
            "项目检查失败",
            len(project_errors),
            left_alert=bool(project_errors),
            right_label="工作表",
            right_value=sheet_name_map[id(record)],
            right_hyperlink=f"#'{sheet_name_map[id(record)]}'!A1",
        )
        row_no = write_summary_pair_row(
            row_no,
            "是否需人工处理",
            "是" if need_manual else "否",
            left_alert=need_manual,
            left_ok=not need_manual,
        )

        write_summary_label(row_no, 1, "处理项")
        summary.merge_cells(start_row=row_no, start_column=2, end_row=row_no, end_column=5)
        reason_cell = summary.cell(row=row_no, column=2, value="；".join(manual_reasons) if manual_reasons else "无")
        reason_cell.alignment = Alignment(vertical="center", wrap_text=True)
        reason_cell.border = BOX_BORDER
        set_box(summary, row_no, 2, row_no, 5, ALERT_ROW_FILL if need_manual else OK_FILL)
        if need_manual:
            reason_cell.font = Font(bold=True)
        row_no += 2

    summary.freeze_panes = None

    for record in records:
        sheet = workbook.create_sheet(title=sheet_name_map[id(record)])
        write_store_sheet(sheet, record, recent_days, enabled)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


class ZclawClient:
    def __init__(self, base_url: str, api_key: str) -> None:
        self.base_url = base_url.rstrip("/")
        self.api_key = api_key

    def _request(self, method: str, path: str, payload: dict[str, Any] | None = None, timeout: float = 60.0) -> Any:
        data = None if payload is None else json.dumps(payload).encode("utf-8")
        headers = {"Content-Type": "application/json; charset=utf-8"}
        if method.upper() == "POST":
            headers["X-ZClaw-Api-Key"] = self.api_key
        request = urllib.request.Request(f"{self.base_url}{path}", data=data, headers=headers, method=method.upper())
        try:
            with urllib.request.urlopen(request, timeout=timeout) as response:
                return json.loads(response.read().decode("utf-8"))
        except urllib.error.HTTPError as error:
            body = error.read().decode("utf-8", errors="ignore")
            raise ZclawError(f"请求失败 {path} -> HTTP {error.code}: {body or error.reason}") from error
        except urllib.error.URLError as error:
            raise ZclawError(f"无法连接 ZClaw：{error.reason}") from error

    def get_tools(self) -> list[dict[str, Any]]:
        response = self._request("GET", "/zclaw/tools", timeout=15.0)
        if response.get("ret") != 0:
            raise ZclawError(f"读取工具列表失败：{response}")
        data = response.get("data")
        if not isinstance(data, list):
            raise ZclawError(f"工具列表格式异常：{response}")
        return data

    def invoke(self, tool: str, args: dict[str, Any], timeout: float = 60.0) -> Any:
        response = self._request("POST", "/zclaw/tools/invoke", {"tool": tool, "args": args}, timeout=timeout)
        if response.get("ret") != 0:
            raise ZclawError(response.get("msg") or f"{tool} 调用失败：{response}")
        data = response.get("data") or {}
        if data.get("ok") is False:
            raise ZclawError(f"{tool} 返回失败：{data}")
        return data.get("data", data)


class ZiniaoTemuInspectorApp:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title("紫鸟 TEMU 巡店工具 v1.52")
        self.root.geometry("1180x760")

        self.base_url = BASE_URL_DEFAULT
        self.api_key = ""
        self.client: ZclawClient | None = None
        self.stores: list[dict[str, Any]] = []
        self.filtered_stores: list[dict[str, Any]] = []
        self.check_vars: dict[str, tk.BooleanVar] = {}
        self.inspection_vars: dict[str, tk.BooleanVar] = {
            key: tk.BooleanVar(value=True) for key, _label in INSPECTION_ITEMS
        }
        self.price_rule_config = load_price_rule_config()
        self.price_rule_editor: tk.Toplevel | None = None
        self.price_rule_rows_container: ttk.Frame | None = None
        self.price_rule_editor_rows: list[dict[str, Any]] = []
        self.price_rule_protect_var = tk.BooleanVar(value=bool(self.price_rule_config.get("protectDiff", True)))
        self.price_rule_limit_var = tk.StringVar(value=str(self.price_rule_config.get("protectDiffLimit", 1.0)))
        self.inspection_checkbuttons: list[ttk.Checkbutton] = []
        self.store_status_vars: dict[str, tk.StringVar] = {}
        self.store_status_colors: dict[str, str] = {}
        self.store_status_labels: dict[str, tk.Label] = {}
        self.store_email_accounts: dict[str, str] = load_store_email_accounts()
        self.store_email_vars: dict[str, tk.StringVar] = {}
        self.store_email_entries: dict[str, ttk.Entry] = {}
        self.active_email_accounts: dict[str, str] = {}
        self.store_alert_flags: dict[str, bool] = {}
        self.result_item_ids: dict[str, str] = {}
        self.result_detail_map: dict[str, str] = {}
        self.latest_report: Path | None = None
        self.report_dir = DESKTOP_DIR
        self.running = False
        self.stop_requested = False
        self.batch_thread: threading.Thread | None = None
        self.recent_days = RECENT_DAYS_DEFAULT

        self.status_var = tk.StringVar(value="正在连接 ZClaw...")
        self.search_var = tk.StringVar()
        self.selected_count_var = tk.StringVar(value="已勾选 0 家")
        self.alert_count_var = tk.StringVar(value="需处理 0 家")
        self.latest_report_var = tk.StringVar(value="暂无报表")
        self.report_dir_var = tk.StringVar(value=str(self.report_dir))
        self.api_key_status_var = tk.StringVar(value="API Key：未检测")
        self.post_check_dock_var = tk.StringVar(value=f"{POST_CHECK_DOCK_LABEL}（核价脚本入口）")
        self.panel_only_var = tk.BooleanVar(value=True)
        self.low_priority_expanded = True
        self.progress_var = tk.DoubleVar(value=0.0)
        self.progress_text_var = tk.StringVar(value="未开始")

        self._build_ui()
        self.refresh_api_key_status()
        self.root.after(200, self.refresh_async)

    def _build_ui(self) -> None:
        top = ttk.Frame(self.root, padding=10)
        top.pack(fill="x")
        ttk.Label(top, text="ZClaw 状态：").pack(side="left")
        ttk.Label(top, textvariable=self.status_var, foreground="#0a5").pack(side="left", padx=(0, 12))
        ttk.Label(top, textvariable=self.api_key_status_var).pack(side="left", padx=(0, 12))
        self.api_key_button = ttk.Button(top, text="设置 API Key", command=self.prompt_set_api_key)
        self.api_key_button.pack(side="left")
        self.refresh_button = ttk.Button(top, text="刷新连接", command=self.refresh_async)
        self.refresh_button.pack(side="left", padx=(8, 0))
        self.refresh_store_button = ttk.Button(top, text="刷新店铺", command=self.refresh_async)
        self.refresh_store_button.pack(side="left", padx=(8, 0))
        self.open_report_button = ttk.Button(top, text="打开最新报表", command=self.open_latest_report)
        self.open_report_button.pack(side="left", padx=(8, 0))

        api_help_bar = ttk.Frame(self.root, padding=(10, 0, 10, 8))
        api_help_bar.pack(fill="x")
        self.api_help_label = ttk.Label(
            api_help_bar,
            text=(
                "使用说明：需先安装紫鸟公测版客户端：https://www.ziniao.com/download  "
                "API Key 获取：https://open.ziniao.com/ziniaoAssistant"
            ),
            foreground="#666666",
            wraplength=1120,
            justify="left",
        )
        self.api_help_label.pack(side="left", anchor="w")

        filter_bar = ttk.Frame(self.root, padding=(10, 0, 10, 10))
        filter_bar.pack(fill="x")
        ttk.Label(filter_bar, text="搜索店铺：").pack(side="left")
        self.search_entry = ttk.Entry(filter_bar, textvariable=self.search_var, width=36)
        self.search_entry.pack(side="left", padx=(4, 12))
        self.search_entry.bind("<KeyRelease>", lambda _e: self.rebuild_store_list())
        self.select_all_button = ttk.Button(filter_bar, text="勾选全部 TEMU 店铺", command=self.select_all_visible)
        self.select_all_button.pack(side="left")
        self.clear_button = ttk.Button(filter_bar, text="清空勾选", command=self.clear_selection)
        self.clear_button.pack(side="left", padx=(8, 0))
        ttk.Label(filter_bar, textvariable=self.selected_count_var).pack(side="left", padx=(12, 0))
        self.selected_button = ttk.Button(filter_bar, text="查询勾选店铺", command=self.run_selected_async)
        self.selected_button.pack(side="left", padx=(20, 0))
        self.all_button = ttk.Button(filter_bar, text="一键巡全部 TEMU 店铺", command=self.run_all_async)
        self.all_button.pack(side="left", padx=(8, 0))
        self.stop_button = ttk.Button(filter_bar, text="停止查询", command=self.stop_batch, state="disabled")
        self.stop_button.pack(side="left", padx=(8, 0))

        inspection_bar = ttk.LabelFrame(self.root, text="巡查项目", padding=(10, 6))
        inspection_bar.pack(fill="x", padx=10, pady=(0, 10))

        primary_group = ttk.Frame(inspection_bar)
        primary_group.pack(fill="x", anchor="w")
        ttk.Label(primary_group, text="常规项目：").pack(side="left")
        for key, label in INSPECTION_ITEMS:
            if key not in HIGH_PRIORITY_INSPECTION_KEYS:
                continue
            button = ttk.Checkbutton(primary_group, text=label, variable=self.inspection_vars[key])
            button.pack(side="left", padx=(10, 0))
            self.inspection_checkbuttons.append(button)
        self.price_rule_button = ttk.Button(primary_group, text="价格规则设置", command=self.open_price_rule_editor)
        self.price_rule_button.pack(side="left", padx=(16, 0))
        self.select_all_checks_button = ttk.Button(primary_group, text="全选项目", command=self.select_all_inspections)
        self.select_all_checks_button.pack(side="left", padx=(16, 0))
        self.clear_checks_button = ttk.Button(primary_group, text="清空项目", command=self.clear_inspections)
        self.clear_checks_button.pack(side="left", padx=(8, 0))
        ttk.Label(primary_group, text="低优先级：").pack(side="left", padx=(16, 4))
        self.low_priority_toggle_button = ttk.Button(primary_group, text="▸", width=3, command=self.toggle_low_priority_visibility)
        self.low_priority_toggle_button.pack(side="left")

        self.low_priority_group_frame = ttk.LabelFrame(inspection_bar, text="低优先级项目", padding=(8, 6))
        for key, label in INSPECTION_ITEMS:
            if key not in LOW_PRIORITY_INSPECTION_KEYS:
                continue
            button = ttk.Checkbutton(
                self.low_priority_group_frame,
                text=label,
                variable=self.inspection_vars[key],
            )
            button.pack(side="left", padx=(10, 0))
            self.inspection_checkbuttons.append(button)
        self.update_low_priority_visibility()

        report_bar = ttk.Frame(self.root, padding=(10, 0, 10, 10))
        report_bar.pack(fill="x")
        self.panel_only_checkbutton = ttk.Checkbutton(
            report_bar,
            text="仅在面板显示（不生成报表）",
            variable=self.panel_only_var,
            command=self.on_output_mode_changed,
        )
        self.panel_only_checkbutton.pack(side="left", padx=(0, 12))
        ttk.Label(report_bar, text="报表保存位置：").pack(side="left")
        ttk.Label(report_bar, textvariable=self.report_dir_var).pack(side="left", fill="x", expand=True, padx=(4, 8))
        self.report_dir_button = ttk.Button(report_bar, text="选择位置", command=self.choose_report_dir)
        self.report_dir_button.pack(side="left")
        ttk.Label(report_bar, text="查后停靠：").pack(side="left", padx=(14, 4))
        ttk.Label(report_bar, textvariable=self.post_check_dock_var).pack(side="left")

        body = ttk.Frame(self.root, padding=(10, 0, 10, 10))
        body.pack(fill="both", expand=True)

        left = ttk.LabelFrame(body, text="TEMU 店铺列表", padding=8)
        left.pack(side="left", fill="both", expand=False)

        header_bar = ttk.Frame(left)
        header_bar.pack(fill="x", pady=(0, 6))
        self.store_count_label = ttk.Label(header_bar, text="0 家店铺")
        self.store_count_label.pack(side="left")
        tk.Label(header_bar, textvariable=self.alert_count_var, fg="#C62828").pack(side="right", padx=(12, 0))
        ttk.Label(header_bar, textvariable=self.selected_count_var).pack(side="right")

        canvas = tk.Canvas(left, width=680, height=480, highlightthickness=0)
        scrollbar = ttk.Scrollbar(left, orient="vertical", command=canvas.yview)
        self.store_frame = ttk.Frame(canvas)
        self.store_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        self.store_frame_window = canvas.create_window((0, 0), window=self.store_frame, anchor="nw")
        canvas.bind("<Configure>", lambda e: canvas.itemconfigure(self.store_frame_window, width=e.width))
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        right = ttk.LabelFrame(body, text="巡查视图", padding=8)
        right.pack(side="left", fill="both", expand=True, padx=(10, 0))
        self.right_notebook = ttk.Notebook(right)
        self.right_notebook.pack(fill="both", expand=True)

        result_tab = ttk.Frame(self.right_notebook, padding=4)
        log_tab = ttk.Frame(self.right_notebook, padding=4)
        self.right_notebook.add(result_tab, text="巡查结果")
        self.right_notebook.add(log_tab, text="运行日志")

        result_paned = ttk.PanedWindow(result_tab, orient=tk.VERTICAL)
        result_paned.pack(fill="both", expand=True)

        result_columns = ("store", "conclusion", "duration")
        result_table_wrap = ttk.Frame(result_paned)
        result_detail_bar = ttk.Frame(result_paned, padding=(0, 6, 0, 0))
        result_paned.add(result_table_wrap, weight=1)
        result_paned.add(result_detail_bar, weight=3)

        self.result_tree = ttk.Treeview(result_table_wrap, columns=result_columns, show="headings", height=8)
        self.result_tree.heading("store", text="店铺")
        self.result_tree.heading("conclusion", text="结论")
        self.result_tree.heading("duration", text="耗时")
        self.result_tree.column("store", width=260, anchor="w")
        self.result_tree.column("conclusion", width=110, anchor="center")
        self.result_tree.column("duration", width=80, anchor="center")
        result_scroll = ttk.Scrollbar(result_table_wrap, orient="vertical", command=self.result_tree.yview)
        self.result_tree.configure(yscrollcommand=result_scroll.set)
        self.result_tree.tag_configure("manual", foreground="#C62828")
        self.result_tree.tag_configure("ok", foreground="#2E7D32")
        self.result_tree.tag_configure("running", foreground="#1565C0")
        self.result_tree.tag_configure("failed", foreground="#B26A00")
        self.result_tree.pack(side="left", fill="both", expand=True)
        result_scroll.pack(side="right", fill="y")
        ttk.Label(result_detail_bar, text="结果详情：").pack(anchor="w")
        self.result_detail_text = tk.Text(result_detail_bar, wrap="word", height=22)
        self.result_detail_text.pack(fill="both", expand=True)
        self.result_detail_text.configure(state="disabled")
        self.result_tree.bind("<<TreeviewSelect>>", self.on_result_tree_select)

        self.log_text = tk.Text(log_tab, wrap="word", height=30)
        self.log_text.pack(fill="both", expand=True)
        self.log_text.configure(state="disabled")

        bottom = ttk.Frame(self.root, padding=(10, 0, 10, 10))
        bottom.pack(fill="x")
        ttk.Label(bottom, text="查询进度：").pack(side="left")
        self.progress_bar = ttk.Progressbar(bottom, variable=self.progress_var, maximum=100, mode="determinate", length=260)
        self.progress_bar.pack(side="left", fill="x", expand=True, padx=(4, 10))
        ttk.Label(bottom, textvariable=self.progress_text_var).pack(side="left", padx=(0, 16))
        ttk.Label(bottom, text="最新报表：").pack(side="left")
        ttk.Label(bottom, textvariable=self.latest_report_var).pack(side="left", padx=(4, 0))
        self.on_output_mode_changed()

    def ui_call(self, func, *args) -> None:
        self.root.after(0, func, *args)

    def log(self, message: str, level: str = "INFO") -> None:
        def _append() -> None:
            normalized_level = (level or "INFO").upper()
            if normalized_level == "INFO":
                if "失败" in message or "错误" in message:
                    normalized_level = "ERROR"
                elif "停止" in message or "等待" in message:
                    normalized_level = "WARN"
            tag_name = f"log_{normalized_level}"
            color_map = {"INFO": "#222222", "WARN": "#B26A00", "ERROR": "#C62828", "SUCCESS": "#2E7D32"}
            self.log_text.configure(state="normal")
            self.log_text.tag_configure(tag_name, foreground=color_map.get(normalized_level, "#222222"))
            self.log_text.insert("end", f"{datetime.now().strftime('%H:%M:%S')}  {message}\n", tag_name)
            self.log_text.see("end")
            self.log_text.configure(state="disabled")

        self.ui_call(_append)

    def set_status(self, message: str) -> None:
        self.ui_call(self.status_var.set, message)

    def set_running(self, running: bool) -> None:
        self.running = running
        self.ui_call(self._update_run_controls, running)

    def _update_run_controls(self, running: bool) -> None:
        run_state = "disabled" if running else "normal"
        stop_state = "normal" if running else "disabled"
        self.selected_button.configure(state=run_state)
        self.all_button.configure(state=run_state)
        self.stop_button.configure(state=stop_state)
        self.low_priority_toggle_button.configure(state=run_state)
        self.api_key_button.configure(state=run_state)
        self.refresh_button.configure(state=run_state)
        self.refresh_store_button.configure(state=run_state)
        self.select_all_button.configure(state=run_state)
        self.clear_button.configure(state=run_state)
        self.price_rule_button.configure(state=run_state)
        self.select_all_checks_button.configure(state=run_state)
        self.clear_checks_button.configure(state=run_state)
        self.search_entry.configure(state=run_state)
        self.panel_only_checkbutton.configure(state=run_state)
        for entry in self.store_email_entries.values():
            entry.configure(state=run_state)
        for button in self.inspection_checkbuttons:
            button.configure(state=run_state)
        self.on_output_mode_changed()

    def set_progress(self, finished: int, total: int, message: str = "") -> None:
        total = max(total, 1)
        progress = finished * 100 / total
        text = f"{finished}/{total}"
        if message:
            text = f"{text}  {message}"
        self.ui_call(self.progress_var.set, progress)
        self.ui_call(self.progress_text_var.set, text)

    def refresh_api_key_status(self) -> None:
        _key, source = get_api_key_state()
        if source == "环境变量":
            text = "API Key：环境变量"
        elif source == "本地配置":
            text = "API Key：已配置"
        else:
            text = "API Key：未配置"
        self.ui_call(self.api_key_status_var.set, text)

    def prompt_set_api_key(self) -> None:
        current_key, source = get_api_key_state()
        if source == "环境变量":
            initial_value = ""
        else:
            initial_value = current_key
        value = simpledialog.askstring(
            "设置 API Key",
            "请输入 ZCLAW_API_KEY\n将保存到 ~/.zclaw/config.json\n环境变量仍然优先于本地配置。",
            initialvalue=initial_value,
            parent=self.root,
        )
        if value is None:
            return
        value = value.strip()
        if not value:
            messagebox.showwarning("提示", "API Key 不能为空。")
            return
        try:
            backup_path = save_api_key_to_config(value)
            self.refresh_api_key_status()
            message = "API Key 已保存到本地配置。"
            if backup_path:
                message += f"\n已备份旧配置：{backup_path}"
            messagebox.showinfo("提示", message)
            self.refresh_async()
        except Exception as error:
            messagebox.showerror("错误", f"保存 API Key 失败：{error}")

    def refresh_async(self) -> None:
        threading.Thread(target=self.refresh_connection, daemon=True).start()

    def refresh_connection(self) -> None:
        self.set_status("正在读取 ZClaw 配置...")
        self.refresh_api_key_status()
        try:
            self.api_key = load_api_key()
            self.client = ZclawClient(self.base_url, self.api_key)
            tool_names = {str(item.get("name") or "") for item in self.client.get_tools()}
            missing = sorted(name for name in TOOLS_REQUIRED if name not in tool_names)
            if missing:
                raise ZclawError(f"ZClaw 缺少必要工具：{', '.join(missing)}")
            raw = self.client.invoke("list_stores", {"all": True, "limit": 500}, timeout=60.0)
            items = raw.get("items") or []
            self.stores = [item for item in items if is_temu_store(item)]
            self.stores.sort(key=lambda item: str(item.get("storeName") or ""))
            self.set_status(f"已连接 {self.base_url}，读取到 {len(self.stores)} 家 TEMU 店铺。")
            self.log(f"已连接 ZClaw：{self.base_url}", "SUCCESS")
            self.log(f"读取到 {len(self.stores)} 家 TEMU 店铺。")
            self.ui_call(self.rebuild_store_list)
        except Exception as error:
            self.refresh_api_key_status()
            self.set_status(f"连接失败：{error}")
            self.log(f"连接失败：{error}", "ERROR")

    def choose_report_dir(self) -> None:
        selected = filedialog.askdirectory(initialdir=str(self.report_dir), title="选择报表保存位置")
        if not selected:
            return
        self.report_dir = Path(selected)
        self.report_dir_var.set(str(self.report_dir))

    def open_latest_report(self) -> None:
        if self.latest_report and self.latest_report.exists():
            os.startfile(str(self.latest_report))
            return
        messagebox.showinfo("提示", "暂无生成的报表。")

    def on_output_mode_changed(self) -> None:
        panel_only = bool(self.panel_only_var.get())
        self.report_dir_button.configure(state="disabled" if self.running or panel_only else "normal")
        open_report_state = "normal"
        if self.running or panel_only or not self.latest_report or not self.latest_report.exists():
            open_report_state = "disabled"
        self.open_report_button.configure(state=open_report_state)
        if panel_only:
            self.latest_report_var.set("本次仅面板显示")
        elif self.latest_report and self.latest_report.exists():
            self.latest_report_var.set(str(self.latest_report))
        else:
            self.latest_report_var.set("暂无报表")

    def clear_result_panel(self) -> None:
        self.result_item_ids.clear()
        self.result_detail_map.clear()
        for item_id in self.result_tree.get_children():
            self.result_tree.delete(item_id)
        self.show_result_detail("")

    def prepare_result_panel(self, stores: list[dict[str, Any]]) -> None:
        self.clear_result_panel()
        for store in stores:
            store_id = str(store.get("storeId") or "")
            store_name = str(store.get("storeName") or "")
            item_id = self.result_tree.insert(
                "",
                "end",
                values=(store_name, "等待中", "-"),
                tags=("waiting",),
            )
            self.result_item_ids[store_id] = item_id
            self.result_detail_map[store_id] = f"店铺：{store_name}\n\n状态：等待开始"
        if stores:
            first_id = str(stores[0].get("storeId") or "")
            item_id = self.result_item_ids.get(first_id)
            if item_id:
                self.result_tree.selection_set(item_id)
                self.result_tree.focus(item_id)
                self.show_result_detail(self.result_detail_map.get(first_id, ""))
        self.right_notebook.select(0)

    def show_result_detail(self, text: str) -> None:
        self.result_detail_text.configure(state="normal")
        self.result_detail_text.delete("1.0", "end")
        self.result_detail_text.insert("1.0", text or "")
        self.result_detail_text.configure(state="disabled")

    def on_result_tree_select(self, _event=None) -> None:
        selection = self.result_tree.selection()
        if not selection:
            return
        selected_item = selection[0]
        for store_id, item_id in self.result_item_ids.items():
            if item_id == selected_item:
                self.show_result_detail(self.result_detail_map.get(store_id, ""))
                break

    def update_result_row(
        self,
        store_id: str,
        store_name: str,
        progress: str,
        summary: str,
        conclusion: str,
        duration_text: str,
        detail: str,
        tag: str,
    ) -> None:
        def _update() -> None:
            item_id = self.result_item_ids.get(store_id)
            if not item_id:
                item_id = self.result_tree.insert("", "end", values=(store_name, conclusion or progress, duration_text))
                self.result_item_ids[store_id] = item_id
            row_conclusion = conclusion or progress
            self.result_tree.item(item_id, values=(store_name, row_conclusion, duration_text), tags=(tag,))
            self.result_detail_map[store_id] = detail
            if self.result_tree.selection() and self.result_tree.selection()[0] == item_id:
                self.show_result_detail(detail)

        self.ui_call(_update)

    def format_record_detail(self, record: dict[str, Any], selected_checks: set[str] | None = None) -> str:
        enabled = normalize_selected_checks(selected_checks)
        target = record.get("target") or {}
        urgent = record.get("urgent") or {}
        urgent_pending_price = record.get("urgentPendingPrice") or {}
        govern = record.get("govern") or {}
        shipping = record.get("shipping") or {}
        price_rule = record.get("priceRule") or {}
        violation = record.get("violation") or {}
        qc = record.get("qc") or {}
        project_errors = record.get("projectErrors") or []
        reasons = build_manual_reasons(record, self.recent_days, enabled)
        lines = [
            f"店铺：{target.get('storeName', '')}",
            f"检查时间：{record.get('checkedAt', '')}",
            f"巡查项目：{format_selected_checks(enabled)}",
            "",
            "结论：",
            f"{'；'.join(reasons) if reasons else '正常，无需人工处理'}",
        ]
        if "qc" in enabled:
            lines.extend(["", "抽检结果：", f"近{self.recent_days}天新增：{to_int(qc.get('recentCount'))}"])
        if "urgent" in enabled:
            lines.extend(
                [
                    "",
                    "JIT逾期：",
                    f"发货已逾期：{to_int(urgent.get('shipOverdue'))}",
                    f"到货已逾期近{ARRIVAL_OVERDUE_RECENT_DAYS}天创建：{to_int(urgent.get('arrivalOverdueRecentCount'))}",
                ]
            )
        if "urgent_declared_price" in enabled:
            if urgent_pending_price.get("error"):
                lines.extend(["", "待发货低申报价：", f"检查失败：{urgent_pending_price.get('error')}"])
            else:
                lines.extend(["", "待发货低申报价：", f"命中条数：{to_int(urgent_pending_price.get('lowPriceCount'))}"])
        if "govern" in enabled:
            lines.extend(["", "合规中心：", f"知识产权投诉：{to_int(govern.get('ipComplaintCount'))}", f"TRO：{to_int(govern.get('troCount'))}"])
        if "shipping" in enabled:
            lines.extend(["", "VMI：", f"超{SHIPPING_STALE_DAYS}天未收货：{to_int(shipping.get('staleCount'))}"])
        if "violation" in enabled:
            lines.extend(["", "违规信息：", f"待处理：{to_int(violation.get('pendingCount'))} 条"])
        if "price_rule" in enabled:
            lines.extend(["", "价格申报：", f"剩余待人工：{to_int(price_rule.get('remainingCount'))} 条"])
        if project_errors:
            lines.extend(["", "项目检查失败："])
            for item in project_errors:
                lines.append(f"- {item.get('label')}: {item.get('message')}")
        return "\n".join(lines)

    def ensure_store_status_state(self, store: dict[str, Any]) -> tk.StringVar:
        store_key = str(store.get("storeId") or "")
        if store_key not in self.store_status_vars:
            self.store_status_vars[store_key] = tk.StringVar(value="未查询")
            self.store_status_colors[store_key] = "#666666"
            self.store_alert_flags.setdefault(store_key, False)
        return self.store_status_vars[store_key]

    def ensure_store_email_var(self, store_id: str) -> tk.StringVar:
        store_key = str(store_id or "")
        if store_key not in self.store_email_vars:
            self.store_email_vars[store_key] = tk.StringVar(value=self.store_email_accounts.get(store_key, ""))
        return self.store_email_vars[store_key]

    def persist_store_email_accounts(self) -> None:
        accounts: dict[str, str] = dict(self.store_email_accounts)
        for store_id, var in self.store_email_vars.items():
            account = var.get().strip()
            if account:
                accounts[str(store_id)] = account
            else:
                accounts.pop(str(store_id), None)
        if accounts == self.store_email_accounts:
            return
        self.store_email_accounts = accounts
        try:
            backup_path = save_store_email_accounts(self.store_email_accounts)
            if backup_path:
                self.log(f"邮箱店铺配置已保存，旧配置已备份：{backup_path}", "SUCCESS")
            else:
                self.log("邮箱店铺配置已保存。", "SUCCESS")
        except Exception as error:
            self.log(f"保存邮箱店铺配置失败：{error}", "ERROR")

    def persist_store_email_account(self, store_id: str) -> None:
        store_key = str(store_id or "")
        var = self.store_email_vars.get(store_key)
        if not var:
            return
        account = var.get().strip()
        current = self.store_email_accounts.get(store_key, "")
        if account == current:
            return
        if account:
            self.store_email_accounts[store_key] = account
        else:
            self.store_email_accounts.pop(store_key, None)
        try:
            backup_path = save_store_email_accounts(self.store_email_accounts)
            if backup_path:
                self.log(f"邮箱店铺配置已保存，旧配置已备份：{backup_path}", "SUCCESS")
            else:
                self.log("邮箱店铺配置已保存。", "SUCCESS")
        except Exception as error:
            self.log(f"保存邮箱店铺配置失败：{error}", "ERROR")

    def snapshot_store_email_accounts(self, stores: list[dict[str, Any]]) -> dict[str, str]:
        accounts: dict[str, str] = {}
        for store in stores:
            store_id = str(store.get("storeId") or "")
            var = self.store_email_vars.get(store_id)
            account = var.get().strip() if var else self.store_email_accounts.get(store_id, "").strip()
            if account:
                accounts[store_id] = account
        return accounts

    def get_active_email_account(self, store_id: str) -> str:
        return self.active_email_accounts.get(str(store_id or ""), "").strip()

    def set_store_runtime_status(self, store_key: str, text: str, color: str, needs_manual: bool | None = None) -> None:
        def _update() -> None:
            status_var = self.store_status_vars.setdefault(store_key, tk.StringVar(value=text))
            status_var.set(text)
            self.store_status_colors[store_key] = color
            label = self.store_status_labels.get(store_key)
            if label:
                label.configure(fg=color)
            if needs_manual is not None:
                self.store_alert_flags[store_key] = needs_manual
            self.refresh_alert_count()

        self.ui_call(_update)

    def refresh_alert_count(self) -> None:
        current_store_keys = {str(store.get("storeId") or "") for store in self.stores}
        alert_count = sum(1 for store_key, needed in self.store_alert_flags.items() if store_key in current_store_keys and needed)
        self.alert_count_var.set(f"需处理 {alert_count} 家")

    def _update_count(self) -> None:
        selected_count = sum(1 for store in self.stores if self.check_vars.get(str(store.get("storeId") or ""), tk.BooleanVar()).get())
        self.selected_count_var.set(f"已勾选 {selected_count} 家")

    def rebuild_store_list(self) -> None:
        for child in self.store_frame.winfo_children():
            child.destroy()
        self.store_email_entries.clear()

        keyword = self.search_var.get().strip().lower()
        self.filtered_stores = []
        for store in self.stores:
            store_id = str(store.get("storeId") or "")
            store_name = str(store.get("storeName") or "")
            platform_name = str(store.get("platformName") or "")
            if keyword and keyword not in store_name.lower() and keyword not in store_id.lower():
                continue
            self.filtered_stores.append(store)
            self.check_vars.setdefault(store_id, tk.BooleanVar(value=False))
            status_var = self.ensure_store_status_state(store)
            email_var = self.ensure_store_email_var(store_id)

            card = tk.Frame(self.store_frame, bg="#D9D9D9", padx=1, pady=1)
            card.pack(fill="x", pady=(0, 8), padx=2)
            inner = tk.Frame(card, bg="#FFFFFF", padx=10, pady=8)
            inner.pack(fill="x")

            top_row = tk.Frame(inner, bg="#FFFFFF")
            top_row.pack(fill="x")
            ttk.Checkbutton(top_row, variable=self.check_vars[store_id], command=self._update_count).pack(side="left")
            store_text = f"{store_name}  |  {platform_name}  |  {store_id}"
            tk.Label(top_row, text=store_text, bg="#FFFFFF", fg="#111111", anchor="w", justify="left").pack(side="left", fill="x", expand=True, padx=(6, 0))

            email_row = tk.Frame(inner, bg="#FFFFFF")
            email_row.pack(fill="x", pady=(6, 0), padx=(26, 0))
            tk.Label(email_row, text="邮箱账号：", bg="#FFFFFF", fg="#444444").pack(side="left")
            email_entry = ttk.Entry(email_row, textvariable=email_var)
            email_entry.pack(side="left", fill="x", expand=True, padx=(4, 8))
            email_entry.bind("<FocusOut>", lambda _e, sid=store_id: self.persist_store_email_account(sid))
            email_entry.bind("<Return>", lambda _e, sid=store_id: self.persist_store_email_account(sid))
            if self.running:
                email_entry.configure(state="disabled")
            tk.Label(email_row, text="邮箱店铺填写，会永久保存", bg="#FFFFFF", fg="#888888").pack(side="left")
            self.store_email_entries[store_id] = email_entry

            status_label = tk.Label(
                inner,
                textvariable=status_var,
                bg="#FFFFFF",
                fg=self.store_status_colors.get(store_id, "#666666"),
                anchor="w",
                justify="left",
                wraplength=620,
            )
            status_label.pack(fill="x", pady=(6, 0), padx=(26, 0))
            self.store_status_labels[store_id] = status_label

        self.store_count_label.configure(text=f"{len(self.filtered_stores)} 家店铺")
        self._update_count()
        self.refresh_alert_count()

    def select_all_visible(self) -> None:
        for store in self.filtered_stores:
            self.check_vars.setdefault(str(store.get("storeId") or ""), tk.BooleanVar(value=False)).set(True)
        self._update_count()

    def clear_selection(self) -> None:
        for var in self.check_vars.values():
            var.set(False)
        self._update_count()

    def get_selected_stores(self) -> list[dict[str, Any]]:
        selected: list[dict[str, Any]] = []
        for store in self.stores:
            store_id = str(store.get("storeId") or "")
            var = self.check_vars.get(store_id)
            if var and var.get():
                selected.append(store)
        return selected

    def get_enabled_checks(self) -> set[str]:
        return normalize_selected_checks(
            {key for key, var in self.inspection_vars.items() if var.get()}
        )

    def toggle_low_priority_visibility(self) -> None:
        self.low_priority_expanded = not self.low_priority_expanded
        self.update_low_priority_visibility()

    def update_low_priority_visibility(self) -> None:
        self.low_priority_toggle_button.configure(text="▾" if self.low_priority_expanded else "▸")
        if self.low_priority_expanded:
            if not self.low_priority_group_frame.winfo_manager():
                self.low_priority_group_frame.pack(fill="x", anchor="w", pady=(8, 0))
        else:
            if self.low_priority_group_frame.winfo_manager():
                self.low_priority_group_frame.pack_forget()

    def get_problem_page_url(self, problem_key: str) -> str:
        if problem_key == "qc":
            return QC_DETAIL_URL
        if problem_key in {"urgent", "urgent_declared_price"}:
            return URGENT_STOCK_URL
        if problem_key == "govern":
            return GOVERN_DASHBOARD_URL
        if problem_key == "shipping":
            return SHIPPING_LIST_URL
        if problem_key == "violation":
            return VIOLATION_MESSAGE_URL
        if problem_key == "price_rule":
            return PRICE_RULE_URL
        if problem_key == "return_order":
            return RETURN_ORDER_URL
        if problem_key == "funds":
            return FUNDS_CENTER_URL
        return ""

    def open_additional_problem_tabs(self, store_id: str, problem_items: list[tuple[str, str]]) -> None:
        pages: list[dict[str, str]] = []
        seen_urls: set[str] = set()
        for problem_key, problem_reason in problem_items:
            url = self.get_problem_page_url(problem_key)
            if not url or url in seen_urls:
                continue
            seen_urls.add(url)
            pages.append({
                "label": INSPECTION_LABELS.get(problem_key, problem_key),
                "reason": problem_reason,
                "url": url,
            })
        if len(pages) <= 1:
            return
        extra_pages = pages[1:]
        script = f"""
(() => {{
  const pages = {json.dumps(extra_pages, ensure_ascii=False)};
  const results = [];
  for (const page of pages) {{
    try {{
      const opened = window.open(page.url, '_blank');
      results.push({{ label: page.label, url: page.url, ok: !!opened }});
    }} catch (error) {{
      results.push({{ label: page.label, url: page.url, ok: false, error: String(error && error.message || error) }});
    }}
  }}
  return results;
}})()
"""
        try:
            results = self.execute_script(store_id, script, timeout_ms=30000) or []
        except Exception as error:
            self.log(f"额外问题页标签打开失败：{error}", "WARN")
            return
        opened = [str(item.get("label") or item.get("url") or "") for item in results if item.get("ok")]
        failed = [str(item.get("label") or item.get("url") or "") for item in results if not item.get("ok")]
        if opened:
            self.log(f"已额外打开问题页标签：{'、'.join(opened)}", "WARN")
        if failed:
            self.log(f"额外问题页未能自动打开：{'、'.join(failed)}，可按面板/报表手动进入。", "WARN")

    def focus_problem_page(self, store_id: str, record: dict[str, Any], selected_checks: set[str] | None = None) -> bool:
        problem_items = build_manual_problem_items(record, self.recent_days, selected_checks)
        if not problem_items:
            return False
        problem_key, problem_reason = problem_items[0]
        label = INSPECTION_LABELS.get(problem_key, problem_key)
        if len(problem_items) > 1:
            self.log(f"该店铺共有 {len(problem_items)} 个需处理项：{'；'.join(reason for _key, reason in problem_items)}", "WARN")
        self.log(f"定位问题页面：{label} -> {problem_reason}", "WARN")
        if problem_key == "qc":
            self.navigate_qc_detail(store_id)
        elif problem_key == "urgent":
            self.navigate_urgent_stock(store_id)
            urgent = record.get("urgent") or {}
            if "到货已逾期" in problem_reason:
                self.navigate_urgent_arrival_overdue(store_id, to_int(urgent.get("arrivalOverdue")))
        elif problem_key == "urgent_declared_price":
            self.navigate_urgent_stock(store_id)
            self.navigate_urgent_pending_price_tab(store_id)
        elif problem_key == "govern":
            self.navigate_govern_dashboard(store_id)
        elif problem_key == "shipping":
            self.navigate_shipping_list(store_id)
        elif problem_key == "violation":
            self.navigate_violation_message(store_id)
        elif problem_key == "price_rule":
            self.navigate_price_rule(store_id)
        elif problem_key == "return_order":
            self.navigate_return_order(store_id)
        elif problem_key == "funds":
            self.navigate_funds_center(store_id)
        self.open_additional_problem_tabs(store_id, problem_items)
        return True

    def select_all_inspections(self) -> None:
        for var in self.inspection_vars.values():
            var.set(True)
        self.low_priority_expanded = True
        self.update_low_priority_visibility()

    def clear_inspections(self) -> None:
        for var in self.inspection_vars.values():
            var.set(False)

    def raise_if_stop_requested(self) -> None:
        if self.stop_requested:
            raise StopRequestedError("用户已停止任务")

    def sleep_with_stop(self, seconds: float, interval: float = 0.2) -> None:
        remaining = max(0.0, seconds)
        while remaining > 0:
            self.raise_if_stop_requested()
            step = min(interval, remaining)
            time.sleep(step)
            remaining -= step

    def start_batch_async(self, stores: list[dict[str, Any]], enabled_checks: set[str]) -> None:
        if self.running or (self.batch_thread and self.batch_thread.is_alive()):
            self.log("批量任务正在执行中，请勿重复启动。", "WARN")
            return
        self.stop_requested = False
        self.persist_store_email_accounts()
        self.active_email_accounts = self.snapshot_store_email_accounts(stores)
        self.prepare_result_panel(stores)
        self.set_running(True)
        self.batch_thread = threading.Thread(target=self.run_batch, args=(stores, enabled_checks), daemon=True)
        self.batch_thread.start()

    def close_price_rule_editor(self) -> None:
        if self.price_rule_editor and self.price_rule_editor.winfo_exists():
            self.price_rule_editor.destroy()
        self.price_rule_editor = None
        self.price_rule_rows_container = None
        self.price_rule_editor_rows = []

    def open_price_rule_editor(self) -> None:
        if self.price_rule_editor and self.price_rule_editor.winfo_exists():
            self.price_rule_editor.lift()
            self.price_rule_editor.focus_force()
            return

        self.price_rule_config = normalize_price_rule_config(self.price_rule_config)
        self.price_rule_protect_var.set(bool(self.price_rule_config.get("protectDiff", True)))
        self.price_rule_limit_var.set(f"{to_float(self.price_rule_config.get('protectDiffLimit', 1.0), 1.0):g}")

        editor = tk.Toplevel(self.root)
        editor.title("价格规则设置")
        editor.geometry("760x520")
        editor.transient(self.root)
        editor.grab_set()
        editor.protocol("WM_DELETE_WINDOW", self.close_price_rule_editor)
        self.price_rule_editor = editor
        self.price_rule_editor_rows = []

        container = ttk.Frame(editor, padding=12)
        container.pack(fill="both", expand=True)

        ttk.Label(
            container,
            text="价格申报自动助手会自动执行规则并延时确认，未命中项会在报表里标记为需人工处理。",
        ).pack(anchor="w")

        common_bar = ttk.Frame(container)
        common_bar.pack(fill="x", pady=(10, 8))
        ttk.Checkbutton(
            common_bar,
            text="开启价差保护（当前价与调整后价差值大于阈值时改为“不调整”）",
            variable=self.price_rule_protect_var,
        ).pack(side="left")
        ttk.Label(common_bar, text="阈值：").pack(side="left", padx=(12, 4))
        ttk.Entry(common_bar, textvariable=self.price_rule_limit_var, width=8).pack(side="left")

        header = ttk.Frame(container)
        header.pack(fill="x", pady=(8, 0))
        ttk.Label(header, text="关键词", width=26, anchor="w").pack(side="left")
        ttk.Label(header, text="最小价", width=12, anchor="w").pack(side="left", padx=(6, 0))
        ttk.Label(header, text="最大价", width=12, anchor="w").pack(side="left", padx=(6, 0))
        ttk.Label(header, text="动作", width=10, anchor="w").pack(side="left", padx=(6, 0))

        rows_wrap = ttk.Frame(container)
        rows_wrap.pack(fill="both", expand=True, pady=(6, 8))

        canvas = tk.Canvas(rows_wrap, highlightthickness=0)
        scrollbar = ttk.Scrollbar(rows_wrap, orient="vertical", command=canvas.yview)
        inner = ttk.Frame(canvas)
        inner.bind("<Configure>", lambda _e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=inner, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        self.price_rule_rows_container = inner

        for rule in self.price_rule_config.get("rules") or []:
            self.add_price_rule_editor_row(rule)
        if not self.price_rule_editor_rows:
            self.add_price_rule_editor_row()

        action_bar = ttk.Frame(container)
        action_bar.pack(fill="x")
        ttk.Button(action_bar, text="添加规则", command=self.add_price_rule_editor_row).pack(side="left")
        ttk.Button(action_bar, text="恢复默认", command=self.reset_price_rule_defaults).pack(side="left", padx=(8, 0))
        ttk.Button(action_bar, text="保存", command=self.save_price_rule_editor).pack(side="right")
        ttk.Button(action_bar, text="关闭", command=self.close_price_rule_editor).pack(side="right", padx=(0, 8))

    def add_price_rule_editor_row(self, rule: dict[str, Any] | None = None) -> None:
        if not self.price_rule_rows_container:
            return
        rule = rule or {}
        frame = ttk.Frame(self.price_rule_rows_container)
        frame.pack(fill="x", pady=3)

        min_value = rule.get("min", "")
        max_value = rule.get("max", "")
        row_data = {
            "frame": frame,
            "kw_var": tk.StringVar(value=str(rule.get("kw") or "")),
            "min_var": tk.StringVar(value="" if min_value in ("", None) else f"{float(min_value):g}"),
            "max_var": tk.StringVar(value="" if max_value in ("", None) else f"{float(max_value):g}"),
            "action_var": tk.StringVar(value="不调整" if str(rule.get("action") or "") == "不调整" else "调整"),
        }

        ttk.Entry(frame, textvariable=row_data["kw_var"], width=28).pack(side="left")
        ttk.Entry(frame, textvariable=row_data["min_var"], width=12).pack(side="left", padx=(6, 0))
        ttk.Entry(frame, textvariable=row_data["max_var"], width=12).pack(side="left", padx=(6, 0))
        ttk.Combobox(
            frame,
            textvariable=row_data["action_var"],
            values=("调整", "不调整"),
            width=8,
            state="readonly",
        ).pack(side="left", padx=(6, 0))
        ttk.Button(frame, text="删除", command=lambda current=row_data: self.remove_price_rule_editor_row(current)).pack(side="left", padx=(8, 0))
        self.price_rule_editor_rows.append(row_data)

    def remove_price_rule_editor_row(self, row_data: dict[str, Any]) -> None:
        if len(self.price_rule_editor_rows) <= 1:
            messagebox.showwarning("提示", "至少保留一条规则。")
            return
        frame = row_data.get("frame")
        if frame and frame.winfo_exists():
            frame.destroy()
        self.price_rule_editor_rows = [item for item in self.price_rule_editor_rows if item is not row_data]

    def reset_price_rule_defaults(self) -> None:
        self.price_rule_protect_var.set(bool(DEFAULT_PRICE_RULE_CONFIG.get("protectDiff", True)))
        self.price_rule_limit_var.set(f"{to_float(DEFAULT_PRICE_RULE_CONFIG.get('protectDiffLimit', 1.0), 1.0):g}")
        for row_data in list(self.price_rule_editor_rows):
            frame = row_data.get("frame")
            if frame and frame.winfo_exists():
                frame.destroy()
        self.price_rule_editor_rows = []
        for rule in DEFAULT_PRICE_RULE_CONFIG.get("rules") or []:
            self.add_price_rule_editor_row(rule)

    def save_price_rule_editor(self) -> None:
        limit_text = self.price_rule_limit_var.get().strip()
        try:
            protect_limit = 1.0 if not limit_text else float(limit_text)
        except ValueError:
            messagebox.showwarning("提示", "价差阈值必须是数字。")
            return
        if protect_limit < 0:
            messagebox.showwarning("提示", "价差阈值不能小于 0。")
            return

        rules: list[dict[str, Any]] = []
        for index, row_data in enumerate(self.price_rule_editor_rows, start=1):
            keyword = row_data["kw_var"].get().strip()
            min_text = row_data["min_var"].get().strip()
            max_text = row_data["max_var"].get().strip()
            action = "不调整" if row_data["action_var"].get().strip() == "不调整" else "调整"

            if not keyword and not min_text and not max_text:
                continue
            if not keyword:
                messagebox.showwarning("提示", f"第 {index} 条规则缺少关键词。")
                return
            try:
                min_value = "" if not min_text else float(min_text)
                max_value = "" if not max_text else float(max_text)
            except ValueError:
                messagebox.showwarning("提示", f"第 {index} 条规则的价格范围格式不对。")
                return
            if min_value != "" and max_value != "" and min_value > max_value:
                messagebox.showwarning("提示", f"第 {index} 条规则的最小价不能大于最大价。")
                return
            rules.append(
                {
                    "kw": keyword,
                    "min": min_value,
                    "max": max_value,
                    "action": action,
                }
            )

        if not rules:
            messagebox.showwarning("提示", "请至少保留一条有效规则。")
            return

        self.price_rule_config = normalize_price_rule_config(
            {
                "protectDiff": self.price_rule_protect_var.get(),
                "protectDiffLimit": protect_limit,
                "rules": rules,
            }
        )
        save_price_rule_config(self.price_rule_config)
        self.log(f"价格规则已保存，共 {len(self.price_rule_config.get('rules') or [])} 条。", "SUCCESS")
        self.close_price_rule_editor()

    def run_selected_async(self) -> None:
        selected = self.get_selected_stores()
        if not selected:
            messagebox.showwarning("提示", "请先勾选至少一家店铺。")
            return
        enabled_checks = self.get_enabled_checks()
        if not enabled_checks:
            messagebox.showwarning("提示", "请至少勾选一个巡查项目。")
            return
        self.start_batch_async(selected, enabled_checks)

    def run_all_async(self) -> None:
        if not self.stores:
            messagebox.showwarning("提示", "当前没有可巡的 TEMU 店铺。")
            return
        enabled_checks = self.get_enabled_checks()
        if not enabled_checks:
            messagebox.showwarning("提示", "请至少勾选一个巡查项目。")
            return
        self.start_batch_async(list(self.stores), enabled_checks)

    def stop_batch(self) -> None:
        self.stop_requested = True
        self.log("已收到停止请求，将尽快停止当前任务。", "WARN")

    def execute_script(self, store_id: str, script: str, timeout_ms: int = 30000) -> Any:
        if not self.client:
            raise ZclawError("ZClaw 未连接。")
        result = self.client.invoke(
            "execute_script",
            {"storeId": store_id, "script": script, "returnByValue": True, "timeoutMs": timeout_ms},
            timeout=max(30.0, timeout_ms / 1000 + 15.0),
        )
        return result.get("result")

    def build_email_login_auth_script(self, email_account: str) -> str:
        account = (email_account or "").strip()
        if not account:
            return EMAIL_LOGIN_AUTH_SCRIPT
        return EMAIL_ACCOUNT_LOGIN_AUTH_SCRIPT_TEMPLATE.replace(
            "__EMAIL_ACCOUNT__",
            json.dumps(account, ensure_ascii=False),
        )

    def add_project_error(
        self,
        project_errors: list[dict[str, Any]],
        key: str,
        error: Exception | str,
    ) -> None:
        message = str(error).strip() or "未知错误"
        label = INSPECTION_LABELS.get(key, key)
        project_errors.append({"key": key, "label": label, "message": message})

    def run_urgent_declared_price_with_retry(self, store_id: str, store_name: str) -> dict[str, Any]:
        last_error: Exception | None = None
        for attempt in range(1, PROJECT_RETRY_COUNT + 1):
            self.raise_if_stop_requested()
            try:
                if attempt > 1:
                    self.log(
                        f"检查待发货低申报价：{store_name} 第 {attempt}/{PROJECT_RETRY_COUNT} 次重试，刷新页面后继续。",
                        "WARN",
                    )
                    self.visit_page_with_retry(store_id, URGENT_STOCK_URL, "检查待发货低申报价页刷新重试")
                    self.sleep_with_stop(3.0)
                self.navigate_urgent_pending_price_tab(store_id)
                return self.extract_urgent_pending_price_rows(store_id)
            except StopRequestedError:
                raise
            except Exception as error:
                last_error = error
                if attempt < PROJECT_RETRY_COUNT:
                    self.log(
                        f"检查待发货低申报价：{store_name} 第 {attempt} 次失败，准备刷新重试：{error}",
                        "WARN",
                    )
                    try:
                        self.visit_page_with_retry(store_id, URGENT_STOCK_URL, "检查待发货低申报价页失败恢复")
                    except Exception as recover_error:
                        self.log(f"检查待发货低申报价：{store_name} 恢复页面失败：{recover_error}", "WARN")
                    self.sleep_with_stop(3.0)
        raise ZclawError(f"检查待发货低申报价连续失败，已跳过该子项目：{last_error}")

    def build_price_rule_execute_script(self) -> str:
        config = normalize_price_rule_config(self.price_rule_config)
        protect_diff = bool(config.get("protectDiff", True))
        protect_limit = max(0.0, to_float(config.get("protectDiffLimit", 1.0), 1.0))
        payload_rules = []
        for rule in config.get("rules") or []:
            payload_rules.append(
                {
                    "kw": str(rule.get("kw") or "").strip(),
                    "min": float("-inf") if rule.get("min") in ("", None) else float(rule.get("min")),
                    "max": float("inf") if rule.get("max") in ("", None) else float(rule.get("max")),
                    "action": "不调整" if str(rule.get("action") or "") == "不调整" else "调整",
                    "protectDiff": protect_diff,
                    "protectDiffLimit": protect_limit,
                }
            )
        payload = {
            "protectDiff": protect_diff,
            "protectDiffLimit": protect_limit,
            "rules": payload_rules,
        }
        return PRICE_RULE_RUN_SCRIPT_TEMPLATE.replace(
            "__PRICE_RULE_PAYLOAD__",
            json.dumps(payload, ensure_ascii=False, allow_nan=True),
        )

    def navigate_price_rule(self, store_id: str) -> dict[str, Any]:
        self.visit_page_with_retry(store_id, PRICE_RULE_URL, "价格申报页")
        deadline = time.monotonic() + 90.0
        last_status = ""
        last_action_at = 0.0
        page_size_attempts = 0
        zero_wait_seen_at: float | None = None
        while time.monotonic() < deadline:
            self.raise_if_stop_requested()
            state = self.execute_script(store_id, PRICE_RULE_PAGE_STATE_SCRIPT, timeout_ms=30000) or {}
            title = normalize_text(str(state.get("title") or ""))
            body = normalize_text(str(state.get("body") or ""))
            active_tab = normalize_text(str(state.get("activeTabText") or ""))
            waiting_count = to_int(state.get("waitingCount"))
            row_count = to_int(state.get("rowCount"))
            page_size = normalize_text(str(state.get("pageSize") or ""))
            target_rows = min(waiting_count, PRICE_RULE_PAGE_SIZE) if waiting_count > 0 else 0
            tabs_ready = bool(state.get("priceTabsReady"))
            waiting_tab_seen = bool(state.get("waitingTabSeen"))
            now = time.monotonic()

            if self.maybe_handle_temu_auth_gate(store_id, "价格申报页", PRICE_RULE_URL, state):
                last_status = ""
                last_action_at = 0.0
                zero_wait_seen_at = None
                continue
            if state.get("loginPrompt"):
                raise ZclawError("跳转到卖家中心登录页，当前登录态未同步。")

            if "价格申报" in title or "待卖家确认" in body or "/adjust-price-manage/order-price" in str(state.get("url") or ""):
                if not tabs_ready:
                    status = "价格申报页已打开，等待标签区加载..."
                    if status != last_status:
                        self.log(status)
                        last_status = status
                    self.sleep_with_stop(1.0)
                    continue

                if waiting_count <= 0 and waiting_tab_seen:
                    if zero_wait_seen_at is None:
                        zero_wait_seen_at = now
                    elif now - zero_wait_seen_at >= 2.0:
                        return {
                            "url": state.get("url"),
                            "title": title,
                            "waitingCount": 0,
                            "pageSize": page_size,
                            "rowCount": row_count,
                            "activeTabText": active_tab,
                        }
                else:
                    zero_wait_seen_at = None

                if waiting_count > 0 and now - last_action_at >= 1.0 and (not active_tab.startswith("待卖家确认") or (row_count == 0 and state.get("hasNoData"))):
                    click_result = self.execute_script(store_id, PRICE_RULE_CLICK_WAITING_TAB_SCRIPT, timeout_ms=30000) or {}
                    self.log(f"价格申报页尝试切换待卖家确认：{click_result}")
                    last_action_at = now
                    self.sleep_with_stop(2.0)
                    continue

                if waiting_count <= 0:
                    status = "价格申报页已打开，但待卖家确认计数还未稳定，继续等待..."
                    if status != last_status:
                        self.log(status)
                        last_status = status
                    self.sleep_with_stop(1.0)
                    continue

                current_page_size = to_int(page_size, 0)
                if active_tab.startswith("待卖家确认") and waiting_count > max(current_page_size, 10) and page_size != str(PRICE_RULE_PAGE_SIZE) and page_size_attempts < 3 and now - last_action_at >= 1.0:
                    size_result = self.execute_script(store_id, PRICE_RULE_SET_PAGE_SIZE_100_SCRIPT, timeout_ms=30000) or {}
                    self.log(f"价格申报页尝试切换每页100条：{size_result}")
                    page_size_attempts += 1
                    last_action_at = now
                    self.sleep_with_stop(2.0)
                    continue

                if active_tab.startswith("待卖家确认") and state.get("hasTable") and state.get("hasBatchButton") and row_count >= max(1, target_rows):
                    return {
                        "url": state.get("url"),
                        "title": title,
                        "waitingCount": waiting_count,
                        "pageSize": page_size,
                        "rowCount": row_count,
                        "activeTabText": active_tab,
                    }

                status = f"价格申报页已打开，等待待卖家确认列表加载：{body[:80]}"
                if status != last_status:
                    self.log(status)
                    last_status = status

            self.sleep_with_stop(2.0)
        raise ZclawError("等待价格申报页超时。")

    def run_price_rule_assistant(self, store_id: str) -> dict[str, Any]:
        page_state = self.execute_script(store_id, PRICE_RULE_PAGE_STATE_SCRIPT, timeout_ms=30000) or {}
        waiting_count = to_int(page_state.get("waitingCount"))
        if waiting_count <= 0:
            return {
                "url": page_state.get("url"),
                "title": page_state.get("title"),
                "waitingCount": 0,
                "matchedCount": 0,
                "actedCount": 0,
                "protectedCount": 0,
                "unmatchedModalRows": 0,
                "remainingCount": 0,
                "selectedPreview": [],
                "stage": "done",
                "confirmPending": False,
                "confirmed": False,
                "message": "待卖家确认为 0",
            }

        self.execute_script(store_id, self.build_price_rule_execute_script(), timeout_ms=30000)
        deadline = time.monotonic() + max(90.0, min(180.0, 45.0 + waiting_count * 1.2))
        last_status = ""
        while time.monotonic() < deadline:
            self.raise_if_stop_requested()
            state = self.execute_script(store_id, PRICE_RULE_FETCH_STATE_SCRIPT, timeout_ms=30000) or {}
            stage = str(state.get("stage") or "")
            message = normalize_text(str(state.get("message") or ""))
            if stage == "error":
                raise ZclawError(f"价格申报自动助手执行失败：{state.get('error') or message or '未知错误'}")
            if stage in {"await_confirm", "confirm_failed", "done"}:
                return {
                    "url": page_state.get("url"),
                    "title": page_state.get("title"),
                    "waitingCount": waiting_count,
                    "matchedCount": to_int(state.get("matchedCount")),
                    "actedCount": to_int(state.get("actedCount")),
                    "protectedCount": to_int(state.get("protectedCount")),
                    "unmatchedModalRows": to_int(state.get("unmatchedModalRows")),
                    "remainingCount": to_int(state.get("remainingCount")),
                    "selectedPreview": state.get("selectedPreview") or [],
                    "stage": stage,
                    "confirmPending": bool(state.get("confirmPending")),
                    "confirmed": bool(state.get("confirmed")),
                    "message": message,
                }
            status = f"价格申报自动助手执行中：{message or stage or '处理中'}"
            if status != last_status:
                self.log(status)
                last_status = status
            self.sleep_with_stop(1.2)
        raise ZclawError("价格申报自动助手执行超时。")

    def visit_page_with_retry(self, store_id: str, url: str, page_name: str, wait_until: str = "load", timeout_ms: int = 120000) -> None:
        if not self.client:
            raise ZclawError("ZClaw 未连接。")
        last_error: Exception | None = None
        for attempt in range(1, NAVIGATION_RETRY_COUNT + 1):
            self.raise_if_stop_requested()
            try:
                self.client.invoke(
                    "visit_page",
                    {"storeId": store_id, "url": url, "waitUntil": wait_until, "timeoutMs": timeout_ms},
                    timeout=max(150.0, timeout_ms / 1000 + 30.0),
                )
                return
            except Exception as error:
                last_error = error
                if attempt >= NAVIGATION_RETRY_COUNT or not is_transient_navigation_error(error):
                    raise
                self.log(
                    f"{page_name} 导航命中扩展页或启动瞬态，等待稳定后重试 {attempt}/{NAVIGATION_RETRY_COUNT - 1}：{error}",
                    "WARN",
                )
                self.sleep_with_stop(NAVIGATION_RETRY_DELAY * attempt)
        if last_error:
            raise last_error

    def dock_store_window_after_check(self, store_id: str, store_name: str) -> None:
        if not self.client:
            return
        self.client.invoke(
            "visit_page",
            {"storeId": store_id, "url": POST_CHECK_DOCK_URL, "waitUntil": "load", "timeoutMs": 120000},
            timeout=150.0,
        )
        self.log(f"查后已停靠核价页面：{store_name} -> {POST_CHECK_DOCK_URL}")

    def reopen_store_for_email_login(self, store_id: str) -> None:
        if not self.client:
            raise ZclawError("ZClaw 未连接。")
        try:
            self.client.invoke("close_store", {"storeId": store_id}, timeout=45.0)
        except Exception as error:
            self.log(f"邮箱登录重试前关闭店铺窗口失败，将继续重开：{error}", "WARN")
        self.sleep_with_stop(2.0)
        self.client.invoke("open_store", {"storeId": store_id, "launchUrl": TEMU_HOME_URL}, timeout=180.0)
        self.sleep_with_stop(3.0)

    def maybe_handle_temu_auth_gate(self, store_id: str, page_name: str, target_url: str, state: dict[str, Any] | None = None) -> bool:
        if not state_looks_like_temu_auth_gate(state):
            return False
        self.log(f"{page_name}跳回登录入口，先回首页确认登录态...", "WARN")
        self.ensure_temu_home(store_id)
        self.visit_page_with_retry(store_id, target_url, page_name)
        self.sleep_with_stop(2.0)
        return True

    def ensure_temu_home(self, store_id: str, email_account: str | None = None) -> None:
        email_account = (email_account if email_account is not None else self.get_active_email_account(store_id)).strip()
        self.visit_page_with_retry(store_id, TEMU_HOME_URL, "TEMU 首页")
        deadline = time.monotonic() + 120.0
        last_status = ""
        region_clicked = False
        auth_clicked = False
        confirm_forward_clicked = False
        auth_stuck_since: float | None = None
        auth_prompt_seen_at: float | None = None
        auth_wait_logged = False
        email_login_retry_count = 0
        email_option_missing_since: float | None = None
        while time.monotonic() < deadline:
            self.raise_if_stop_requested()
            state = self.execute_script(store_id, TEMU_STATE_SCRIPT, timeout_ms=30000) or {}
            body = normalize_text(str(state.get("body") or ""))
            now = time.monotonic()
            if state.get("homeReady"):
                return
            if state.get("hasRegionPage") and state.get("regionButtons"):
                if not region_clicked:
                    status = "检测到商家中心入口，正在尝试进入首页..."
                    if status != last_status:
                        self.log(status)
                        last_status = status
                    self.execute_script(store_id, CLICK_REGION_SCRIPT, timeout_ms=30000)
                    region_clicked = True
                    auth_stuck_since = time.monotonic()
                    self.sleep_with_stop(3.0)
                    continue
                if auth_stuck_since is not None and time.monotonic() - auth_stuck_since >= 8.0:
                    raise ZclawError("首页仍停留在商家中心入口页，请先在弹出的登录窗完成授权登录后重试。")
            if state.get("hasConfirmForwardButton") or state.get("hasConfirmForwardPrompt"):
                if not confirm_forward_clicked and state.get("hasConfirmForwardButton"):
                    status = "检测到确认授权弹窗，正在确认并前往..."
                    if status != last_status:
                        self.log(status)
                        last_status = status
                    self.execute_script(store_id, CLICK_CONFIRM_FORWARD_SCRIPT, timeout_ms=30000)
                    confirm_forward_clicked = True
                    auth_stuck_since = time.monotonic()
                    self.sleep_with_stop(5.0)
                    continue
                if auth_stuck_since is not None and time.monotonic() - auth_stuck_since >= 8.0:
                    raise ZclawError("首页仍停留在确认授权弹窗，请先在弹出的登录窗完成确认后重试。")
            if state.get("phoneFormatError") or state.get("needsEmailLogin") or state.get("emailLoginPage"):
                result = self.execute_script(store_id, self.build_email_login_auth_script(email_account), timeout_ms=30000) or {}
                stage = str(result.get("stage") or "")
                if stage in {"switched_to_email", "email_selected", "email_account_filled"}:
                    email_option_missing_since = None
                    if stage == "switched_to_email":
                        status = "检测到短号账号，已切换邮箱登录。"
                    elif stage == "email_account_filled":
                        status = f"已填写邮箱账号：{result.get('accountValue') or email_account}"
                    else:
                        status = f"已选择邮箱账号：{result.get('emailText') or ''}"
                    if status != last_status:
                        self.log(status)
                        last_status = status
                    self.sleep_with_stop(1.5)
                    continue
                if stage == "email_auth_clicked":
                    status = "邮箱账号已就绪，正在授权登录..."
                    if status != last_status:
                        self.log(status)
                        last_status = status
                    auth_clicked = True
                    auth_stuck_since = now
                    email_option_missing_since = None
                    self.sleep_with_stop(5.0)
                    continue
                if stage == "email_account_protected":
                    raise ZclawError("网页拒绝手输邮箱账号：仅支持登录下拉框内的账号，需要人工处理。")
                if stage == "email_account_empty":
                    raise ZclawError("该店铺需要邮箱登录，请先在店铺卡片里填写邮箱账号。")
                if stage == "email_option_missing":
                    if email_option_missing_since is None:
                        email_option_missing_since = now
                        self.log("邮箱登录账号选项暂未出现，等待后重试...", "WARN")
                        self.sleep_with_stop(2.0)
                        continue
                    email_login_retry_count += 1
                    if email_login_retry_count > EMAIL_LOGIN_RETRY_COUNT:
                        raise ZclawError(f"邮箱登录账号选项连续 {EMAIL_LOGIN_RETRY_COUNT} 次未出现，需要人工处理。")
                    self.log(f"邮箱登录账号选项未出现，重开店铺窗口重试 {email_login_retry_count}/{EMAIL_LOGIN_RETRY_COUNT}。", "WARN")
                    self.reopen_store_for_email_login(store_id)
                    region_clicked = False
                    auth_clicked = False
                    confirm_forward_clicked = False
                    auth_stuck_since = None
                    auth_prompt_seen_at = None
                    auth_wait_logged = False
                    email_option_missing_since = None
                    last_status = ""
                    continue
                if stage not in {"not_email_login_page", ""}:
                    raise ZclawError(f"邮箱登录处理失败：{stage or result}")
            if state.get("hasAuthButton"):
                if auth_prompt_seen_at is None:
                    auth_prompt_seen_at = now
                    auth_wait_logged = False
                if not auth_clicked:
                    auth_ready = bool(state.get("authCredentialsReady"))
                    auth_visible_long_enough = now - auth_prompt_seen_at >= 6.0
                    if auth_ready or auth_visible_long_enough:
                        status = "检测到授权页，账号密码已就绪，正在尝试授权登录..."
                        if status != last_status:
                            self.log(status)
                            last_status = status
                        self.execute_script(store_id, CLICK_AUTH_SCRIPT, timeout_ms=30000)
                        auth_clicked = True
                        auth_stuck_since = now
                        self.sleep_with_stop(5.0)
                        continue
                    if not auth_wait_logged:
                        filled = f"{to_int(state.get('authFilledCount'))}/{to_int(state.get('authInputCount'))}"
                        self.log(f"检测到授权页，等待账号密码自动填充后再授权登录：已填 {filled}", "WARN")
                        auth_wait_logged = True
                if auth_stuck_since is not None and time.monotonic() - auth_stuck_since >= 8.0:
                    raise ZclawError("首页仍停留在授权登录页，请先在弹出的登录窗完成授权登录后重试。")
            else:
                auth_prompt_seen_at = None
                auth_wait_logged = False
            if state.get("loginPrompt"):
                if auth_stuck_since is None:
                    auth_stuck_since = now
                elif now - auth_stuck_since >= 8.0:
                    raise ZclawError(f"登录未完成：{body[:160]}")
            status = f"等待首页加载：{body[:80]}"
            if status != last_status:
                self.log(status)
                last_status = status
            self.sleep_with_stop(2.0)
        raise ZclawError("等待 TEMU 首页超时。")

    def navigate_qc_detail(self, store_id: str) -> None:
        self.visit_page_with_retry(store_id, QC_DETAIL_URL, "抽检结果明细页")
        deadline = time.monotonic() + 60.0
        last_status = ""
        while time.monotonic() < deadline:
            self.raise_if_stop_requested()
            state = self.execute_script(store_id, QC_READY_SCRIPT, timeout_ms=30000) or {}
            title = str(state.get("title") or "")
            text = normalize_text(str(state.get("body") or ""))
            url = str(state.get("url") or "")
            row_count = int(state.get("rowCount") or 0)
            if self.maybe_handle_temu_auth_gate(store_id, "抽检结果明细页", QC_DETAIL_URL, state):
                last_status = ""
                continue
            if state.get("loginPrompt"):
                raise ZclawError("跳转到卖家中心登录页，当前登录态未同步。")
            if "抽检结果明细" in title or "抽检结果明细" in text:
                if row_count > 0 or state.get("hasNoData") or state.get("hasTotal"):
                    return
                status = f"抽检页已打开，等待表格数据渲染：{text[:80]}"
                if status != last_status:
                    self.log(status)
                    last_status = status
            self.sleep_with_stop(2.0)
        raise ZclawError("等待抽检结果明细页超时。")

    def navigate_urgent_stock(self, store_id: str) -> None:
        self.visit_page_with_retry(store_id, URGENT_STOCK_URL, "检查JIT是否逾期页")
        deadline = time.monotonic() + 60.0
        last_status = ""
        while time.monotonic() < deadline:
            self.raise_if_stop_requested()
            state = self.execute_script(store_id, URGENT_METRICS_SCRIPT, timeout_ms=30000) or {}
            title = str(state.get("title") or "")
            text = normalize_text(str(state.get("body") or ""))
            ship_overdue = state.get("shipOverdue")
            arrival_overdue = state.get("arrivalOverdue")
            if self.maybe_handle_temu_auth_gate(store_id, "检查JIT是否逾期页", URGENT_STOCK_URL, state):
                last_status = ""
                continue
            if state.get("loginPrompt"):
                raise ZclawError("跳转到卖家中心登录页，当前登录态未同步。")
            if state.get("isUrgentPage"):
                if state.get("ready") and ship_overdue is not None and arrival_overdue is not None:
                    return
                status = f"检查JIT是否逾期页已打开，等待指标加载：{text[:80]}"
                if status != last_status:
                    self.log(status)
                    last_status = status
            self.sleep_with_stop(2.0)
        raise ZclawError("等待检查JIT是否逾期页超时。")

    def extract_urgent_metrics(self, store_id: str) -> dict[str, Any]:
        deadline = time.monotonic() + 20.0
        raw: dict[str, Any] = {}
        warned = False
        while time.monotonic() < deadline:
            self.raise_if_stop_requested()
            raw = self.execute_script(store_id, URGENT_METRICS_SCRIPT, timeout_ms=30000) or {}
            ship_overdue = raw.get("shipOverdue")
            arrival_overdue = raw.get("arrivalOverdue")
            if ship_overdue is not None and arrival_overdue is not None:
                return {
                    "url": raw.get("url"),
                    "title": raw.get("title"),
                    "shipOverdue": to_int(ship_overdue),
                    "arrivalOverdue": to_int(arrival_overdue),
                    "arrivalOverdueRecentCount": 0,
                    "arrivalOverdueRows": [],
                    "bodyPreview": raw.get("section") or raw.get("body"),
                }
            if not warned:
                self.log("检查JIT是否逾期页已打开，但指标还没渲染完，继续等待...", "WARN")
                warned = True
            self.sleep_with_stop(2.0)
        raise ZclawError("检查JIT是否逾期页指标读取失败。")

    def navigate_urgent_arrival_overdue(self, store_id: str, expected_count: int) -> None:
        if expected_count <= 0:
            return
        deadline = time.monotonic() + 60.0
        last_status = ""
        last_action_at = 0.0
        while time.monotonic() < deadline:
            self.raise_if_stop_requested()
            state = self.execute_script(store_id, URGENT_ARRIVAL_OVERDUE_STATE_SCRIPT, timeout_ms=30000) or {}
            body = normalize_text(str(state.get("body") or ""))
            actual_row_count = to_int(state.get("actualRowCount"))
            selected = bool(state.get("selected"))
            if self.maybe_handle_temu_auth_gate(store_id, "到货已逾期筛选", URGENT_STOCK_URL, state):
                last_status = ""
                last_action_at = 0.0
                continue
            if state.get("loginPrompt"):
                raise ZclawError("跳转到卖家中心登录页，当前登录态未同步。")
            if selected and actual_row_count > 0 and actual_row_count <= expected_count:
                return
            if selected and state.get("hasNoData") and expected_count == 0:
                return
            now = time.monotonic()
            if state.get("hasCard") and not selected and now - last_action_at >= 1.5:
                mark = self.execute_script(store_id, URGENT_MARK_ARRIVAL_OVERDUE_CARD_SCRIPT, timeout_ms=30000) or {}
                if not mark.get("ok"):
                    raise ZclawError("未找到到货已逾期筛选按钮。")
                if not self.client:
                    raise ZclawError("ZClaw 未连接。")
                self.client.invoke(
                    "click_element",
                    {"storeId": store_id, "selector": "#codex-arrival-overdue-card", "timeoutMs": 30000},
                    timeout=45.0,
                )
                last_action_at = now
                self.sleep_with_stop(2.0)
                continue
            status = f"到货已逾期筛选处理中，等待按钮变蓝并切表：{body[:80]}"
            if status != last_status:
                self.log(status)
                last_status = status
            self.sleep_with_stop(2.0)
        raise ZclawError("等待到货已逾期列表切换超时。")

    def extract_urgent_arrival_overdue_rows(self, store_id: str, recent_days: int = ARRIVAL_OVERDUE_RECENT_DAYS) -> dict[str, Any]:
        deadline = time.monotonic() + 25.0
        raw: dict[str, Any] = {}
        warned = False
        while time.monotonic() < deadline:
            self.raise_if_stop_requested()
            raw = self.execute_script(store_id, URGENT_ARRIVAL_OVERDUE_EXTRACT_SCRIPT, timeout_ms=30000) or {}
            rows = raw.get("rows") or []
            body_preview = normalize_text(str(raw.get("bodyPreview") or ""))
            if rows or raw.get("hasNoData") or "暂无数据" in body_preview:
                break
            if not warned:
                self.log("到货已逾期列表已打开，但表格行还没渲染完，继续等待...", "WARN")
                warned = True
            self.sleep_with_stop(2.0)

        parsed_rows = []
        for item in raw.get("rows") or []:
            mapped = {normalize_text(str(key)): str(value or "").strip() for key, value in (item.get("mapped") or {}).items()}
            cells = [str(cell or "").strip() for cell in item.get("cells") or []]
            prepare_order_no = pick_mapping_value(mapped, "备货单号") or (cells[1] if len(cells) > 1 else "")
            if not prepare_order_no or prepare_order_no == "-":
                continue
            created_time = pick_mapping_value(mapped, "备货单创建时间")
            if not created_time:
                continue
            product_info = pick_mapping_value(mapped, "商品信息") or (cells[2] if len(cells) > 2 else "")
            status = pick_mapping_value(mapped, "状态") or (cells[3] if len(cells) > 3 else "")
            sku_info = pick_mapping_value(mapped, "SKU信息") or (cells[4] if len(cells) > 4 else "")
            declared_price_text = pick_mapping_value(mapped, "申报价格(CNY)", "申报价格") or (cells[5] if len(cells) > 5 else "")
            delivery_info = pick_mapping_value(mapped, "送货/入库数") or (cells[7] if len(cells) > 7 else "")
            created_at = parse_datetime_text(created_time)
            parsed_rows.append(
                {
                    "prepareOrderNo": prepare_order_no,
                    "productInfo": product_info,
                    "status": status,
                    "skuInfo": sku_info,
                    "declaredPrice": round(to_float(declared_price_text, 0.0), 2),
                    "deliveryInfo": delivery_info,
                    "createdTime": created_time,
                    "isRecent": is_recent(created_at, recent_days),
                }
            )

        recent_rows = [item for item in parsed_rows if item.get("isRecent")]
        return {
            "url": raw.get("url"),
            "title": raw.get("title"),
            "bodyPreview": raw.get("bodyPreview"),
            "rows": parsed_rows,
            "recentCount": len(recent_rows),
            "recentDays": recent_days,
        }

    def navigate_urgent_pending_price_tab(self, store_id: str) -> None:
        deadline = time.monotonic() + 60.0
        last_status = ""
        last_action_at = 0.0
        while time.monotonic() < deadline:
            self.raise_if_stop_requested()
            state = self.execute_script(store_id, URGENT_PENDING_PRICE_STATE_SCRIPT, timeout_ms=30000) or {}
            body = normalize_text(str(state.get("body") or ""))
            active_tab = normalize_text(str(state.get("activeTabText") or ""))
            now = time.monotonic()
            if self.maybe_handle_temu_auth_gate(store_id, "检查待发货低申报价页", URGENT_STOCK_URL, state):
                last_status = ""
                last_action_at = 0.0
                continue
            if state.get("loginPrompt"):
                raise ZclawError("跳转到卖家中心登录页，当前登录态未同步。")
            if not state.get("isUrgentPage"):
                status = f"检查待发货低申报价页未进入紧急备货建议主页面，继续等待：{body[:80]}"
                if status != last_status:
                    self.log(status, "WARN")
                    last_status = status
                self.sleep_with_stop(2.0)
                continue
            if state.get("pendingTabSeen") and not active_tab.startswith("待发货") and now - last_action_at >= 1.0:
                self.execute_script(store_id, URGENT_CLICK_PENDING_TAB_SCRIPT, timeout_ms=30000)
                last_action_at = now
                self.sleep_with_stop(2.0)
                continue
            if active_tab.startswith("待发货") and state.get("hasUrgentTable") and state.get("hasPriceHeader"):
                if not state.get("loading") and (
                    to_int(state.get("actualRowCount")) > 0
                    or state.get("hasNoData")
                    or to_int(state.get("pendingCount")) == 0
                ):
                    return
                status = f"检查待发货低申报价页已打开，等待列表加载：{body[:80]}"
                if status != last_status:
                    self.log(status)
                    last_status = status
            self.sleep_with_stop(2.0)
        raise ZclawError("等待检查待发货低申报价页超时。")

    def extract_urgent_pending_price_rows(self, store_id: str) -> dict[str, Any]:
        deadline = time.monotonic() + 25.0
        raw: dict[str, Any] = {}
        warned = False
        missing_table_warned = False
        while time.monotonic() < deadline:
            self.raise_if_stop_requested()
            raw = self.execute_script(store_id, URGENT_PENDING_PRICE_EXTRACT_SCRIPT, timeout_ms=30000) or {}
            rows = raw.get("rows") or []
            body_preview = normalize_text(str(raw.get("bodyPreview") or ""))
            if not raw.get("isUrgentPage") and body_preview:
                raise ZclawError("检查待发货低申报价时未停留在紧急备货建议页面。")
            if not raw.get("tableFound") and not raw.get("hasNoData"):
                if not missing_table_warned:
                    headers_preview = "、".join(str(item) for item in (raw.get("headersPreview") or raw.get("headers") or [])[:12] if item)
                    suffix = f"；已看到表头：{headers_preview}" if headers_preview else ""
                    self.log(f"检查待发货低申报价主表暂未稳定，继续等待{suffix}", "WARN")
                    missing_table_warned = True
                self.sleep_with_stop(2.0)
                continue
            if rows or raw.get("hasNoData") or "暂无数据" in body_preview:
                break
            if not warned:
                self.log("检查待发货低申报价页已打开，但表格行还没渲染完，继续等待...", "WARN")
                warned = True
            self.sleep_with_stop(2.0)

        if raw and not raw.get("tableFound") and not raw.get("hasNoData"):
            headers_preview = "、".join(str(item) for item in (raw.get("headersPreview") or raw.get("headers") or [])[:12] if item)
            table_count = to_int(raw.get("tableCount"), -1)
            details = []
            if table_count >= 0:
                details.append(f"可见表格{table_count}个")
            if headers_preview:
                details.append(f"已看到表头：{headers_preview}")
            message_suffix = f"（{'；'.join(details)}）" if details else ""
            raise ZclawError(f"检查待发货低申报价主表等待后仍未找到，当前页面结构与预期不符{message_suffix}。")

        parsed_rows = []
        for item in raw.get("rows") or []:
            prepare_order_no = normalize_text(str(item.get("prepareOrderNo") or ""))
            if not re.search(r"\bWB\d{6,}\b", prepare_order_no):
                continue
            declared_price_text = normalize_text(str(item.get("declaredPriceText") or ""))
            if not looks_like_currency_price(declared_price_text):
                continue
            declared_price = to_float(declared_price_text, -1.0)
            if declared_price < 0 or declared_price >= 10:
                continue
            parsed_rows.append(
                {
                    "prepareOrderNo": prepare_order_no,
                    "productInfo": normalize_text(str(item.get("productInfo") or "")),
                    "status": normalize_text(str(item.get("status") or "")),
                    "skuInfo": normalize_text(str(item.get("skuInfo") or "")),
                    "declaredPrice": round(declared_price, 2),
                    "createdTime": normalize_text(str(item.get("createdTime") or "")),
                    "remark": "申报价格低于10",
                    "isLowPrice": True,
                }
            )

        return {
            "url": raw.get("url"),
            "title": raw.get("title"),
            "bodyPreview": raw.get("bodyPreview"),
            "rows": parsed_rows,
            "lowPriceCount": len(parsed_rows),
        }

    def navigate_govern_dashboard(self, store_id: str) -> None:
        self.visit_page_with_retry(store_id, GOVERN_DASHBOARD_URL, "合规中心首页")
        deadline = time.monotonic() + 60.0
        last_status = ""
        while time.monotonic() < deadline:
            self.raise_if_stop_requested()
            state = self.execute_script(store_id, GOVERN_METRICS_SCRIPT, timeout_ms=30000) or {}
            title = str(state.get("title") or "")
            text = normalize_text(str(state.get("body") or ""))
            ip_complaint_count = state.get("ipComplaintCount")
            tro_count = state.get("troCount")
            if self.maybe_handle_temu_auth_gate(store_id, "合规中心首页", GOVERN_DASHBOARD_URL, state):
                last_status = ""
                continue
            if state.get("loginPrompt"):
                raise ZclawError("跳转到卖家中心登录页，当前登录态未同步。")
            if "合规中心" in title or "涉嫌违反政策" in text:
                if state.get("ready") and ip_complaint_count is not None and tro_count is not None:
                    return
                status = f"合规中心已打开，等待涉嫌违反政策指标加载：{text[:80]}"
                if status != last_status:
                    self.log(status)
                    last_status = status
            self.sleep_with_stop(2.0)
        raise ZclawError("等待合规中心首页超时。")

    def extract_govern_metrics(self, store_id: str) -> dict[str, Any]:
        deadline = time.monotonic() + 20.0
        raw: dict[str, Any] = {}
        warned = False
        while time.monotonic() < deadline:
            self.raise_if_stop_requested()
            raw = self.execute_script(store_id, GOVERN_METRICS_SCRIPT, timeout_ms=30000) or {}
            ip_complaint_count = raw.get("ipComplaintCount")
            tro_count = raw.get("troCount")
            if ip_complaint_count is not None and tro_count is not None:
                return {
                    "url": raw.get("url"),
                    "title": raw.get("title"),
                    "ipComplaintCount": to_int(ip_complaint_count),
                    "troCount": to_int(tro_count),
                    "bodyPreview": raw.get("section") or raw.get("body"),
                }
            if not warned:
                self.log("合规中心已打开，但涉嫌违反政策指标还没渲染完，继续等待...", "WARN")
                warned = True
            self.sleep_with_stop(2.0)
        raise ZclawError("合规中心涉嫌违反政策指标读取失败。")

    def navigate_shipping_list(self, store_id: str) -> None:
        self.visit_page_with_retry(store_id, SHIPPING_LIST_URL, f"检查VMI超{SHIPPING_STALE_DAYS}天未收货页")
        deadline = time.monotonic() + 90.0
        last_action_at = 0.0
        query_clicked_at = 0.0
        last_status = ""
        while time.monotonic() < deadline:
            self.raise_if_stop_requested()
            state = self.execute_script(store_id, SHIPPING_FILTER_STATE_SCRIPT, timeout_ms=30000) or {}
            title = normalize_text(str(state.get("title") or ""))
            text = normalize_text(str(state.get("body") or ""))
            jit_value = normalize_text(str(state.get("jitValue") or ""))
            active_tab = normalize_text(str(state.get("activeTab") or ""))
            now = time.monotonic()
            if self.maybe_handle_temu_auth_gate(store_id, f"检查VMI超{SHIPPING_STALE_DAYS}天未收货页", SHIPPING_LIST_URL, state):
                last_status = ""
                last_action_at = 0.0
                query_clicked_at = 0.0
                continue
            if state.get("loginPrompt"):
                raise ZclawError("跳转到卖家中心登录页，当前登录态未同步。")
            if not state.get("isExpanded") and now - last_action_at >= 1.0:
                self.execute_script(store_id, SHIPPING_CLICK_EXPAND_SCRIPT, timeout_ms=30000)
                last_action_at = now
                self.sleep_with_stop(1.5)
                continue
            if state.get("isExpanded") and jit_value != "否" and now - last_action_at >= 1.0:
                self.execute_script(store_id, SHIPPING_OPEN_JIT_SELECT_SCRIPT, timeout_ms=30000)
                self.sleep_with_stop(0.8)
                self.execute_script(store_id, SHIPPING_SELECT_JIT_NO_SCRIPT, timeout_ms=30000)
                last_action_at = now
                query_clicked_at = 0.0
                self.sleep_with_stop(1.2)
                continue
            if state.get("isExpanded") and jit_value == "否" and query_clicked_at == 0.0 and now - last_action_at >= 1.0:
                self.execute_script(store_id, SHIPPING_CLICK_QUERY_SCRIPT, timeout_ms=30000)
                last_action_at = now
                query_clicked_at = now
                self.sleep_with_stop(2.0)
                continue
            if state.get("isExpanded") and jit_value == "否" and active_tab != "待仓库收货" and now - last_action_at >= 1.0:
                self.execute_script(store_id, SHIPPING_CLICK_WAITING_TAB_SCRIPT, timeout_ms=30000)
                last_action_at = now
                self.sleep_with_stop(2.0)
                continue
            if ("发货单列表" in title or "发货单列表" in text) and jit_value == "否" and active_tab == "待仓库收货":
                if query_clicked_at > 0 and not state.get("loading") and (
                    state.get("rowCount", 0) > 0 or state.get("hasNoData") or state.get("totalRowsText") is not None
                ):
                    return
                status = f"检查VMI超{SHIPPING_STALE_DAYS}天未收货页已打开，等待筛选结果加载：{text[:80]}"
                if status != last_status:
                    self.log(status)
                    last_status = status
            self.sleep_with_stop(2.0)
        raise ZclawError(f"等待检查VMI超{SHIPPING_STALE_DAYS}天未收货页筛选结果超时。")

    def extract_shipping_rows(self, store_id: str) -> dict[str, Any]:
        deadline = time.monotonic() + 25.0
        raw: dict[str, Any] = {}
        warned = False
        while time.monotonic() < deadline:
            self.raise_if_stop_requested()
            raw = self.execute_script(store_id, SHIPPING_LIST_EXTRACT_SCRIPT, timeout_ms=30000) or {}
            rows = raw.get("rows") or []
            body_preview = normalize_text(str(raw.get("bodyPreview") or ""))
            total_rows_text = raw.get("totalRowsText")
            if rows or "暂无数据" in body_preview or total_rows_text == 0:
                break
            if not warned:
                self.log(f"检查VMI超{SHIPPING_STALE_DAYS}天未收货页筛选已完成，但表格行还没渲染完，继续等待...", "WARN")
                warned = True
            self.sleep_with_stop(2.0)

        parsed_rows = []
        for item in raw.get("rows") or []:
            cells = [str(cell or "") for cell in item.get("cells") or []]
            node_info = cells[9] if len(cells) > 9 else ""
            status = cells[10] if len(cells) > 10 else ""
            product_info = cells[4] if len(cells) > 4 else ""
            logistics_info = cells[2] if len(cells) > 2 else ""
            shipping_order_info = cells[3] if len(cells) > 3 else ""
            shipping_order_no = shipping_order_info.split()[0] if shipping_order_info else ""
            prepare_order_no = extract_text_after_label(product_info, "备货单号")
            ship_time = parse_shipping_datetime(node_info, "发货时间")
            receive_time = parse_shipping_datetime(node_info, "收货时间")
            parsed_rows.append(
                {
                    "batchNo": cells[1] if len(cells) > 1 else "",
                    "shippingOrderNo": shipping_order_no,
                    "prepareOrderNo": prepare_order_no,
                    "productInfo": product_info,
                    "shippingInfo": logistics_info,
                    "nodeInfo": node_info,
                    "status": status,
                    "shipTimeText": ship_time.strftime("%Y-%m-%d %H:%M:%S") if ship_time else "",
                    "receiveTimeText": receive_time.strftime("%Y-%m-%d %H:%M:%S") if receive_time else "",
                    "isStale": is_older_than_days(ship_time, SHIPPING_STALE_DAYS),
                }
            )

        return {
            "url": raw.get("url"),
            "title": raw.get("title"),
            "totalRowsText": raw.get("totalRowsText"),
            "bodyPreview": raw.get("bodyPreview"),
            "rows": parsed_rows,
            "staleCount": sum(1 for row in parsed_rows if row["isStale"]),
        }

    def navigate_return_order(self, store_id: str) -> None:
        self.visit_page_with_retry(store_id, RETURN_ORDER_URL, "退货包裹查询页")
        deadline = time.monotonic() + 90.0
        query_clicked = False
        last_status = ""
        while time.monotonic() < deadline:
            self.raise_if_stop_requested()
            state = self.execute_script(store_id, RETURN_ORDER_STATE_SCRIPT, timeout_ms=30000) or {}
            title = normalize_text(str(state.get("title") or ""))
            text = normalize_text(str(state.get("body") or ""))
            url = str(state.get("url") or "")
            if self.maybe_handle_temu_auth_gate(store_id, "退货包裹查询页", RETURN_ORDER_URL, state):
                last_status = ""
                query_clicked = False
                continue
            if state.get("loginPrompt"):
                raise ZclawError("跳转到卖家中心登录页，当前登录态未同步。")
            if "/return-order-mgt" in url or "退货包裹" in title or "退货包裹" in text:
                if state.get("hasQueryButton") and not query_clicked:
                    self.execute_script(store_id, RETURN_ORDER_CLICK_QUERY_SCRIPT, timeout_ms=30000)
                    query_clicked = True
                    self.sleep_with_stop(2.0)
                    continue
                if not state.get("loading") and (
                    state.get("rowCount", 0) > 0 or state.get("hasNoData") or state.get("totalRowsText") == 0
                ):
                    return
                status = f"退货包裹查询页已打开，等待查询结果加载：{text[:80]}"
                if status != last_status:
                    self.log(status)
                    last_status = status
            self.sleep_with_stop(2.0)
        raise ZclawError("等待退货包裹查询页超时。")

    def extract_return_order_rows(self, store_id: str) -> dict[str, Any]:
        deadline = time.monotonic() + 25.0
        raw: dict[str, Any] = {}
        warned = False
        while time.monotonic() < deadline:
            self.raise_if_stop_requested()
            raw = self.execute_script(store_id, RETURN_ORDER_EXTRACT_SCRIPT, timeout_ms=30000) or {}
            rows = raw.get("rows") or []
            body_preview = normalize_text(str(raw.get("bodyPreview") or ""))
            total_rows_text = raw.get("totalRowsText")
            if rows or "暂无数据" in body_preview or "暂无结果" in body_preview or total_rows_text == 0:
                break
            if not warned:
                self.log("退货包裹查询页已打开，但表格行还没渲染完，继续等待...", "WARN")
                warned = True
            self.sleep_with_stop(2.0)

        parsed_rows = []
        for item in raw.get("rows") or []:
            mapped = {normalize_text(str(key)): str(value or "").strip() for key, value in (item.get("mapped") or {}).items()}
            cells = [str(cell or "").strip() for cell in item.get("cells") or []]
            tracking_no = pick_mapping_value(mapped, "运单号", "快递单号") or (cells[3] if len(cells) > 3 else "")
            return_package_no = pick_mapping_value(mapped, "退货包裹号") or (cells[2] if len(cells) > 2 else "")
            carrier = pick_mapping_value(mapped, "物流商", "快递公司") or (cells[10] if len(cells) > 10 else "")
            status = pick_mapping_value(mapped, "状态") or (cells[6] if len(cells) > 6 else "")
            pack_complete_time = pick_mapping_value(mapped, "打包完成时间") or (cells[11] if len(cells) > 11 else "")
            outbound_time = pick_mapping_value(mapped, "出库时间") or (cells[12] if len(cells) > 12 else "")
            parsed_rows.append(
                {
                    "returnPackageNo": return_package_no,
                    "trackingNo": tracking_no,
                    "carrier": carrier,
                    "status": status,
                    "packCompleteTime": pack_complete_time,
                    "outboundTime": outbound_time,
                }
            )

        return {
            "url": raw.get("url"),
            "title": raw.get("title"),
            "totalRowsText": raw.get("totalRowsText"),
            "bodyPreview": raw.get("bodyPreview"),
            "rows": parsed_rows,
            "count": len(parsed_rows),
        }

    def navigate_funds_center(self, store_id: str) -> None:
        self.visit_page_with_retry(store_id, FUNDS_CENTER_URL, "资金中心页")
        deadline = time.monotonic() + 60.0
        last_status = ""
        while time.monotonic() < deadline:
            self.raise_if_stop_requested()
            state = self.execute_script(store_id, FUNDS_BALANCE_SCRIPT, timeout_ms=30000) or {}
            title = normalize_text(str(state.get("title") or ""))
            text = normalize_text(str(state.get("body") or ""))
            url = str(state.get("url") or "")
            if self.maybe_handle_temu_auth_gate(store_id, "资金中心页", FUNDS_CENTER_URL, state):
                last_status = ""
                continue
            if state.get("loginPrompt"):
                raise ZclawError("跳转到卖家中心登录页，当前登录态未同步。")
            if "/labor/account" in url or "资金中心" in title or "可用余额" in text:
                if state.get("ready"):
                    return
                status = f"资金中心页已打开，等待余额加载：{text[:80]}"
                if status != last_status:
                    self.log(status)
                    last_status = status
            self.sleep_with_stop(2.0)
        raise ZclawError("等待资金中心页超时。")

    def extract_funds_balance(self, store_id: str) -> dict[str, Any]:
        deadline = time.monotonic() + 20.0
        raw: dict[str, Any] = {}
        warned = False
        while time.monotonic() < deadline:
            self.raise_if_stop_requested()
            raw = self.execute_script(store_id, FUNDS_BALANCE_SCRIPT, timeout_ms=30000) or {}
            available_balance = raw.get("availableBalance")
            if available_balance is not None:
                break
            if not warned:
                self.log("资金中心页已打开，但余额还没渲染完，继续等待...", "WARN")
                warned = True
            self.sleep_with_stop(2.0)
        return {
            "url": raw.get("url"),
            "title": raw.get("title"),
            "bodyPreview": raw.get("body"),
            "availableBalance": to_float(raw.get("availableBalance")),
            "balanceText": str(raw.get("balanceText") or ""),
            "needWithdraw": to_float(raw.get("availableBalance")) > WITHDRAW_ALERT_THRESHOLD,
        }

    def navigate_violation_message(self, store_id: str) -> None:
        self.visit_page_with_retry(store_id, VIOLATION_MESSAGE_URL, "违规信息页")
        deadline = time.monotonic() + 60.0
        last_status = ""
        while time.monotonic() < deadline:
            self.raise_if_stop_requested()
            state = self.execute_script(store_id, VIOLATION_MESSAGE_STATE_SCRIPT, timeout_ms=30000) or {}
            title = normalize_text(str(state.get("title") or ""))
            text = normalize_text(str(state.get("body") or ""))
            url = str(state.get("url") or "")
            if self.maybe_handle_temu_auth_gate(store_id, "违规信息页", VIOLATION_MESSAGE_URL, state):
                last_status = ""
                continue
            if state.get("loginPrompt"):
                raise ZclawError("跳转到卖家中心登录页，当前登录态未同步。")
            if "/violation-message" in url or "违规信息" in title or "违规信息" in text:
                if state.get("ready") and not state.get("loading") and (
                    state.get("rowCount", 0) > 0 or state.get("hasNoData") or state.get("totalRowsText") == 0
                ):
                    return
                status = f"违规信息页已打开，等待列表加载：{text[:80]}"
                if status != last_status:
                    self.log(status)
                    last_status = status
            self.sleep_with_stop(2.0)
        raise ZclawError("等待违规信息页超时。")

    def extract_violation_rows(self, store_id: str) -> dict[str, Any]:
        deadline = time.monotonic() + 25.0
        raw: dict[str, Any] = {}
        warned = False
        while time.monotonic() < deadline:
            self.raise_if_stop_requested()
            raw = self.execute_script(store_id, VIOLATION_MESSAGE_EXTRACT_SCRIPT, timeout_ms=30000) or {}
            rows = raw.get("rows") or []
            body_preview = normalize_text(str(raw.get("bodyPreview") or ""))
            total_rows_text = raw.get("totalRowsText")
            if rows or "暂无数据" in body_preview or "暂无结果" in body_preview or total_rows_text == 0:
                break
            if not warned:
                self.log("违规信息页已打开，但表格行还没渲染完，继续等待...", "WARN")
                warned = True
            self.sleep_with_stop(2.0)

        rows = []
        for item in raw.get("rows") or []:
            mapped = {normalize_text(str(key)): str(value or "").strip() for key, value in (item.get("mapped") or {}).items()}
            cells = [str(cell or "").strip() for cell in item.get("cells") or []]
            violation_time = pick_mapping_value(mapped, "违规发起时间") or (cells[6] if len(cells) > 6 else "")
            progress = pick_mapping_value(mapped, "进度") or (cells[10] if len(cells) > 10 else "")
            rows.append(
                {
                    "violationNo": pick_mapping_value(mapped, "违规编号") or (cells[2] if len(cells) > 2 else ""),
                    "prepareOrderNo": pick_mapping_value(mapped, "备货单") or (cells[3] if len(cells) > 3 else ""),
                    "prepareOrderType": pick_mapping_value(mapped, "备货单类型") or (cells[4] if len(cells) > 4 else ""),
                    "violationType": pick_mapping_value(mapped, "违规类型") or (cells[5] if len(cells) > 5 else ""),
                    "violationTime": violation_time,
                    "amount": pick_mapping_value(mapped, "违规金额(CNY)") or (cells[7] if len(cells) > 7 else ""),
                    "reducedAmount": pick_mapping_value(mapped, "减免后违规金额") or (cells[9] if len(cells) > 9 else ""),
                    "progress": progress,
                    "action": pick_mapping_value(mapped, "操作") or (cells[11] if len(cells) > 11 else ""),
                    "needsManual": violation_needs_manual(progress),
                }
            )

        return {
            "url": raw.get("url"),
            "title": raw.get("title"),
            "totalRowsText": raw.get("totalRowsText"),
            "bodyPreview": raw.get("bodyPreview"),
            "rows": rows,
            "pendingCount": sum(1 for row in rows if row.get("needsManual")),
        }

    def extract_qc_bad_items(self, store_id: str) -> dict[str, Any]:
        deadline = time.monotonic() + 20.0
        raw: dict[str, Any] = {}
        warned = False
        while time.monotonic() < deadline:
            self.raise_if_stop_requested()
            raw = self.execute_script(store_id, QC_EXTRACT_SCRIPT, timeout_ms=30000) or {}
            rows = raw.get("rows") or []
            body_preview = normalize_text(str(raw.get("bodyPreview") or ""))
            total_rows_text = raw.get("totalRowsText")
            if rows or "暂无数据" in body_preview or total_rows_text == 0:
                break
            if not warned:
                self.log("抽检页已打开，但表格行还没渲染完，继续等待...", "WARN")
                warned = True
            self.sleep_with_stop(2.0)
        rows = []
        for item in raw.get("rows") or []:
            row = {
                "productInfo": str(item.get("productInfo") or ""),
                "skuInfo": str(item.get("skuInfo") or ""),
                "prepareOrderNo": str(item.get("prepareOrderNo") or ""),
                "latestQcTime": str(item.get("latestQcTime") or ""),
                "operation": str(item.get("operation") or ""),
            }
            row["isRecent"] = is_recent(parse_datetime_text(row["latestQcTime"]), self.recent_days)
            rows.append(row)
        return {
            "url": raw.get("url"),
            "title": raw.get("title"),
            "totalRowsText": raw.get("totalRowsText"),
            "bodyPreview": raw.get("bodyPreview"),
            "rows": rows,
            "recentCount": sum(1 for row in rows if row["isRecent"]),
        }

    def inspect_store(self, store: dict[str, Any], selected_checks: set[str] | None = None) -> dict[str, Any]:
        if not self.client:
            raise ZclawError("ZClaw 未连接。")
        enabled = normalize_selected_checks(selected_checks)
        store_id = str(store.get("storeId") or "")
        store_name = str(store.get("storeName") or "")
        email_account = self.get_active_email_account(store_id)
        record: dict[str, Any] | None = None
        self.raise_if_stop_requested()
        self.client.invoke("open_store", {"storeId": store_id, "launchUrl": TEMU_HOME_URL}, timeout=180.0)
        try:
            self.ensure_temu_home(store_id, email_account)
            urgent: dict[str, Any] = {}
            urgent_pending_price: dict[str, Any] = {}
            govern: dict[str, Any] = {}
            shipping: dict[str, Any] = {}
            price_rule: dict[str, Any] = {}
            return_order: dict[str, Any] = {}
            funds: dict[str, Any] = {}
            violation: dict[str, Any] = {}
            qc: dict[str, Any] = {}
            project_errors: list[dict[str, Any]] = []

            if "urgent" in enabled or "urgent_declared_price" in enabled:
                self.navigate_urgent_stock(store_id)
                if "urgent" in enabled:
                    urgent = self.extract_urgent_metrics(store_id)
                    arrival_overdue = to_int(urgent.get("arrivalOverdue"))
                    if arrival_overdue > 0:
                        self.navigate_urgent_arrival_overdue(store_id, arrival_overdue)
                        arrival_result = self.extract_urgent_arrival_overdue_rows(store_id)
                        urgent["arrivalOverdueRows"] = arrival_result.get("rows") or []
                        urgent["arrivalOverdueRecentCount"] = to_int(arrival_result.get("recentCount"))
                        if to_int(urgent.get("arrivalOverdueRecentCount")) > 0:
                            self.log(
                                f"检查JIT是否逾期：{store.get('storeName')} 到货已逾期 {arrival_overdue} 条，其中近{ARRIVAL_OVERDUE_RECENT_DAYS}天创建 {urgent['arrivalOverdueRecentCount']} 条。",
                                "WARN",
                            )
                        else:
                            self.log(
                                f"检查JIT是否逾期：{store.get('storeName')} 到货已逾期 {arrival_overdue} 条，但近{ARRIVAL_OVERDUE_RECENT_DAYS}天创建 0 条，不提示人工。",
                            )
                if "urgent_declared_price" in enabled:
                    try:
                        urgent_pending_price = self.run_urgent_declared_price_with_retry(store_id, store_name)
                        low_price_count = to_int(urgent_pending_price.get("lowPriceCount"))
                        if low_price_count > 0:
                            self.log(f"检查待发货低申报价：{store_name} 发现 {low_price_count} 条申报价格低于10。", "WARN")
                        else:
                            self.log(f"检查待发货低申报价：{store_name} 未发现低于10的申报价格。")
                    except StopRequestedError:
                        raise
                    except Exception as error:
                        self.add_project_error(project_errors, "urgent_declared_price", error)
                        urgent_pending_price = {
                            "error": str(error),
                            "rows": [],
                            "lowPriceCount": 0,
                        }
                        self.log(
                            f"检查待发货低申报价：{store_name} 连续失败，已跳过该子项目并继续后续巡查：{error}",
                            "ERROR",
                        )
            if "govern" in enabled:
                self.navigate_govern_dashboard(store_id)
                govern = self.extract_govern_metrics(store_id)
            if "shipping" in enabled:
                self.navigate_shipping_list(store_id)
                shipping = self.extract_shipping_rows(store_id)
            if "return_order" in enabled:
                self.navigate_return_order(store_id)
                return_order = self.extract_return_order_rows(store_id)
            if "funds" in enabled:
                self.navigate_funds_center(store_id)
                funds = self.extract_funds_balance(store_id)
            if "violation" in enabled:
                self.navigate_violation_message(store_id)
                violation = self.extract_violation_rows(store_id)
                violation_pending_count = to_int(violation.get("pendingCount"))
                if violation_pending_count > 0:
                    self.log(f"违规信息：{store.get('storeName')} 命中待处理 {violation_pending_count} 条。", "WARN")
                else:
                    self.log(f"违规信息：{store.get('storeName')} 未命中待处理项。")
            if "qc" in enabled:
                self.navigate_qc_detail(store_id)
                qc = self.extract_qc_bad_items(store_id)
            if "price_rule" in enabled:
                self.navigate_price_rule(store_id)
                price_rule = self.run_price_rule_assistant(store_id)
                waiting_count = to_int(price_rule.get("waitingCount"))
                matched_count = to_int(price_rule.get("matchedCount"))
                acted_count = to_int(price_rule.get("actedCount"))
                remaining_count = to_int(price_rule.get("remainingCount"))
                stage = str(price_rule.get("stage") or "")
                if waiting_count <= 0:
                    self.log(f"价格申报自动助手：{store.get('storeName')} 待卖家确认 0 条。")
                elif bool(price_rule.get("confirmPending")):
                    self.log(
                        f"价格申报自动助手：{store.get('storeName')} 自动确认失败，待卖家确认 {waiting_count} 条，规则命中 {matched_count} 条，弹窗已处理 {acted_count} 条，需人工确认。",
                        "WARN",
                    )
                elif remaining_count > 0:
                    self.log(
                        f"价格申报自动助手：{store.get('storeName')} 已自动确认，待卖家确认 {waiting_count} 条中已处理 {acted_count} 条，剩余 {remaining_count} 条需人工处理。",
                        "WARN",
                    )
                else:
                    self.log(
                        f"价格申报自动助手：{store.get('storeName')} 已自动确认完成，待卖家确认 {waiting_count} 条，规则命中 {matched_count} 条，状态 {stage or 'done'}。",
                    )
            record = {
                "checkedAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "target": {
                    "storeId": store_id,
                    "storeName": str(store.get("storeName") or ""),
                    "platformName": str(store.get("platformName") or ""),
                    "ip": str(store.get("ip") or ""),
                },
                "urgent": urgent,
                "urgentPendingPrice": urgent_pending_price,
                "govern": govern,
                "shipping": shipping,
                "priceRule": price_rule,
                "returnOrder": return_order,
                "funds": funds,
                "violation": violation,
                "qc": qc,
                "projectErrors": project_errors,
            }
            return record
        except StopRequestedError:
            raise
        except Exception:
            raise
        finally:
            if self.stop_requested:
                self.log(f"任务已停止，保留店铺窗口当前页面：{store.get('storeName')}", "WARN")
            else:
                try:
                    self.dock_store_window_after_check(store_id, store_name)
                except Exception as error:
                    self.log(f"查后停靠核价页面失败：{store.get('storeName')} -> {error}", "WARN")

    def summarize_record_status(self, record: dict[str, Any], selected_checks: set[str] | None = None) -> tuple[bool, str, str]:
        enabled = normalize_selected_checks(selected_checks)
        reasons = build_manual_reasons(record, self.recent_days, enabled)
        if reasons:
            return True, f"需处理: {'；'.join(reasons)}", "#C62828"
        return False, f"正常: 已查{len(enabled)}项", "#2E7D32"

    def run_batch(self, selected_stores: list[dict[str, Any]], selected_checks: set[str] | None = None) -> None:
        total = len(selected_stores)
        enabled_checks = normalize_selected_checks(selected_checks)
        records: list[dict[str, Any]] = []
        failed: list[str] = []
        finished_count = 0
        panel_only = bool(self.panel_only_var.get())

        try:
            if not self.client:
                self.refresh_connection()
                if not self.client:
                    self.set_status("连接失败")
                    return

            self.set_status("巡店中，请稍候...")
            self.set_progress(0, total, "准备开始")
            self.latest_report = None
            self.ui_call(self.on_output_mode_changed)
            self.log(f"开始批量巡 TEMU 店铺，共 {total} 家。")
            self.log(f"本次巡查项目：{format_selected_checks(enabled_checks)}")
            self.log(f"查后停靠：{POST_CHECK_DOCK_LABEL}（核价脚本入口）")
            if self.active_email_accounts:
                self.log(f"邮箱登录账号：本轮已填写 {len(self.active_email_accounts)} 家。")
            if panel_only:
                self.log("输出方式：仅在面板显示，本次不生成报表。")
            else:
                self.log(f"输出方式：生成Excel报表，保存到 {self.report_dir}")

            for index, store in enumerate(selected_stores, start=1):
                if self.stop_requested:
                    self.log("检测到停止请求，后续店铺不再查询。", "WARN")
                    break

                store_id = str(store.get("storeId") or "")
                store_name = str(store.get("storeName") or "")
                started_at = time.monotonic()
                self.set_store_runtime_status(store_id, "查询中...", "#1565C0", needs_manual=False)
                self.update_result_row(
                    store_id,
                    store_name,
                    "巡查中",
                    "正在查询",
                    "-",
                    "-",
                    f"店铺：{store_name}\n\n状态：巡查中\n\n说明：正在执行当前店铺的巡查项目，请稍候。",
                    "running",
                )
                self.set_progress(finished_count, total, f"正在查询 {store_name}")
                self.log(f"[{index}/{total}] 开始查询：{store_name}")
                try:
                    record = self.inspect_store(store, enabled_checks)
                    records.append(record)
                    finished_count += 1
                    needs_manual, status_text, status_color = self.summarize_record_status(record, enabled_checks)
                    duration_text = f"{time.monotonic() - started_at:.1f}s"
                    summary_text = "；".join(build_manual_reasons(record, self.recent_days, enabled_checks)) or "正常"
                    conclusion_text = "需处理" if needs_manual else "正常"
                    self.set_store_runtime_status(store_id, status_text, status_color, needs_manual=needs_manual)
                    self.update_result_row(
                        store_id,
                        store_name,
                        "已完成",
                        summary_text,
                        conclusion_text,
                        duration_text,
                        self.format_record_detail(record, enabled_checks),
                        "manual" if needs_manual else "ok",
                    )
                    self.set_progress(finished_count, total, f"已完成 {store_name}")
                    self.log(f"[{index}/{total}] 完成：{store_name}")
                    if needs_manual:
                        self.log(f"[{index}/{total}] 命中需人工处理：{store_name} -> {status_text}", "WARN")
                except StopRequestedError:
                    finished_count += 1
                    duration_text = f"{time.monotonic() - started_at:.1f}s"
                    self.set_store_runtime_status(store_id, "已停止", "#B26A00", needs_manual=False)
                    self.update_result_row(
                        store_id,
                        store_name,
                        "已停止",
                        "用户停止",
                        "未完成",
                        duration_text,
                        f"店铺：{store_name}\n\n状态：已停止\n\n说明：用户手动停止了本次巡查。",
                        "failed",
                    )
                    self.set_progress(finished_count, total, f"已停止 {store_name}")
                    self.log(f"[{index}/{total}] 已停止：{store_name}", "WARN")
                    break
                except Exception as error:
                    finished_count += 1
                    duration_text = f"{time.monotonic() - started_at:.1f}s"
                    failed.append(f"{store_name}（{error}）")
                    self.set_store_runtime_status(store_id, "失败待处理", "#B26A00", needs_manual=True)
                    self.update_result_row(
                        store_id,
                        store_name,
                        "失败",
                        str(error),
                        "待人工",
                        duration_text,
                        f"店铺：{store_name}\n\n状态：失败待处理\n\n错误：{error}",
                        "failed",
                    )
                    self.set_progress(finished_count, total, f"失败 {store_name}")
                    self.log(f"[{index}/{total}] 失败：{store_name} -> {error}", "ERROR")

            if not records:
                if self.stop_requested:
                    self.log("任务已停止，本次没有成功结果，未生成报表。", "WARN")
                    self.set_status("巡店已停止")
                else:
                    self.log("本次没有成功结果，未生成报表。", "ERROR")
                    self.set_status("巡店失败")
                if failed:
                    self.log(f"失败店铺：{', '.join(failed)}", "ERROR")
                return

            if panel_only:
                self.ui_call(self.right_notebook.select, 0)
                self.latest_report = None
                self.ui_call(self.on_output_mode_changed)
                self.log("本次已按仅面板显示模式完成，未生成报表。", "SUCCESS")
                if failed:
                    self.log(f"失败店铺：{', '.join(failed)}", "ERROR")
                if self.stop_requested:
                    self.set_status("巡店已停止")
                else:
                    self.set_status("巡店完成（仅面板显示）")
                return

            report_name = f"紫鸟TEMU巡店结果_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            report_path = self.report_dir / report_name
            try:
                generate_report(records, report_path, self.recent_days, enabled_checks)
            except Exception as error:
                self.log(f"生成报表失败：{report_path} -> {error}", "ERROR")
                if failed:
                    self.log(f"失败店铺：{', '.join(failed)}", "ERROR")
                self.set_status("报表生成失败")
                return
            self.latest_report = report_path
            self.ui_call(self.on_output_mode_changed)
            self.log(f"报表已生成：{report_path}", "SUCCESS")
            if failed:
                self.log(f"失败店铺：{', '.join(failed)}", "ERROR")
            if self.stop_requested:
                self.set_status("巡店已停止")
            else:
                self.set_status("巡店完成")
        finally:
            self.batch_thread = None
            self.stop_requested = False
            self.set_running(False)
            self.set_progress(finished_count, total, "已结束")


def main() -> None:
    root = tk.Tk()
    app = ZiniaoTemuInspectorApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
